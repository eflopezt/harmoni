"""
Tests para integraciones/contables.py — generadores de asientos contables.

Cubre:
- _empresa_label() — helper para labels de empresa
- _monto_str()    — formateo de Decimal
- _get_plan_cuentas() — fallback al default
- Estructura de Concar CSV (columnas + cuadre DEBE/HABER)
- Estructura de Siscont (columnas + fila TOTAL)
- Estructura de SIRE PLE (13 campos por linea)
- Multi-empresa: con/sin empresa, sufijos en numero de asiento

Algunos tests usan SimpleNamespace para mockear los modelos.
"""
from __future__ import annotations

from decimal import Decimal
from types import SimpleNamespace
from unittest.mock import patch, MagicMock

import pytest

from integraciones.contables import (
    _empresa_label,
    _monto_str,
    _get_plan_cuentas,
    PLAN_CUENTAS_DEFAULT,
)


# ════════════════════════════════════════════════════════════════════════════
# 1. Helpers — no DB
# ════════════════════════════════════════════════════════════════════════════

class TestEmpresaLabel:
    """_empresa_label devuelve (codigo, nombre) para usar en CC y glosa."""

    def test_empresa_none(self):
        """Sin empresa devuelve ('', '')."""
        cod, nom = _empresa_label(None)
        assert cod == ""
        assert nom == ""

    def test_empresa_con_subdominio(self):
        """Si tiene subdominio, lo usa como codigo."""
        emp = SimpleNamespace(
            subdominio="demo",
            ruc="20100100100",
            razon_social="EDO Sushi Bar SAC",
            nombre_comercial="EDO Sushi Bar",
            pk=1,
        )
        cod, nom = _empresa_label(emp)
        assert cod == "demo"
        assert "EDO Sushi Bar" in nom

    def test_empresa_sin_subdominio_usa_ruc(self):
        """Sin subdominio, usa el RUC."""
        emp = SimpleNamespace(
            subdominio="",
            ruc="20100100200",
            razon_social="EDO Premium SAC",
            nombre_comercial="",
            pk=2,
        )
        cod, nom = _empresa_label(emp)
        assert cod == "20100100200"
        assert "EDO Premium" in nom

    def test_empresa_sin_subdominio_ni_ruc_usa_pk(self):
        """Sin subdominio ni ruc, usa el pk."""
        emp = SimpleNamespace(
            subdominio="",
            ruc="",
            razon_social="Sin RUC SAC",
            nombre_comercial="",
            pk=42,
        )
        cod, nom = _empresa_label(emp)
        assert cod == "42"

    def test_empresa_label_trunca_codigo_a_11_chars(self):
        """El codigo no debe exceder 11 chars (suficiente para un RUC)."""
        emp = SimpleNamespace(
            subdominio="muy-largo-subdominio-supera-once",
            ruc="",
            razon_social="X",
            nombre_comercial="",
            pk=1,
        )
        cod, _ = _empresa_label(emp)
        assert len(cod) <= 11

    def test_empresa_label_trunca_nombre_a_40_chars(self):
        """Nombre comercial maximo 40 chars."""
        emp = SimpleNamespace(
            subdominio="x",
            ruc="20",
            razon_social="",
            nombre_comercial="X" * 100,
            pk=1,
        )
        _, nom = _empresa_label(emp)
        assert len(nom) <= 40


# ════════════════════════════════════════════════════════════════════════════
# 2. _monto_str — formato 2 decimales sin separador
# ════════════════════════════════════════════════════════════════════════════

class TestMontoStr:
    def test_decimal_simple(self):
        assert _monto_str(Decimal("100.00")) == "100.00"

    def test_decimal_con_centavos(self):
        assert _monto_str(Decimal("145333.01")) == "145333.01"

    def test_redondeo_2_decimales(self):
        """5+ decimales se cortan a 2."""
        assert _monto_str(Decimal("100.12345")) == "100.12"

    def test_none_devuelve_cero(self):
        assert _monto_str(None) == "0.00"

    def test_cero_devuelve_cero(self):
        assert _monto_str(Decimal("0")) == "0.00"

    def test_negativo(self):
        assert _monto_str(Decimal("-50.5")) == "-50.50"

    def test_string_se_convierte(self):
        """Acepta string-like, lo convierte a Decimal."""
        assert _monto_str("123.45") == "123.45"


# ════════════════════════════════════════════════════════════════════════════
# 3. _get_plan_cuentas — fallback al default
# ════════════════════════════════════════════════════════════════════════════

class TestPlanCuentas:
    def test_plan_default_tiene_cuentas_obligatorias(self):
        plan = _get_plan_cuentas()
        # Cuentas obligatorias para una planilla peruana
        assert "sueldos_debe" in plan
        assert "essalud_debe" in plan
        assert "gratif_debe" in plan
        assert "rem_pagar_haber" in plan
        assert "afp_pagar_haber" in plan
        assert "onp_pagar_haber" in plan
        assert "essalud_pagar_haber" in plan

    def test_plan_default_codigos_pcge_correctos(self):
        plan = _get_plan_cuentas()
        # Estos codigos son del PCGE 2020 Peru — no inventados
        assert plan["sueldos_debe"][0] == "6211"   # Sueldos y salarios
        assert plan["gratif_debe"][0] == "6215"    # Gratificaciones provision
        assert plan["essalud_debe"][0] == "6271"   # EsSalud aporte empleador
        assert plan["rem_pagar_haber"][0] == "4111"  # Remuneraciones por pagar
        assert plan["afp_pagar_haber"][0] == "4031"  # AFP por pagar
        assert plan["onp_pagar_haber"][0] == "4011"  # ONP por pagar
        assert plan["essalud_pagar_haber"][0] == "4032"  # EsSalud por pagar

    def test_default_es_inmutable(self):
        """PLAN_CUENTAS_DEFAULT no debe modificarse al llamar _get_plan_cuentas."""
        plan = _get_plan_cuentas()
        before = dict(PLAN_CUENTAS_DEFAULT)
        plan["sueldos_debe"] = ("9999", "fake")
        # PLAN_CUENTAS_DEFAULT debe seguir intacto
        assert PLAN_CUENTAS_DEFAULT == before


# ════════════════════════════════════════════════════════════════════════════
# 4. Asiento Concar — estructura CSV
# ════════════════════════════════════════════════════════════════════════════

def _mock_periodo(anio=2026, mes=5, total_bruto=100000, total_neto=80000):
    """Construye un mock de PeriodoNomina sin BD."""
    from datetime import date
    return SimpleNamespace(
        pk=1,
        anio=anio,
        mes=mes,
        mes_nombre={
            1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
            5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
            9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
        }.get(mes, "Mes"),
        fecha_fin=date(anio, mes, 30 if mes != 2 else 28),
        total_bruto=Decimal(str(total_bruto)),
        total_neto=Decimal(str(total_neto)),
    )


class TestConcarCSV:
    """Validacion de estructura del CSV CONCAR sin BD."""

    @patch("integraciones.contables._get_totales_periodo")
    def test_concar_tiene_headers_correctos(self, mock_totales):
        """La primera fila debe tener los 20 headers de CONCAR."""
        mock_totales.return_value = {
            "bruto": Decimal("100000"),
            "neto": Decimal("80000"),
            "essalud": Decimal("9000"),
            "afp": Decimal("11000"),
            "onp": Decimal("0"),
            "gratif_prov": Decimal("16667"),
        }
        from integraciones.contables import generar_asiento_concar
        contenido, count = generar_asiento_concar(_mock_periodo())
        primera_linea = contenido.split("\n")[0]
        assert "SubDiario" in primera_linea
        assert "NroAsiento" in primera_linea
        assert "FechaAsiento" in primera_linea
        assert "CodCuenta" in primera_linea
        assert "CodCentroCosto" in primera_linea
        assert "TipoMov" in primera_linea
        assert "MontoMN" in primera_linea
        assert "CodMoneda" in primera_linea
        # Debe tener exactamente 20 columnas (separadas por coma)
        assert primera_linea.count(",") == 19

    @patch("integraciones.contables._get_totales_periodo")
    def test_concar_genera_filas_debe_haber(self, mock_totales):
        """Debe tener filas D (debe) y H (haber)."""
        mock_totales.return_value = {
            "bruto": Decimal("100000"),
            "neto": Decimal("80000"),
            "essalud": Decimal("9000"),
            "afp": Decimal("11000"),
            "onp": Decimal("0"),
            "gratif_prov": Decimal("16667"),
        }
        from integraciones.contables import generar_asiento_concar
        contenido, _ = generar_asiento_concar(_mock_periodo())
        lineas = [l for l in contenido.split("\n") if l]
        # Header + al menos 6 lineas (3 D + 3 H minimo)
        assert len(lineas) >= 7
        # Hay filas con tipo D
        assert any(",D," in l for l in lineas[1:])
        # Hay filas con tipo H
        assert any(",H," in l for l in lineas[1:])

    @patch("integraciones.contables._get_totales_periodo")
    def test_concar_con_empresa_agrega_sufijo(self, mock_totales):
        """Si pasamos empresa, el NroAsiento incluye sufijo."""
        mock_totales.return_value = {
            "bruto": Decimal("100000"),
            "neto": Decimal("80000"),
            "essalud": Decimal("9000"),
            "afp": Decimal("11000"),
            "onp": Decimal("0"),
            "gratif_prov": Decimal("16667"),
        }
        from integraciones.contables import generar_asiento_concar
        emp = SimpleNamespace(
            subdominio="demo",
            ruc="20100100100",
            razon_social="EDO Sushi Bar",
            nombre_comercial="EDO Sushi Bar",
            pk=1,
        )
        contenido, _ = generar_asiento_concar(_mock_periodo(), empresa=emp)
        # Numero de asiento debe contener "-demo"
        assert "-demo" in contenido
        # Glosa debe mencionar la empresa
        assert "EDO Sushi Bar" in contenido

    @patch("integraciones.contables._get_totales_periodo")
    def test_concar_cuadre_debe_haber(self, mock_totales):
        """Suma de D = suma de H en el asiento generado (validacion contable)."""
        mock_totales.return_value = {
            "bruto":       Decimal("100000.00"),
            "neto":        Decimal("80000.00"),
            "essalud":     Decimal("9000.00"),
            "afp":         Decimal("11000.00"),
            "onp":         Decimal("0"),
            "gratif_prov": Decimal("16667.00"),
        }
        from integraciones.contables import generar_asiento_concar
        contenido, _ = generar_asiento_concar(_mock_periodo())
        debe = Decimal("0")
        haber = Decimal("0")
        for l in contenido.split("\n")[1:]:
            if not l.strip():
                continue
            cols = l.split(",")
            if len(cols) < 14:
                continue
            tipo_mov = cols[10]
            try:
                monto = Decimal(cols[13])
            except Exception:
                continue
            if tipo_mov == "D":
                debe += monto
            elif tipo_mov == "H":
                haber += monto
        # Cuadre dentro de tolerancia (centavos)
        assert abs(debe - haber) < Decimal("0.05"), f"DEBE={debe} != HABER={haber}"


# ════════════════════════════════════════════════════════════════════════════
# 5. Asiento Siscont — estructura Excel
# ════════════════════════════════════════════════════════════════════════════

class TestSiscontExcel:
    @patch("integraciones.contables._get_totales_periodo")
    def test_siscont_genera_bytes_xlsx(self, mock_totales):
        """Devuelve bytes de un archivo Excel válido."""
        mock_totales.return_value = {
            "bruto": Decimal("100000"),
            "neto": Decimal("80000"),
            "essalud": Decimal("9000"),
            "afp": Decimal("11000"),
            "onp": Decimal("0"),
            "gratif_prov": Decimal("16667"),
        }
        from integraciones.contables import generar_asiento_siscont
        contenido, count = generar_asiento_siscont(_mock_periodo())
        assert isinstance(contenido, bytes)
        # Excel xlsx starts with PK (zip format)
        assert contenido.startswith(b"PK")
        assert count == 1

    @patch("integraciones.contables._get_totales_periodo")
    def test_siscont_con_empresa_label_en_archivo(self, mock_totales):
        """El glosa con nombre empresa va dentro del Excel."""
        mock_totales.return_value = {
            "bruto": Decimal("100000"), "neto": Decimal("80000"),
            "essalud": Decimal("9000"), "afp": Decimal("11000"),
            "onp": Decimal("0"), "gratif_prov": Decimal("16667"),
        }
        from integraciones.contables import generar_asiento_siscont
        import openpyxl
        from io import BytesIO

        emp = SimpleNamespace(
            subdominio="demo", ruc="20100100100",
            razon_social="EDO Sushi Bar SAC",
            nombre_comercial="EDO Sushi Bar", pk=1,
        )
        contenido, _ = generar_asiento_siscont(_mock_periodo(), empresa=emp)
        wb = openpyxl.load_workbook(BytesIO(contenido))
        ws = wb.active
        # Headers correctos
        headers = [ws.cell(row=1, column=i).value for i in range(1, 11)]
        assert "Fecha" in headers
        assert "Cuenta" in headers
        assert "Glosa" in headers
        assert "Debe (S/.)" in headers
        assert "Haber (S/.)" in headers
        # Glosa con nombre empresa
        glosa_any = False
        for row in ws.iter_rows(min_row=2, max_row=10, values_only=True):
            for cell in row:
                if cell and "EDO Sushi Bar" in str(cell):
                    glosa_any = True
                    break
        assert glosa_any, "No se encontro 'EDO Sushi Bar' en la glosa"


# ════════════════════════════════════════════════════════════════════════════
# 6. Asiento SIRE PLE — formato SUNAT 13 campos
# ════════════════════════════════════════════════════════════════════════════

class TestSirePLE:
    @patch("integraciones.contables._get_totales_periodo")
    def test_sire_lineas_tienen_13_campos(self, mock_totales):
        """Cada linea del PLE Libro Diario debe tener 13 campos separados por |."""
        mock_totales.return_value = {
            "bruto": Decimal("100000"), "neto": Decimal("80000"),
            "essalud": Decimal("9000"), "afp": Decimal("11000"),
            "onp": Decimal("0"), "gratif_prov": Decimal("16667"),
        }
        from integraciones.contables import generar_sire_libro_diario
        contenido, _ = generar_sire_libro_diario(_mock_periodo())
        for linea in contenido.strip().split("\n"):
            if not linea.strip():
                continue
            campos = linea.split("|")
            assert len(campos) == 13, f"Linea con {len(campos)} campos: {linea}"

    @patch("integraciones.contables._get_totales_periodo")
    def test_sire_periodo_formato_correcto(self, mock_totales):
        """Campo 1 debe ser AAAAMM00 (formato SUNAT)."""
        mock_totales.return_value = {
            "bruto": Decimal("100000"), "neto": Decimal("80000"),
            "essalud": Decimal("9000"), "afp": Decimal("11000"),
            "onp": Decimal("0"), "gratif_prov": Decimal("16667"),
        }
        from integraciones.contables import generar_sire_libro_diario
        contenido, _ = generar_sire_libro_diario(_mock_periodo(anio=2026, mes=5))
        primera_linea = contenido.strip().split("\n")[0]
        periodo_campo = primera_linea.split("|")[0]
        assert periodo_campo == "20260500"


# ════════════════════════════════════════════════════════════════════════════
# 7. Asiento SIGO — formato pipe-delimitado
# ════════════════════════════════════════════════════════════════════════════

class TestProvisionesExcel:
    """Asiento de provisiones contables separadas."""

    @patch("integraciones.contables._get_totales_periodo")
    def test_provisiones_genera_xlsx(self, mock_totales):
        """Debe generar bytes de un Excel valido."""
        mock_totales.return_value = {
            "bruto": Decimal("60000"), "neto": Decimal("48000"),
            "essalud": Decimal("5400"), "afp": Decimal("6600"),
            "onp": Decimal("0"), "gratif_prov": Decimal("10000"),
        }
        from integraciones.contables import generar_asiento_provisiones
        contenido, count = generar_asiento_provisiones(_mock_periodo(total_bruto=60000))
        assert isinstance(contenido, bytes)
        assert contenido.startswith(b"PK")  # zip header de xlsx
        assert count == 4  # 4 asientos generados

    @patch("integraciones.contables._get_totales_periodo")
    def test_provisiones_tiene_5_hojas(self, mock_totales):
        """Debe tener Resumen + 4 hojas de asiento."""
        mock_totales.return_value = {
            "bruto": Decimal("60000"), "neto": Decimal("48000"),
            "essalud": Decimal("5400"), "afp": Decimal("6600"),
            "onp": Decimal("0"), "gratif_prov": Decimal("10000"),
        }
        from integraciones.contables import generar_asiento_provisiones
        import openpyxl
        from io import BytesIO

        contenido, _ = generar_asiento_provisiones(_mock_periodo(total_bruto=60000))
        wb = openpyxl.load_workbook(BytesIO(contenido))
        nombres_hojas = set(wb.sheetnames)
        assert "Resumen" in nombres_hojas
        assert "Asiento Gratificación" in nombres_hojas
        assert "Asiento CTS" in nombres_hojas
        assert "Asiento Vacaciones" in nombres_hojas
        # Bonif 9% solo aparece si prov_bonif9 > 0 (siempre debe aparecer en este test)
        assert "Asiento Bonif 9%" in nombres_hojas

    @patch("integraciones.contables._get_totales_periodo")
    def test_provisiones_calculos_correctos(self, mock_totales):
        """Verifica que los cálculos de provisión cumplen la normativa peruana."""
        bruto = Decimal("60000.00")
        mock_totales.return_value = {
            "bruto": bruto, "neto": Decimal("48000"),
            "essalud": Decimal("5400"), "afp": Decimal("6600"),
            "onp": Decimal("0"), "gratif_prov": Decimal("10000"),
        }
        from integraciones.contables import generar_asiento_provisiones
        import openpyxl
        from io import BytesIO

        contenido, _ = generar_asiento_provisiones(_mock_periodo(total_bruto=60000))
        wb = openpyxl.load_workbook(BytesIO(contenido))

        # Hoja Asiento Gratificación: 1/6 del bruto = 10,000
        ws = wb["Asiento Gratificación"]
        debe_grat = ws.cell(row=2, column=4).value  # Fila DEBE, columna Debe
        assert abs(debe_grat - 10000.00) < 0.05

        # Hoja Asiento Vacaciones: 1/12 del bruto = 5,000
        ws = wb["Asiento Vacaciones"]
        debe_vac = ws.cell(row=2, column=4).value
        assert abs(debe_vac - 5000.00) < 0.05

        # Hoja Asiento CTS: 1/12 + 1/72 = 0.0833 + 0.01389 = 0.0972 del bruto = 5833.33
        ws = wb["Asiento CTS"]
        debe_cts = ws.cell(row=2, column=4).value
        # 60000 * (1/12 + 1/72) = 60000 * 0.0972 = 5833.33
        assert abs(debe_cts - 5833.33) < 0.10

    @patch("integraciones.contables._get_totales_periodo")
    def test_provisiones_cuentas_pcge_correctas(self, mock_totales):
        """Las cuentas usadas deben ser PCGE 2020."""
        mock_totales.return_value = {
            "bruto": Decimal("60000"), "neto": Decimal("48000"),
            "essalud": Decimal("5400"), "afp": Decimal("6600"),
            "onp": Decimal("0"), "gratif_prov": Decimal("10000"),
        }
        from integraciones.contables import generar_asiento_provisiones
        import openpyxl
        from io import BytesIO

        contenido, _ = generar_asiento_provisiones(_mock_periodo(total_bruto=60000))
        wb = openpyxl.load_workbook(BytesIO(contenido))

        # Gratif: DEBE 6215 → HABER 4151
        ws = wb["Asiento Gratificación"]
        assert ws.cell(row=2, column=1).value == "6215"  # DEBE
        assert ws.cell(row=3, column=1).value == "4151"  # HABER

        # CTS: DEBE 6291 → HABER 4152
        ws = wb["Asiento CTS"]
        assert ws.cell(row=2, column=1).value == "6291"
        assert ws.cell(row=3, column=1).value == "4152"

        # Vacaciones: DEBE 6214 → HABER 4115
        ws = wb["Asiento Vacaciones"]
        assert ws.cell(row=2, column=1).value == "6214"
        assert ws.cell(row=3, column=1).value == "4115"


class TestSigoTxt:
    @patch("integraciones.contables._get_totales_periodo")
    def test_sigo_lineas_15_campos(self, mock_totales):
        """SIGO usa 15 campos pipe-delimitados sin header."""
        mock_totales.return_value = {
            "bruto": Decimal("100000"), "neto": Decimal("80000"),
            "essalud": Decimal("9000"), "afp": Decimal("11000"),
            "onp": Decimal("0"), "gratif_prov": Decimal("16667"),
        }
        from integraciones.contables import generar_asiento_sigo
        contenido, _ = generar_asiento_sigo(_mock_periodo())
        for linea in contenido.strip().split("\n"):
            if not linea.strip():
                continue
            campos = linea.split("|")
            assert len(campos) == 15

    @patch("integraciones.contables._get_totales_periodo")
    def test_sigo_con_empresa_usa_cc(self, mock_totales):
        """El CentroCosto (campo 10) debe ser el codigo de la empresa."""
        mock_totales.return_value = {
            "bruto": Decimal("100000"), "neto": Decimal("80000"),
            "essalud": Decimal("9000"), "afp": Decimal("11000"),
            "onp": Decimal("0"), "gratif_prov": Decimal("16667"),
        }
        from integraciones.contables import generar_asiento_sigo
        emp = SimpleNamespace(
            subdominio="demo", ruc="X", razon_social="EDO",
            nombre_comercial="EDO", pk=1,
        )
        contenido, _ = generar_asiento_sigo(_mock_periodo(), empresa=emp)
        primera_linea = contenido.strip().split("\n")[0]
        # campo 10 (índice 9) es CentroCosto
        cc = primera_linea.split("|")[9]
        assert cc == "demo"
