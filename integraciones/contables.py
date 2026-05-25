"""
Integraciones Contables - Generadores de asientos de planilla.

Formatos soportados:
  - Concar  (CSV - el mas usado en Peru para medianas empresas)
  - SIGO    (TXT pipe-delimitado - empresas publicas y mineras)
  - SAP/Generico (Excel - compatible con SAP FICO y otros ERPs contables)
  - SIRE SUNAT (PLE - Libros Electronicos Libro Diario, Registro 5.1)

Plan de cuentas PCGE 2020 (simplificado - configurable por empresa):
  6211  Sueldos y salarios
  6215  Gratificaciones (provision)
  6271  EsSalud (empleador)
  4111  Remuneraciones por pagar
  4031  AFP por pagar
  4011  ONP (SCN) por pagar
  4032  EsSalud por pagar
"""
import csv
import io
from datetime import date
from decimal import Decimal

# Plan de cuentas default (PCGE 2020)

PLAN_CUENTAS_DEFAULT = {
    "sueldos_debe":        ("6211", "Sueldos y salarios"),
    "gratif_debe":         ("6215", "Gratificaciones (provision)"),
    "essalud_debe":        ("6271", "EsSalud - aporte empleador"),
    "rem_pagar_haber":     ("4111", "Remuneraciones por pagar"),
    "afp_pagar_haber":     ("4031", "AFP por pagar"),
    "onp_pagar_haber":     ("4011", "ONP por pagar"),
    "essalud_pagar_haber": ("4032", "EsSalud por pagar"),
    # Provisiones contables separadas (PCGE 2020)
    "cts_debe":            ("6291", "Compensacion tiempo de servicios (provision)"),
    "cts_pagar_haber":     ("4152", "CTS por pagar"),
    "vac_debe":            ("6214", "Vacaciones (provision)"),
    "vac_pagar_haber":     ("4115", "Vacaciones por pagar"),
    "gratif_pagar_haber":  ("4151", "Gratificaciones por pagar"),
    "bonif_9_debe":        ("6294", "Bonificacion extraordinaria Ley 29351"),
    "bonif_9_pagar_haber": ("4118", "Bonificacion extraordinaria por pagar"),
}


def _get_plan_cuentas(empresa=None):
    """Obtiene el plan de cuentas con prioridad:
       1. Empresa.plan_cuentas (JSON por RUC, si existe)
       2. ConfiguracionSistema.plan_cuentas (global, si existe)
       3. PLAN_CUENTAS_DEFAULT (constante PCGE 2020).

    Cada nivel mergea sobre el anterior, así puedes sobreescribir solo
    las cuentas que difieran del default.

    Siempre devuelve una COPIA del dict (no la referencia).
    """
    import json
    plan = dict(PLAN_CUENTAS_DEFAULT)

    # Nivel 2: configuración global
    try:
        from core.models import ConfiguracionSistema
        cfg = ConfiguracionSistema.objects.first()
        if cfg and hasattr(cfg, "plan_cuentas") and cfg.plan_cuentas:
            try:
                global_pc = json.loads(cfg.plan_cuentas)
                # Convertir listas [codigo, nombre] a tuplas para mantener el formato
                global_pc = {k: tuple(v) if isinstance(v, list) else v for k, v in global_pc.items()}
                plan.update(global_pc)
            except (json.JSONDecodeError, TypeError, ValueError):
                pass
    except Exception:
        pass

    # Nivel 1: empresa específica (sobre-escribe global)
    if empresa is not None and hasattr(empresa, 'plan_cuentas') and empresa.plan_cuentas:
        try:
            emp_pc = json.loads(empresa.plan_cuentas) if isinstance(empresa.plan_cuentas, str) else empresa.plan_cuentas
            emp_pc = {k: tuple(v) if isinstance(v, list) else v for k, v in emp_pc.items()}
            plan.update(emp_pc)
        except (json.JSONDecodeError, TypeError, ValueError, AttributeError):
            pass

    return plan


def _get_totales_periodo(periodo, empresa=None):
    """
    Calcula los totales del periodo desagregados por tipo de descuento/ingreso.
    Retorna dict: bruto, neto, afp, onp, essalud, gratif_prov.

    Args:
        periodo: PeriodoNomina instance
        empresa: opcional, Empresa instance o id. Si se pasa, filtra solo
                 los RegistroNomina cuyo personal pertenece a esa empresa.
                 Util para generar asientos contables desagregados en
                 escenarios multi-RUC (ej: Sabores del Sur con 24 razones sociales).

    Si empresa es None, agrega TODOS los registros del periodo (consolidado,
    comportamiento legacy).

    Prioriza los totales precalculados en PeriodoNomina.total_bruto/total_neto
    solo si empresa es None (los precalculados son consolidados).
    """
    from django.db.models import Sum
    from nominas.models import RegistroNomina, LineaNomina

    # Filtro base
    reg_qs   = RegistroNomina.objects.filter(periodo=periodo)
    linea_qs = LineaNomina.objects.filter(registro__periodo=periodo)
    if empresa is not None:
        emp_id = empresa.pk if hasattr(empresa, "pk") else empresa
        reg_qs   = reg_qs.filter(personal__empresa_id=emp_id)
        linea_qs = linea_qs.filter(registro__personal__empresa_id=emp_id)

    # Solo usar precalculados si NO hay filtro de empresa
    bruto_pre = (periodo.total_bruto or Decimal("0")) if empresa is None else Decimal("0")
    neto_pre  = (periodo.total_neto  or Decimal("0")) if empresa is None else Decimal("0")

    if bruto_pre > 0:
        agg = reg_qs.aggregate(essalud=Sum("aporte_essalud"))
        essalud = (agg["essalud"] or Decimal("0")).quantize(Decimal("0.01"))

        afp_total = linea_qs.filter(
            concepto__formula__in=["AFP_APORTE", "AFP_COMISION", "AFP_SEGURO"],
        ).aggregate(s=Sum("monto"))["s"] or Decimal("0")

        onp_total = linea_qs.filter(
            concepto__formula="ONP",
        ).aggregate(s=Sum("monto"))["s"] or Decimal("0")

        gratif_prov = (bruto_pre / Decimal("6")).quantize(Decimal("0.01"))

        return {
            "bruto":       bruto_pre.quantize(Decimal("0.01")),
            "neto":        neto_pre.quantize(Decimal("0.01")),
            "essalud":     essalud,
            "afp":         afp_total.quantize(Decimal("0.01")),
            "onp":         onp_total.quantize(Decimal("0.01")),
            "gratif_prov": gratif_prov,
        }

    # Agregar desde registros individuales (con filtro empresa si aplica)
    agg = reg_qs.aggregate(
        bruto=Sum("total_ingresos"),
        neto=Sum("neto_a_pagar"),
        essalud=Sum("aporte_essalud"),
    )

    bruto   = agg["bruto"]   or Decimal("0")
    neto    = agg["neto"]    or Decimal("0")
    essalud = agg["essalud"] or Decimal("0")

    afp_total = linea_qs.filter(
        concepto__formula__in=["AFP_APORTE", "AFP_COMISION", "AFP_SEGURO"],
    ).aggregate(s=Sum("monto"))["s"] or Decimal("0")

    onp_total = linea_qs.filter(
        concepto__formula="ONP",
    ).aggregate(s=Sum("monto"))["s"] or Decimal("0")

    gratif_prov = (bruto / Decimal("6")).quantize(Decimal("0.01"))

    return {
        "bruto":       bruto.quantize(Decimal("0.01")),
        "neto":        neto.quantize(Decimal("0.01")),
        "essalud":     essalud.quantize(Decimal("0.01")),
        "afp":         afp_total.quantize(Decimal("0.01")),
        "onp":         onp_total.quantize(Decimal("0.01")),
        "gratif_prov": gratif_prov,
    }


def _empresa_label(empresa):
    """Devuelve un label corto para usar en glosa/centro de costo."""
    if empresa is None:
        return ("", "")
    cod = (getattr(empresa, 'subdominio', '') or getattr(empresa, 'ruc', '') or str(empresa.pk))[:11]
    nom = (getattr(empresa, 'nombre_comercial', '') or getattr(empresa, 'razon_social', '') or '')[:40]
    return (cod, nom)


def _monto_str(value):
    """Formatea Decimal con 2 decimales sin separador de miles."""
    if not value:
        return "0.00"
    return f"{Decimal(str(value)):.2f}"


# ====================================================================
# CONCAR - Formato CSV
# ====================================================================

def generar_asiento_concar(periodo, empresa=None):
    """
    Genera el asiento contable de planilla en formato CONCAR (CSV).

    Args:
        periodo: PeriodoNomina instance
        empresa: opcional, Empresa instance. Si se pasa, el asiento se
                 desagrega solo con los trabajadores de esa empresa,
                 usando su codigo como CentroCosto y glosa con el nombre.

    Columnas Concar (SubDiario 01 = Libro Diario, TipoMov D=Debe H=Haber):
    SubDiario | NroAsiento | FechaAsiento | CodCuenta | CodCentroCosto |
    TipoDocumento | NroDocumento | FecRef1 | CodDocRef | NroDocRef |
    TipoMov | Glosa | TipoNota | MontoMN | MontoME | TipoCambio |
    Pendiente | Marcador | FecVcto | CodMoneda
    """
    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow([
        "SubDiario", "NroAsiento", "FechaAsiento", "CodCuenta", "CodCentroCosto",
        "TipoDocumento", "NroDocumento", "FecRef1", "CodDocRef", "NroDocRef",
        "TipoMov", "Glosa", "TipoNota", "MontoMN", "MontoME", "TipoCambio",
        "Pendiente", "Marcador", "FecVcto", "CodMoneda",
    ])

    plan        = _get_plan_cuentas(empresa=empresa)
    tots        = _get_totales_periodo(periodo, empresa=empresa)
    fecha       = (periodo.fecha_fin or date.today()).strftime("%d/%m/%Y")
    cc, emp_nom = _empresa_label(empresa)
    # Asiento numero: sufijo de empresa si aplica para evitar colisiones
    suf         = f"-{cc}" if cc else ""
    nro_asiento = f"P{periodo.anio}{periodo.mes:02d}001{suf}"
    glosa       = (
        f"Planilla {periodo.mes_nombre} {periodo.anio}"
        + (f" - {emp_nom}" if emp_nom else "")
    )

    def row(cuenta_key, tipo_mov, monto):
        cod, _nombre = plan.get(cuenta_key, ("0000", ""))
        return [
            "01", nro_asiento, fecha, cod, cc,
            "PL", nro_asiento, fecha, "", "",
            tipo_mov, glosa, "", _monto_str(monto), "0.00", "1.000",
            "N", "", "", "MN",
        ]

    # DEBE
    writer.writerow(row("sueldos_debe", "D", tots["bruto"]))
    writer.writerow(row("essalud_debe", "D", tots["essalud"]))
    writer.writerow(row("gratif_debe",  "D", tots["gratif_prov"]))

    # HABER
    writer.writerow(row("rem_pagar_haber",     "H", tots["neto"]))
    if tots["afp"] > 0:
        writer.writerow(row("afp_pagar_haber", "H", tots["afp"]))
    if tots["onp"] > 0:
        writer.writerow(row("onp_pagar_haber", "H", tots["onp"]))
    writer.writerow(row("essalud_pagar_haber", "H", tots["essalud"]))

    # Cuadre: diferencia de gratif_prov -> cuenta 4151 Gratificaciones por pagar
    diferencia = (
        tots["bruto"] + tots["essalud"] + tots["gratif_prov"]
        - tots["neto"] - tots["afp"] - tots["onp"] - tots["essalud"]
    )
    if abs(diferencia) > Decimal("0.02"):
        writer.writerow([
            "01", nro_asiento, fecha, "4151", cc,
            "PL", nro_asiento, fecha, "", "",
            "H", f"{glosa} - Provision gratificaciones", "",
            _monto_str(diferencia), "0.00", "1.000",
            "N", "", "", "MN",
        ])

    return output.getvalue(), 1


# ====================================================================
# SIGO - Formato TXT pipe-delimitado
# ====================================================================

def generar_asiento_sigo(periodo, empresa=None):
    """
    Genera el asiento contable en formato SIGO (Sistema Integrado de Gestion).
    Usado por empresas publicas y sector minero. Sin encabezado, pipe-delimitado.

    Args:
        periodo: PeriodoNomina instance
        empresa: opcional, Empresa instance para desagregar por RUC.

    Campos por linea:
    TipoReg|CodEmpresa|Periodo|NumAsiento|Fecha|CodCuenta|Glosa|
    Debe|Haber|CentroCosto|TipoDoc|NroDoc|CodMoneda|TipoCambio|Estado
    """
    output      = io.StringIO()
    plan        = _get_plan_cuentas(empresa=empresa)
    tots        = _get_totales_periodo(periodo, empresa=empresa)
    fecha       = (periodo.fecha_fin or date.today()).strftime("%d%m%Y")
    periodo_str = f"{periodo.anio}{periodo.mes:02d}"
    cc, emp_nom = _empresa_label(empresa)
    suf         = f"{cc[:3]}" if cc else "001"
    num_asiento = f"PL{periodo_str}{suf}"
    glosa_base  = f"PLANILLA {periodo.mes_nombre.upper()} {periodo.anio}"
    glosa       = (glosa_base + (f" - {emp_nom.upper()}" if emp_nom else "")).strip()
    cc_field    = cc or "001"

    def linea(cuenta_key, debe, haber):
        cod, _ = plan.get(cuenta_key, ("0000", ""))
        return [
            "6", "01", periodo_str, num_asiento, fecha, cod, glosa,
            _monto_str(debe), _monto_str(haber),
            cc_field, "PL", num_asiento, "MN", "1.000", "C",
        ]

    filas = [
        linea("sueldos_debe",        tots["bruto"],       Decimal("0")),
        linea("essalud_debe",        tots["essalud"],     Decimal("0")),
        linea("gratif_debe",         tots["gratif_prov"], Decimal("0")),
        linea("rem_pagar_haber",     Decimal("0"),        tots["neto"]),
        linea("essalud_pagar_haber", Decimal("0"),        tots["essalud"]),
    ]
    if tots["afp"] > 0:
        filas.append(linea("afp_pagar_haber", Decimal("0"), tots["afp"]))
    if tots["onp"] > 0:
        filas.append(linea("onp_pagar_haber", Decimal("0"), tots["onp"]))

    for f in filas:
        output.write("|".join(f) + "\n")

    return output.getvalue(), 1


# ====================================================================
# SAP / GENERICO - Excel (openpyxl)
# ====================================================================

def generar_asiento_sap_excel(periodo, empresa=None):
    """
    Genera el asiento contable en formato Excel generico compatible
    con SAP FICO, Oracle y otros ERPs contables.

    Args:
        periodo: PeriodoNomina
        empresa: opcional, filtro por empresa.

    Hoja 1: Asiento resumen por cuenta contable.
    Hoja 2: Detalle por trabajador con columnas de nomina.
    Retorna bytes (Excel .xlsx).
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    plan        = _get_plan_cuentas(empresa=empresa)
    tots        = _get_totales_periodo(periodo, empresa=empresa)
    fecha       = periodo.fecha_fin or date.today()
    cc, emp_nom = _empresa_label(empresa)

    wb = openpyxl.Workbook()

    # Hoja 1: Asiento resumen
    ws1       = wb.active
    ws1.title = "Asiento Planilla"

    TEAL     = "FF0D2B27"
    HDR_FONT = Font(color="FFFFFFFF", bold=True, size=10)
    HDR_FILL = PatternFill(fill_type="solid", fgColor=TEAL)

    headers = [
        "Cuenta", "Descripcion", "Centro Costo", "Debe (S/)", "Haber (S/)",
        "Glosa", "Fecha", "Tipo Doc", "Referencia",
    ]
    for col, h in enumerate(headers, 1):
        cell           = ws1.cell(row=1, column=col, value=h)
        cell.font      = HDR_FONT
        cell.fill      = HDR_FILL
        cell.alignment = Alignment(horizontal="center")

    glosa = (
        f"Planilla {periodo.mes_nombre} {periodo.anio}"
        + (f" - {emp_nom}" if emp_nom else "")
    )
    ref   = f"PL{periodo.anio}{periodo.mes:02d}" + (f"-{cc}" if cc else "")

    filas_asiento = [
        (plan["sueldos_debe"][0],        plan["sueldos_debe"][1],        tots["bruto"],        Decimal("0")),
        (plan["essalud_debe"][0],        plan["essalud_debe"][1],        tots["essalud"],      Decimal("0")),
        (plan["gratif_debe"][0],         plan["gratif_debe"][1],         tots["gratif_prov"],  Decimal("0")),
        (plan["rem_pagar_haber"][0],     plan["rem_pagar_haber"][1],     Decimal("0"),         tots["neto"]),
        (plan["afp_pagar_haber"][0],     plan["afp_pagar_haber"][1],     Decimal("0"),         tots["afp"]),
        (plan["onp_pagar_haber"][0],     plan["onp_pagar_haber"][1],     Decimal("0"),         tots["onp"]),
        (plan["essalud_pagar_haber"][0], plan["essalud_pagar_haber"][1], Decimal("0"),         tots["essalud"]),
    ]

    cc_field = cc or "001"
    for i, (cuenta, desc, debe, haber) in enumerate(filas_asiento, 2):
        ws1.cell(row=i, column=1, value=cuenta)
        ws1.cell(row=i, column=2, value=desc)
        ws1.cell(row=i, column=3, value=cc_field)
        ws1.cell(row=i, column=4, value=float(debe))
        ws1.cell(row=i, column=5, value=float(haber))
        ws1.cell(row=i, column=6, value=glosa)
        ws1.cell(row=i, column=7, value=fecha)
        ws1.cell(row=i, column=8, value="PL")
        ws1.cell(row=i, column=9, value=ref)

    last = len(filas_asiento) + 2
    ws1.cell(row=last, column=3, value="TOTAL").font = Font(bold=True)
    ws1.cell(row=last, column=4, value=f"=SUM(D2:D{last-1})").font = Font(bold=True)
    ws1.cell(row=last, column=5, value=f"=SUM(E2:E{last-1})").font = Font(bold=True)

    ws1.column_dimensions["A"].width = 10
    ws1.column_dimensions["B"].width = 35
    ws1.column_dimensions["D"].width = 14
    ws1.column_dimensions["E"].width = 14
    ws1.column_dimensions["F"].width = 30
    ws1.column_dimensions["G"].width = 14

    # Hoja 2: Detalle por trabajador
    ws2 = wb.create_sheet("Detalle por Trabajador")
    headers2 = [
        "DNI", "Apellidos y Nombres", "Cargo", "Area", "Grupo",
        "Sueldo Base", "Total Ingresos", "AFP/ONP", "EsSalud",
        "Total Descuentos", "Neto a Pagar",
    ]
    for col, h in enumerate(headers2, 1):
        cell      = ws2.cell(row=1, column=col, value=h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL

    from nominas.models import RegistroNomina
    registros = RegistroNomina.objects.filter(
        periodo=periodo
    ).select_related("personal__subarea__area", "personal__empresa").order_by(
        "personal__apellidos_nombres"
    )
    if empresa is not None:
        emp_id = empresa.pk if hasattr(empresa, "pk") else empresa
        registros = registros.filter(personal__empresa_id=emp_id)

    for i, reg in enumerate(registros, 2):
        p = reg.personal
        # AFP/ONP estimado: total descuentos menos prestamos y descuentos manuales
        afp_onp = (
            float(reg.total_descuentos or 0)
            - float(reg.descuento_prestamo or 0)
            - float(reg.otros_descuentos  or 0)
        )
        ws2.cell(row=i, column=1,  value=p.nro_doc)
        ws2.cell(row=i, column=2,  value=p.apellidos_nombres)
        ws2.cell(row=i, column=3,  value=p.cargo or "")
        ws2.cell(row=i, column=4,  value=p.subarea.area.nombre if p.subarea and p.subarea.area else "")
        ws2.cell(row=i, column=5,  value=reg.grupo or getattr(p, "grupo_tareo", ""))
        ws2.cell(row=i, column=6,  value=float(reg.sueldo_base or 0))
        ws2.cell(row=i, column=7,  value=float(reg.total_ingresos or 0))
        ws2.cell(row=i, column=8,  value=max(afp_onp, 0))
        ws2.cell(row=i, column=9,  value=float(reg.aporte_essalud or 0))
        ws2.cell(row=i, column=10, value=float(reg.total_descuentos or 0))
        ws2.cell(row=i, column=11, value=float(reg.neto_a_pagar or 0))

    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 40
    ws2.column_dimensions["C"].width = 25
    ws2.column_dimensions["D"].width = 25

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue(), registros.count()


# ====================================================================
# SIRE SUNAT - PLE Libro Diario (Registro 5.1)
# ====================================================================

def generar_sire_libro_diario(periodo, empresa=None):
    """
    Genera el archivo PLE del Libro Diario en formato SUNAT SIRE.

    Args:
        periodo: PeriodoNomina
        empresa: opcional, Empresa para desagregar por RUC. Util porque
                 cada RUC debe presentar su propio PLE a SUNAT.

    Formato: TXT pipe-delimitado, Registro 5.1 del PLE.
    Campos (13 por linea):
    01 Periodo (AAAAMM00)
    02 CUO - Codigo Unico de Operacion
    03 Correlativo del asiento (3 digitos)
    04 Fecha de la operacion (DD/MM/AAAA)
    05 Glosa
    06 Referencia / numero de documento
    07 Codigo libro (05=Libro Diario)
    08 Cuenta contable PCGE
    09 Descripcion de la cuenta
    10 Tipo de moneda (1=PEN)
    11 Monto Debe en soles
    12 Monto Haber en soles
    13 Indicador estado (1=activo, 2=anulado, 9=cierre)
    """
    output      = io.StringIO()
    plan        = _get_plan_cuentas(empresa=empresa)
    tots        = _get_totales_periodo(periodo, empresa=empresa)
    fecha       = (periodo.fecha_fin or date.today()).strftime("%d/%m/%Y")
    periodo_ple = f"{periodo.anio}{periodo.mes:02d}00"
    cc, emp_nom = _empresa_label(empresa)
    suf         = f"{cc[:3]}" if cc else "001"
    cuo         = f"PL{periodo.anio}{periodo.mes:02d}{suf}"
    glosa       = (
        f"Planilla {periodo.mes_nombre} {periodo.anio}"
        + (f" - {emp_nom}" if emp_nom else "")
    )

    def linea(cuenta_key, debe, haber, correlativo):
        cod, nombre = plan.get(cuenta_key, ("0000", ""))
        return [
            periodo_ple,
            cuo,
            str(correlativo).zfill(3),
            fecha,
            glosa,
            cuo,
            "05",
            cod,
            nombre,
            "1",
            _monto_str(debe),
            _monto_str(haber),
            "1",
        ]

    asientos = [
        ("sueldos_debe",        tots["bruto"],       Decimal("0")),
        ("essalud_debe",        tots["essalud"],     Decimal("0")),
        ("gratif_debe",         tots["gratif_prov"], Decimal("0")),
        ("rem_pagar_haber",     Decimal("0"),        tots["neto"]),
        ("essalud_pagar_haber", Decimal("0"),        tots["essalud"]),
    ]
    if tots["afp"] > 0:
        asientos.append(("afp_pagar_haber", Decimal("0"), tots["afp"]))
    if tots["onp"] > 0:
        asientos.append(("onp_pagar_haber", Decimal("0"), tots["onp"]))

    for i, (cuenta_key, debe, haber) in enumerate(asientos, 1):
        output.write("|".join(linea(cuenta_key, debe, haber, i)) + "\n")

    return output.getvalue(), 1


# ====================================================================
# SISCONT Smart - Formato Excel para importar a Siscont
# ====================================================================

def generar_asiento_siscont(periodo, empresa=None):
    """
    Genera asiento contable de planilla en formato Excel compatible con
    SISCONT Smart (popular en pymes peruanas tras Concar).

    Estructura Siscont Smart importa con columnas:
    Fecha | TipoComp | NroComp | Cuenta | Glosa | Debe S/. | Haber S/. |
    Centro Costo | Anexo | Referencia
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    plan        = _get_plan_cuentas(empresa=empresa)
    tots        = _get_totales_periodo(periodo, empresa=empresa)
    fecha       = periodo.fecha_fin or date.today()
    cc, emp_nom = _empresa_label(empresa)
    suf         = f"-{cc[:3]}" if cc else ""
    nro_comp    = f"PL{periodo.anio}{periodo.mes:02d}001{suf}"
    glosa       = (
        f"Planilla {periodo.mes_nombre} {periodo.anio}"
        + (f" - {emp_nom}" if emp_nom else "")
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Asiento Planilla"

    HDR_FONT = Font(color="FFFFFFFF", bold=True, size=10)
    HDR_FILL = PatternFill(fill_type="solid", fgColor="FF0D2B27")
    headers = [
        "Fecha", "TipoComp", "NroComp", "Cuenta", "Glosa",
        "Debe (S/.)", "Haber (S/.)", "Centro Costo", "Anexo", "Referencia",
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center")

    cc_field = cc or "001"

    def row_data(cuenta_key, debe, haber):
        cod, _ = plan.get(cuenta_key, ("0000", ""))
        return [
            fecha, "PL", nro_comp, cod, glosa,
            float(debe), float(haber), cc_field, "", nro_comp,
        ]

    filas = [
        row_data("sueldos_debe",        tots["bruto"],       Decimal("0")),
        row_data("essalud_debe",        tots["essalud"],     Decimal("0")),
        row_data("gratif_debe",         tots["gratif_prov"], Decimal("0")),
        row_data("rem_pagar_haber",     Decimal("0"),        tots["neto"]),
        row_data("essalud_pagar_haber", Decimal("0"),        tots["essalud"]),
    ]
    if tots["afp"] > 0:
        filas.append(row_data("afp_pagar_haber", Decimal("0"), tots["afp"]))
    if tots["onp"] > 0:
        filas.append(row_data("onp_pagar_haber", Decimal("0"), tots["onp"]))

    # Provision gratificacion (cuadre)
    diff = (
        tots["bruto"] + tots["essalud"] + tots["gratif_prov"]
        - tots["neto"] - tots["afp"] - tots["onp"] - tots["essalud"]
    )
    if abs(diff) > Decimal("0.02"):
        filas.append([
            fecha, "PL", nro_comp, "4151",
            f"{glosa} - Provision gratificaciones",
            0.0, float(diff), cc_field, "", nro_comp,
        ])

    for i, fila in enumerate(filas, 2):
        for col, val in enumerate(fila, 1):
            ws.cell(row=i, column=col, value=val)

    # Totales debajo
    last = len(filas) + 2
    ws.cell(row=last, column=5, value="TOTAL").font = Font(bold=True)
    ws.cell(row=last, column=6, value=f"=SUM(F2:F{last-1})").font = Font(bold=True)
    ws.cell(row=last, column=7, value=f"=SUM(G2:G{last-1})").font = Font(bold=True)

    # Anchos
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["E"].width = 40
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue(), 1


# ====================================================================
# PROVISIONES CONTABLES SEPARADAS — Excel
# ====================================================================

def generar_asiento_provisiones(periodo, empresa=None):
    """
    Genera UN asiento separado por TIPO de provisión contable:
      - Gratificación  (DEBE 6215 → HABER 4151)
      - CTS            (DEBE 6291 → HABER 4152)
      - Vacaciones     (DEBE 6214 → HABER 4115)
      - Bonif. 9% Ley 29351 (DEBE 6294 → HABER 4118) si aplica al mes

    Devuelve Excel xlsx con una hoja por provisión + hoja resumen.

    Por qué separado: muchos contadores prefieren tener cada provisión
    como un asiento contable distinto para facilitar la conciliación
    con la cuenta de pasivo correspondiente al momento del pago real.

    El asiento principal (planilla regular en CONCAR/Siscont/etc) sigue
    incluyendo la provisión de gratificación como hasta ahora — este
    formato es ADICIONAL, no reemplaza al consolidado.

    Provisiones calculadas según normativa peruana:
      - Gratificación: 1/6 del bruto del mes
      - CTS:           1/12 del bruto + 1/72 del bruto (1/6 de gratif provisional)
      - Vacaciones:    1/12 del bruto
      - Bonif 9%:      9% del bruto si es mes que va con gratificación
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    plan        = _get_plan_cuentas(empresa=empresa)
    tots        = _get_totales_periodo(periodo, empresa=empresa)
    fecha       = periodo.fecha_fin or date.today()
    cc, emp_nom = _empresa_label(empresa)
    suf         = f"-{cc[:3]}" if cc else ""
    nro_base    = f"PRV{periodo.anio}{periodo.mes:02d}"
    bruto       = tots["bruto"]
    glosa_emp   = f" — {emp_nom}" if emp_nom else ""

    # Cálculos de cada provisión (normativa peruana)
    prov_gratif = (bruto / Decimal("6")).quantize(Decimal("0.01"))
    prov_vac    = (bruto / Decimal("12")).quantize(Decimal("0.01"))
    # CTS: 1/12 sueldo + 1/72 (1/6 de 1/12 = provisión gratif del mes)
    prov_cts    = ((bruto / Decimal("12")) + (bruto / Decimal("72"))).quantize(Decimal("0.01"))
    # Bonif 9% solo en meses que devengan gratif (semestres jul/dic). Tomamos
    # el cálculo conservador: 9% de la provisión gratif mensual.
    prov_bonif9 = (prov_gratif * Decimal("0.09")).quantize(Decimal("0.01"))

    wb = openpyxl.Workbook()

    TEAL     = "FF0D2B27"
    HDR_FONT = Font(color="FFFFFFFF", bold=True, size=10)
    HDR_FILL = PatternFill(fill_type="solid", fgColor=TEAL)
    cc_field = cc or "001"

    def _hoja_asiento(nombre_hoja, glosa, nro_voucher, cuenta_debe, monto_debe, cuenta_haber, monto_haber):
        """Crea una hoja con un asiento simple DEBE → HABER."""
        ws = wb.create_sheet(nombre_hoja)
        headers = ["Cuenta", "Descripcion", "Centro Costo", "Debe (S/)", "Haber (S/)",
                   "Glosa", "Fecha", "Tipo Doc", "Voucher"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = HDR_FONT; c.fill = HDR_FILL
            c.alignment = Alignment(horizontal="center")
        # Fila DEBE
        cod_d, nom_d = plan.get(cuenta_debe, ("0000", ""))
        ws.append([cod_d, nom_d, cc_field, float(monto_debe), 0.0, glosa, fecha, "PL", nro_voucher])
        # Fila HABER
        cod_h, nom_h = plan.get(cuenta_haber, ("0000", ""))
        ws.append([cod_h, nom_h, cc_field, 0.0, float(monto_haber), glosa, fecha, "PL", nro_voucher])
        # Total
        ws.append(["", "TOTAL", "", f"=SUM(D2:D3)", f"=SUM(E2:E3)", "", "", "", ""])
        # Anchos
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 38
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["F"].width = 40

    # Resumen (hoja 0)
    ws_res = wb.active
    ws_res.title = "Resumen"
    ws_res.append(["Provisiones Contables del Periodo"])
    ws_res["A1"].font = Font(bold=True, size=14, color="FF0D2B27")
    ws_res.append([])
    ws_res.append([f"Periodo: {periodo.mes_nombre} {periodo.anio}"])
    if emp_nom:
        ws_res.append([f"Empresa: {emp_nom}"])
    ws_res.append([f"Base de cálculo (bruto mensual): S/ {bruto:,.2f}"])
    ws_res.append([])
    ws_res.append(["Provision", "Cuenta DEBE", "Cuenta HABER", "Monto S/"])
    for cell in ws_res[ws_res.max_row]:
        cell.font = HDR_FONT; cell.fill = HDR_FILL
    ws_res.append(["Gratificación (1/6)",     plan["gratif_debe"][0],  plan["gratif_pagar_haber"][0], float(prov_gratif)])
    ws_res.append(["CTS (1/12 + 1/72)",       plan["cts_debe"][0],     plan["cts_pagar_haber"][0],    float(prov_cts)])
    ws_res.append(["Vacaciones (1/12)",       plan["vac_debe"][0],     plan["vac_pagar_haber"][0],    float(prov_vac)])
    ws_res.append(["Bonif 9% Ley 29351",      plan["bonif_9_debe"][0], plan["bonif_9_pagar_haber"][0], float(prov_bonif9)])
    ws_res.append([])
    total_row = ws_res.max_row + 1
    ws_res.cell(row=total_row, column=1, value="TOTAL PROVISIONES").font = Font(bold=True)
    ws_res.cell(row=total_row, column=4, value=float(prov_gratif + prov_cts + prov_vac + prov_bonif9)).font = Font(bold=True)
    ws_res.column_dimensions["A"].width = 32
    ws_res.column_dimensions["B"].width = 14
    ws_res.column_dimensions["C"].width = 14
    ws_res.column_dimensions["D"].width = 16

    # 4 hojas de asientos
    _hoja_asiento(
        "Asiento Gratificación",
        f"Provision gratificacion {periodo.mes_nombre} {periodo.anio}{glosa_emp}",
        f"{nro_base}-GRAT{suf}",
        "gratif_debe", prov_gratif, "gratif_pagar_haber", prov_gratif,
    )
    _hoja_asiento(
        "Asiento CTS",
        f"Provision CTS {periodo.mes_nombre} {periodo.anio}{glosa_emp}",
        f"{nro_base}-CTS{suf}",
        "cts_debe", prov_cts, "cts_pagar_haber", prov_cts,
    )
    _hoja_asiento(
        "Asiento Vacaciones",
        f"Provision vacaciones {periodo.mes_nombre} {periodo.anio}{glosa_emp}",
        f"{nro_base}-VAC{suf}",
        "vac_debe", prov_vac, "vac_pagar_haber", prov_vac,
    )
    if prov_bonif9 > 0:
        _hoja_asiento(
            "Asiento Bonif 9%",
            f"Provision bonif 9% Ley 29351 {periodo.mes_nombre} {periodo.anio}{glosa_emp}",
            f"{nro_base}-B9{suf}",
            "bonif_9_debe", prov_bonif9, "bonif_9_pagar_haber", prov_bonif9,
        )

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue(), 4


# ====================================================================
# EXCEL UNIVERSAL - Formato simple compatible con cualquier ERP
# ====================================================================

def generar_asiento_excel_universal(periodo, empresa=None):
    """
    Genera el asiento contable en formato Excel UNIVERSAL — UNA hoja,
    columnas mínimas y estándar que cualquier sistema contable acepta.

    Pensado para clientes que NO usan Concar/Siscont/SAP específicamente
    sino su propio ERP custom o un sistema legacy (ej. Spring de Sabores del Sur,
    sistemas hechos a medida, hojas de cálculo del contador).

    Columnas (ningún campo opcional, todo simple):
      Fecha | Nro Asiento | Cuenta | Descripción Cuenta | Centro Costo |
      Glosa | Debe | Haber

    Una sola hoja, sin formatos fancy, sin fórmulas, sin merge cells.
    Importable directamente a 99% de los ERPs del mercado.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    plan        = _get_plan_cuentas(empresa=empresa)
    tots        = _get_totales_periodo(periodo, empresa=empresa)
    fecha       = (periodo.fecha_fin or date.today())
    cc, emp_nom = _empresa_label(empresa)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Asiento Planilla"

    TEAL     = "FF0D2B27"
    HDR_FONT = Font(color="FFFFFFFF", bold=True, size=10)
    HDR_FILL = PatternFill(fill_type="solid", fgColor=TEAL)

    headers = [
        "Fecha", "Nro Asiento", "Cuenta", "Descripción Cuenta",
        "Centro Costo", "Glosa", "Debe", "Haber",
    ]
    for col, h in enumerate(headers, 1):
        c           = ws.cell(row=1, column=col, value=h)
        c.font      = HDR_FONT
        c.fill      = HDR_FILL
        c.alignment = Alignment(horizontal="center")

    suf         = f"-{cc}" if cc else ""
    nro_asiento = f"PL{periodo.anio}{periodo.mes:02d}{suf}"
    glosa       = (
        f"Planilla {periodo.mes_nombre} {periodo.anio}"
        + (f" - {emp_nom}" if emp_nom else "")
    )
    cc_field    = cc or ""

    filas = [
        # (cuenta_key, debe, haber)
        ("sueldos_debe",        tots["bruto"],       Decimal("0")),
        ("essalud_debe",        tots["essalud"],     Decimal("0")),
        ("gratif_debe",         tots["gratif_prov"], Decimal("0")),
        ("rem_pagar_haber",     Decimal("0"),        tots["neto"]),
        ("afp_pagar_haber",     Decimal("0"),        tots["afp"]),
        ("onp_pagar_haber",     Decimal("0"),        tots["onp"]),
        ("essalud_pagar_haber", Decimal("0"),        tots["essalud"]),
    ]

    # Cuadre de la provisión gratificación al haber (4151)
    diferencia = (
        tots["bruto"] + tots["essalud"] + tots["gratif_prov"]
        - tots["neto"] - tots["afp"] - tots["onp"] - tots["essalud"]
    )
    if abs(diferencia) > Decimal("0.02"):
        filas.append(("gratif_pagar_haber", Decimal("0"), diferencia))

    row = 2
    for cuenta_key, debe, haber in filas:
        # Filtrar líneas en cero (más limpio)
        if debe == 0 and haber == 0:
            continue
        cod, nom = plan.get(cuenta_key, ("0000", "—"))
        ws.cell(row=row, column=1, value=fecha)
        ws.cell(row=row, column=2, value=nro_asiento)
        ws.cell(row=row, column=3, value=cod)
        ws.cell(row=row, column=4, value=nom)
        ws.cell(row=row, column=5, value=cc_field)
        ws.cell(row=row, column=6, value=glosa)
        ws.cell(row=row, column=7, value=float(debe) if debe else None)
        ws.cell(row=row, column=8, value=float(haber) if haber else None)
        row += 1

    # Fila de TOTAL (sin fórmulas, valores estáticos — más portable)
    total_debe  = sum(d for _, d, _ in filas)
    total_haber = sum(h for _, _, h in filas)
    if diferencia > Decimal("0.02"):
        total_haber += diferencia
    ws.cell(row=row, column=6, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=7, value=float(total_debe)).font  = Font(bold=True)
    ws.cell(row=row, column=8, value=float(total_haber)).font = Font(bold=True)

    # Anchos de columna
    widths = [12, 18, 10, 35, 14, 38, 13, 13]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    # Formato fecha y decimal
    for r in range(2, row + 1):
        ws.cell(row=r, column=1).number_format = "DD/MM/YYYY"
        ws.cell(row=r, column=7).number_format = "#,##0.00"
        ws.cell(row=r, column=8).number_format = "#,##0.00"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue(), 1
