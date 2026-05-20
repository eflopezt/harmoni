"""
Generador de Flujo de Caja detallado — Consorcio SRT
Combina Planilla + BBSS, cruza con Partidas, divide por Casa/GG/Intervencion
"""
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# ========================================================
# 1. LEER DATOS
# ========================================================
print("Leyendo archivos...")

# Partidas de control
partidas = pd.read_excel(
    r'C:\Users\EDWIN LOPEZ\Downloads\Partidas de control Staff.xlsx', header=1
)
partidas.columns = ['Codigo', 'Nombre', 'DNI', 'Ocupacion', 'Partida', 'Clasificacion']
partidas['Codigo'] = partidas['Codigo'].astype(str).str.strip()
partidas['Clasificacion'] = partidas['Clasificacion'].fillna('Sin Clasificar')
partida_map = dict(zip(partidas['Codigo'], partidas['Clasificacion']))
partida_cuenta_map = dict(zip(partidas['Codigo'], partidas['Partida']))
print(f"  Partidas: {len(partidas)} empleados")
print(f"  Clasificaciones: {partidas['Clasificacion'].value_counts().to_dict()}")

# Planilla Combinada (ya tiene los 3 meses)
planilla = pd.read_excel(
    r'C:\Users\EDWIN LOPEZ\OneDrive - RIPCON\Escritorio\Planilla enero marzo empleados.xlsx'
)
planilla.columns = [c.strip() for c in planilla.columns]
# Normalizar codigo
cod_col = [c for c in planilla.columns if 'digo' in c.lower()][0]
planilla['Codigo'] = planilla[cod_col].astype(str).str.strip()
# Extraer periodo numerico
planilla['PeriodoNum'] = planilla['Periodo'].str.extract(r'(\d{2})$')[0]
planilla['PeriodoNum'] = '2026' + planilla['PeriodoNum']
print(f"  Planilla combinada: {len(planilla)} filas, periodos: {planilla['PeriodoNum'].unique()}")

# BBSS - combinar los 3 meses
bbss_files = [
    (r'C:\Users\EDWIN LOPEZ\OneDrive - RIPCON\Escritorio\Planillas\bbss enero_eflopez.XLSX', '202601'),
    (r'C:\Users\EDWIN LOPEZ\OneDrive - RIPCON\Escritorio\Planillas\bbss febrero_eflopez.XLSX', '202602'),
    (r'C:\Users\EDWIN LOPEZ\OneDrive - RIPCON\Escritorio\Planillas\bbss marzo_eflopez.XLSX', '202603'),
]

bbss_dfs = []
for path, periodo in bbss_files:
    df = pd.read_excel(path)
    df.columns = [c.strip() for c in df.columns]
    cod_col_bb = [c for c in df.columns if 'digo' in c.lower()][0]
    df['Codigo'] = df[cod_col_bb].astype(str).str.strip()
    df['PeriodoNum'] = periodo
    bbss_dfs.append(df)

bbss = pd.concat(bbss_dfs, ignore_index=True, sort=False)
print(f"  BBSS combinado: {len(bbss)} filas")

# ========================================================
# 2. ASIGNAR CLASIFICACION
# ========================================================
print("\nAsignando clasificaciones...")
planilla['Clasificacion'] = planilla['Codigo'].map(partida_map).fillna('Sin Clasificar')
planilla['PartidaCuenta'] = planilla['Codigo'].map(partida_cuenta_map).fillna('')
bbss['Clasificacion'] = bbss['Codigo'].map(partida_map).fillna('Sin Clasificar')
bbss['PartidaCuenta'] = bbss['Codigo'].map(partida_cuenta_map).fillna('')

print(f"  Planilla clasificada: {planilla['Clasificacion'].value_counts().to_dict()}")
print(f"  BBSS clasificada: {bbss['Clasificacion'].value_counts().to_dict()}")

# ========================================================
# 3. IDENTIFICAR COLUMNAS MONETARIAS
# ========================================================
# Planilla: columnas con (S/) son montos, columnas con (Hora) o (Dia) son cantidades
planilla_meta_cols = ['ArchivoOrigen', 'Proyecto', 'Periodo', 'Proceso', cod_col,
                      'N\u00ba Documento Identidad', 'Nombre', 'Ocupaci\u00f3n',
                      'Fecha Ingreso', 'Fecha Cese', 'Cuspp', 'Regimen Pensi\u00f3n',
                      'Codigo', 'PeriodoNum', 'Clasificacion', 'PartidaCuenta']
planilla_data_cols = [c for c in planilla.columns if c not in planilla_meta_cols]

bbss_meta_cols = ['Proyecto', 'Periodo', 'Proceso', 'Nombre',
                  'Tipo Doc. Identidad', 'N\u00ba Documento Identidad', 'Nacionalidad',
                  'Codigo', 'PeriodoNum', 'Clasificacion', 'PartidaCuenta']
bbss_meta_cols = [c for c in bbss_meta_cols if c in bbss.columns]
bbss_data_cols = [c for c in bbss.columns if c not in bbss_meta_cols and c != bbss.columns[3]]

print(f"\n  Planilla conceptos: {len(planilla_data_cols)}")
print(f"  BBSS conceptos: {len(bbss_data_cols)}")

# ========================================================
# 4. CREAR EXCEL
# ========================================================
print("\nGenerando Excel...")

wb = Workbook()

BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
FILL_HEADER = PatternFill('solid', fgColor='0F766E')
FILL_GG = PatternFill('solid', fgColor='DDEBF7')
FILL_CASA = PatternFill('solid', fgColor='E2EFDA')
FILL_INTERV = PatternFill('solid', fgColor='FCE4D6')
FILL_TOTAL = PatternFill('solid', fgColor='1F3864')
FNT_HDR = Font(bold=True, color='FFFFFF', name='Arial', size=9)
FNT_BOLD = Font(bold=True, name='Arial', size=9)
FNT_NORMAL = Font(name='Arial', size=9)
FNT_TOTAL = Font(bold=True, color='FFFFFF', name='Arial', size=9)
NUM_FMT = '#,##0.00'

def write_sheet(ws, df, data_cols, sheet_type='PLANILLA'):
    """Write a detailed sheet with all workers, periods, and classification."""
    # Prepare output columns
    id_cols = ['PeriodoNum', 'Codigo', 'Nombre', 'Ocupacion_display', 'Clasificacion', 'PartidaCuenta']

    # Prepare data
    if 'Ocupaci\u00f3n' in df.columns:
        df['Ocupacion_display'] = df['Ocupaci\u00f3n']
    elif 'Ocupacion' in df.columns:
        df['Ocupacion_display'] = df['Ocupacion']
    else:
        df['Ocupacion_display'] = ''

    # Sort
    df_sorted = df.sort_values(['Clasificacion', 'PeriodoNum', 'Nombre']).reset_index(drop=True)

    # Headers
    headers = ['PERIODO', 'CODIGO', 'NOMBRE', 'CARGO', 'CLASIFICACION', 'PARTIDA CUENTA'] + list(data_cols)
    widths = [10, 11, 35, 30, 16, 40] + [15] * len(data_cols)

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    c = ws.cell(1, 1, f'{sheet_type} REAL DETALLADO — CONSORCIO SRT')
    c.font = Font(bold=True, color='FFFFFF', name='Arial', size=12)
    c.fill = PatternFill('solid', fgColor='0A2F2B')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # Header row
    for j, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(2, j, h)
        cell.font = FNT_HDR
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[2].height = 32

    # Data rows
    row = 3
    fills_map = {'Casa': FILL_CASA, 'Gastos Generales': FILL_GG, 'Intervencion': FILL_INTERV}

    for _, r in df_sorted.iterrows():
        clasif = r.get('Clasificacion', '')
        fill = fills_map.get(clasif, None)

        vals = [
            r.get('PeriodoNum', ''),
            r.get('Codigo', ''),
            r.get('Nombre', ''),
            r.get('Ocupacion_display', ''),
            clasif,
            r.get('PartidaCuenta', ''),
        ]
        for dc in data_cols:
            v = r.get(dc, 0)
            vals.append(v if pd.notna(v) else 0)

        for j, v in enumerate(vals, 1):
            cell = ws.cell(row, j, v)
            cell.font = FNT_NORMAL
            cell.border = BORDER
            if j <= 6:
                cell.alignment = Alignment(horizontal='left' if j in [3,4,6] else 'center')
            else:
                cell.number_format = NUM_FMT
                cell.alignment = Alignment(horizontal='right')
            if fill and j <= 5:
                cell.fill = fill
        row += 1

    # Freeze panes and filters
    ws.freeze_panes = 'G3'
    ws.auto_filter.ref = f'A2:{get_column_letter(len(headers))}{row - 1}'

    return row


# --- Hoja 1: PLANILLA REAL ---
ws1 = wb.active
ws1.title = 'PLANILLA REAL'
ws1.sheet_properties.tabColor = '0F766E'
end_row = write_sheet(ws1, planilla, planilla_data_cols, 'PLANILLA')
print(f"  PLANILLA REAL: {end_row - 3} filas")

# --- Hoja 2: BBSS REAL ---
ws2 = wb.create_sheet('BBSS REAL')
ws2.sheet_properties.tabColor = '2E75B6'
end_row2 = write_sheet(ws2, bbss, bbss_data_cols, 'BBSS / LIQUIDACIONES')
print(f"  BBSS REAL: {end_row2 - 3} filas")

# --- Hoja 3: RESUMEN POR CLASIFICACION Y PERIODO ---
ws3 = wb.create_sheet('RESUMEN')
ws3.sheet_properties.tabColor = '1F3864'

ws3.merge_cells('A1:H1')
c = ws3.cell(1, 1, 'RESUMEN FLUJO DE CAJA POR CLASIFICACION Y PERIODO')
c.font = Font(bold=True, color='FFFFFF', name='Arial', size=12)
c.fill = PatternFill('solid', fgColor='0A2F2B')
c.alignment = Alignment(horizontal='center')
ws3.row_dimensions[1].height = 28

# Planilla resumen
resumen_cols = ['TOTAL REMUNERACION (BM) (S/)', 'TOTAL DESCUENTO (BM) (S/)', 'TOTAL NETO (BM) (S/)',
                'ESSALUD (S/)', 'TOTAL APORTES (BM) (S/)']
resumen_cols_exist = [c for c in resumen_cols if c in planilla.columns]

if resumen_cols_exist:
    pivot = planilla.groupby(['PeriodoNum', 'Clasificacion'])[resumen_cols_exist].sum().reset_index()
    pivot_sorted = pivot.sort_values(['PeriodoNum', 'Clasificacion'])

    headers3 = ['PERIODO', 'CLASIFICACION', 'PERSONAS'] + resumen_cols_exist
    for j, h in enumerate(headers3, 1):
        cell = ws3.cell(3, j, h)
        cell.font = FNT_HDR
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = BORDER
        ws3.column_dimensions[get_column_letter(j)].width = 20 if j > 2 else 14

    row = 4
    for _, r in pivot_sorted.iterrows():
        per = r['PeriodoNum']
        clasif = r['Clasificacion']
        count = len(planilla[(planilla['PeriodoNum'] == per) & (planilla['Clasificacion'] == clasif)])
        fill = {'Casa': FILL_CASA, 'Gastos Generales': FILL_GG, 'Intervencion': FILL_INTERV}.get(clasif, None)

        vals = [per, clasif, count] + [r[c] for c in resumen_cols_exist]
        for j, v in enumerate(vals, 1):
            cell = ws3.cell(row, j, v)
            cell.font = FNT_NORMAL
            cell.border = BORDER
            if j > 2:
                cell.number_format = NUM_FMT
            if fill:
                cell.fill = fill
        row += 1

    # Totals per period
    for per in sorted(planilla['PeriodoNum'].unique()):
        sub = planilla[planilla['PeriodoNum'] == per]
        vals = [per, 'TOTAL', len(sub)] + [sub[c].sum() for c in resumen_cols_exist]
        for j, v in enumerate(vals, 1):
            cell = ws3.cell(row, j, v)
            cell.font = FNT_TOTAL
            cell.fill = FILL_TOTAL
            cell.border = BORDER
            if j > 2:
                cell.number_format = NUM_FMT
        row += 1

# BBSS resumen
row += 2
ws3.cell(row, 1, 'BBSS / LIQUIDACIONES POR CLASIFICACION').font = Font(bold=True, name='Arial', size=11, color='0F766E')
row += 1

bbss_sum_cols = [c for c in bbss.columns if '(S/)' in c and c not in ['Codigo', 'PeriodoNum', 'Clasificacion', 'PartidaCuenta']]
if bbss_sum_cols:
    bbss_pivot = bbss.groupby(['PeriodoNum', 'Clasificacion'])[bbss_sum_cols].sum().reset_index()

    headers_bb = ['PERIODO', 'CLASIFICACION', 'PERSONAS'] + bbss_sum_cols[:8]
    for j, h in enumerate(headers_bb, 1):
        cell = ws3.cell(row, j, h[:30])
        cell.font = FNT_HDR
        cell.fill = PatternFill('solid', fgColor='2E75B6')
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = BORDER
    row += 1

    for _, r in bbss_pivot.sort_values(['PeriodoNum', 'Clasificacion']).iterrows():
        per = r['PeriodoNum']
        clasif = r['Clasificacion']
        count = len(bbss[(bbss['PeriodoNum'] == per) & (bbss['Clasificacion'] == clasif)])
        fill = {'Casa': FILL_CASA, 'Gastos Generales': FILL_GG, 'Intervencion': FILL_INTERV}.get(clasif, None)
        vals = [per, clasif, count] + [r[c] for c in bbss_sum_cols[:8]]
        for j, v in enumerate(vals, 1):
            cell = ws3.cell(row, j, v)
            cell.font = FNT_NORMAL
            cell.border = BORDER
            if j > 2:
                cell.number_format = NUM_FMT
            if fill:
                cell.fill = fill
        row += 1

# --- Hoja 4: PARTIDAS CONTROL ---
ws4 = wb.create_sheet('PARTIDAS CONTROL')
ws4.sheet_properties.tabColor = 'E74C3C'

ws4.cell(1, 1, 'CODIGO').font = FNT_HDR
ws4.cell(1, 1).fill = FILL_HEADER
ws4.cell(1, 2, 'NOMBRE').font = FNT_HDR
ws4.cell(1, 2).fill = FILL_HEADER
ws4.cell(1, 3, 'DNI').font = FNT_HDR
ws4.cell(1, 3).fill = FILL_HEADER
ws4.cell(1, 4, 'OCUPACION').font = FNT_HDR
ws4.cell(1, 4).fill = FILL_HEADER
ws4.cell(1, 5, 'PARTIDA').font = FNT_HDR
ws4.cell(1, 5).fill = FILL_HEADER
ws4.cell(1, 6, 'CLASIFICACION').font = FNT_HDR
ws4.cell(1, 6).fill = FILL_HEADER

ws4.column_dimensions['A'].width = 12
ws4.column_dimensions['B'].width = 36
ws4.column_dimensions['C'].width = 12
ws4.column_dimensions['D'].width = 32
ws4.column_dimensions['E'].width = 45
ws4.column_dimensions['F'].width = 18

for i, (_, r) in enumerate(partidas.iterrows(), 2):
    fill = {'Casa': FILL_CASA, 'Gastos Generales': FILL_GG, 'Intervencion': FILL_INTERV}.get(r['Clasificacion'], None)
    for j, v in enumerate([r['Codigo'], r['Nombre'], r['DNI'], r['Ocupacion'], r['Partida'], r['Clasificacion']], 1):
        cell = ws4.cell(i, j, v)
        cell.font = FNT_NORMAL
        cell.border = BORDER
        if fill:
            cell.fill = fill

ws4.auto_filter.ref = f'A1:F{len(partidas)+1}'

# ========================================================
# 5. GUARDAR
# ========================================================
out_path = r'C:\Users\EDWIN LOPEZ\OneDrive - RIPCON\Documentos\FLUJO DE CAJA DETALLADO - CONSORCIO SRT.xlsx'
wb.save(out_path)
print(f"\nGuardado: {out_path}")
print("COMPLETADO")
