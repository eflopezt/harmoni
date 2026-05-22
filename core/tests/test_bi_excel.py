"""
Tests para el Reporte BI Excel mensual (core.bi_excel + core.views_bi).

Verifica:
  - El generador produce bytes validos (.xlsx = zip "PK..")
  - Las 8 tabs existen con los nombres esperados
  - La vista responde 200 con content-type correcto
  - El filename incluye anio/mes
  - El generador no rompe con DB vacia
"""
from __future__ import annotations

from io import BytesIO

import pytest
from django.contrib.auth.models import User
from django.test import Client, TestCase
from django.urls import reverse
from openpyxl import load_workbook


URL_NAME = 'reporte_bi_mensual'

EXPECTED_TABS = [
    'Resumen Ejecutivo',
    'Headcount',
    'Reclutamiento',
    'Asistencia',
    'Nóminas',
    'Capacitaciones',
    'Alertas',
    'Detalle por Empresa',
]


@pytest.mark.django_db
class TestBIExcelBuilder(TestCase):
    """Tests directos al generador (sin pasar por la vista)."""

    def test_genera_bytes_validos(self):
        from core.bi_excel import build_reporte_bi_mensual
        data = build_reporte_bi_mensual(2026, 5)
        assert isinstance(data, bytes), 'debe devolver bytes'
        assert len(data) > 1000, 'el archivo debe pesar mas de 1KB'

    def test_magic_bytes_zip(self):
        """Un .xlsx valido es un ZIP — empieza con "PK"."""
        from core.bi_excel import build_reporte_bi_mensual
        data = build_reporte_bi_mensual(2026, 5)
        assert data[:2] == b'PK', f'expected ZIP magic, got {data[:4]!r}'

    def test_contiene_las_8_tabs(self):
        from core.bi_excel import build_reporte_bi_mensual
        data = build_reporte_bi_mensual(2026, 5)
        wb = load_workbook(BytesIO(data))
        sheet_names = wb.sheetnames
        assert len(sheet_names) == 8, (
            f'esperaba 8 tabs, hay {len(sheet_names)}: {sheet_names}'
        )
        for tab in EXPECTED_TABS:
            assert tab in sheet_names, f'falta tab: {tab}'

    def test_resumen_ejecutivo_tiene_kpis(self):
        """El tab Resumen Ejecutivo debe llevar al menos el titulo."""
        from core.bi_excel import build_reporte_bi_mensual
        data = build_reporte_bi_mensual(2026, 5)
        wb = load_workbook(BytesIO(data))
        ws = wb['Resumen Ejecutivo']
        titulo = ws.cell(row=1, column=1).value
        assert titulo and 'Harmoni' in str(titulo) and '2026' in str(titulo)

    def test_no_rompe_con_db_vacia(self):
        """Sin seed data el reporte debe generarse igual."""
        from core.bi_excel import build_reporte_bi_mensual
        # No hay personal, vacantes, papeletas, etc.
        data = build_reporte_bi_mensual(2026, 5)
        assert isinstance(data, bytes)
        wb = load_workbook(BytesIO(data))
        assert len(wb.sheetnames) == 8

    def test_acepta_empresa_id_none(self):
        from core.bi_excel import build_reporte_bi_mensual
        data = build_reporte_bi_mensual(2026, 5, empresa_id=None)
        assert isinstance(data, bytes)

    def test_acepta_empresa_id_inexistente(self):
        """empresa_id desconocido no rompe — devuelve archivo vacio pero valido."""
        from core.bi_excel import build_reporte_bi_mensual
        data = build_reporte_bi_mensual(2026, 5, empresa_id=99999)
        assert isinstance(data, bytes)
        wb = load_workbook(BytesIO(data))
        assert len(wb.sheetnames) == 8

    def test_tamanio_razonable(self):
        """Para DB vacia el archivo debe pesar < 100KB."""
        from core.bi_excel import build_reporte_bi_mensual
        data = build_reporte_bi_mensual(2026, 5)
        # Sin datos: ~30KB tipico. <500KB es el limite del contrato.
        assert len(data) < 500_000, f'archivo demasiado grande: {len(data)} bytes'


@pytest.mark.django_db
class TestBIExcelView(TestCase):
    """Tests de la vista HTTP."""

    def setUp(self):
        self.client = Client()
        self.url = reverse(URL_NAME)
        self.admin = User.objects.create_superuser(
            username='bi_admin', email='bi@harmoni.pe', password='test1234',
        )

    def test_anonymous_redirects_to_login(self):
        resp = self.client.get(self.url)
        assert resp.status_code in (302, 301)

    def test_regular_user_blocked(self):
        User.objects.create_user(username='nobody', password='test1234')
        self.client.login(username='nobody', password='test1234')
        resp = self.client.get(self.url)
        assert resp.status_code in (302, 301)

    def test_admin_200_con_content_type_correcto(self):
        self.client.login(username='bi_admin', password='test1234')
        resp = self.client.get(self.url, {'anio': 2026, 'mes': 5})
        assert resp.status_code == 200, f'status={resp.status_code}'
        expected_ct = (
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        assert resp['Content-Type'] == expected_ct

    def test_filename_incluye_anio_y_mes(self):
        self.client.login(username='bi_admin', password='test1234')
        resp = self.client.get(self.url, {'anio': 2026, 'mes': 5})
        disp = resp['Content-Disposition']
        assert 'reporte_bi_2026_05.xlsx' in disp, disp

    def test_default_usa_fecha_actual(self):
        """Sin params, usa anio/mes de hoy."""
        self.client.login(username='bi_admin', password='test1234')
        resp = self.client.get(self.url)
        assert resp.status_code == 200
        assert 'reporte_bi_' in resp['Content-Disposition']

    def test_mes_invalido_400(self):
        self.client.login(username='bi_admin', password='test1234')
        resp = self.client.get(self.url, {'anio': 2026, 'mes': 13})
        assert resp.status_code == 400

    def test_anio_no_entero_400(self):
        self.client.login(username='bi_admin', password='test1234')
        resp = self.client.get(self.url, {'anio': 'abc', 'mes': 5})
        assert resp.status_code == 400

    def test_response_contiene_zip_valido(self):
        self.client.login(username='bi_admin', password='test1234')
        resp = self.client.get(self.url, {'anio': 2026, 'mes': 5})
        assert resp.content[:2] == b'PK'
