"""
Reporte BI Mensual — Excel ejecutivo cross-modulo.

Genera un Excel con 8 tabs que consolida metricas de TODAS las areas de RRHH:
  1. Resumen Ejecutivo (12 KPIs)
  2. Headcount (areas, edades, contratos)
  3. Reclutamiento (funnel, top vacantes, conversion)
  4. Asistencia (papeletas, tipos, top trabajadores, briefings)
  5. Nominas (bruto/neto, comparativo, top cargos)
  6. Capacitaciones (BPM/HACCP, vencidos)
  7. Alertas y acciones urgentes
  8. Detalle por empresa (multi-tenant)

Disenio defensivo: cada bloque va en try/except — si un modelo no existe,
falla un import o un campo cambio, se devuelve "—" o 0 sin romper el archivo.
"""
from __future__ import annotations

import calendar
import logging
from collections import OrderedDict
from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

logger = logging.getLogger('core.bi_excel')


# ── Paleta Harmoni ────────────────────────────────────────────────────
COLOR_TEAL     = '0D2B27'   # Harmoni primary
COLOR_TEAL_LT  = '0F766E'
COLOR_AMBAR    = 'F59E0B'
COLOR_ROJO     = 'DC2626'
COLOR_VERDE    = '16A34A'
COLOR_AZUL     = '1D4ED8'
COLOR_VIOLETA  = '7C3AED'
COLOR_GRIS     = '64748B'
COLOR_BG_SOFT  = 'F8FAFC'
COLOR_BORDER   = 'E2E8F0'

# Estilos reutilizables
HEADER_FILL   = PatternFill('solid', fgColor=COLOR_TEAL)
HEADER_FONT   = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
SUBHEADER_FILL = PatternFill('solid', fgColor=COLOR_TEAL_LT)
SUBHEADER_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
TITLE_FONT    = Font(name='Calibri', size=18, bold=True, color=COLOR_TEAL)
KPI_FONT      = Font(name='Calibri', size=22, bold=True, color=COLOR_TEAL)
KPI_LABEL_FONT = Font(name='Calibri', size=10, bold=True, color=COLOR_GRIS)
ALERT_FONT    = Font(name='Calibri', size=11, bold=True, color=COLOR_ROJO)
WARN_FILL     = PatternFill('solid', fgColor='FEF3C7')
DANGER_FILL   = PatternFill('solid', fgColor='FEE2E2')
OK_FILL       = PatternFill('solid', fgColor='DCFCE7')
SOFT_FILL     = PatternFill('solid', fgColor=COLOR_BG_SOFT)

THIN_BORDER = Border(
    left=Side(style='thin', color=COLOR_BORDER),
    right=Side(style='thin', color=COLOR_BORDER),
    top=Side(style='thin', color=COLOR_BORDER),
    bottom=Side(style='thin', color=COLOR_BORDER),
)

MESES_ES = [
    '', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
    'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre',
]


# ══════════════════════════════════════════════════════════════════════
# Helpers de estilo / utilidades
# ══════════════════════════════════════════════════════════════════════

def _safe(fn, default=None):
    """Ejecuta fn() y devuelve default si lanza cualquier excepcion."""
    try:
        return fn()
    except Exception as exc:  # noqa: BLE001
        logger.warning('bi_excel: bloque fallo: %s', exc)
        return default


def _apply_header(ws, row, cols, text, fill=HEADER_FILL, font=HEADER_FONT, height=30):
    """Aplica fila de header (titulo de tab o seccion)."""
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill = fill
    cell.font = font
    cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    ws.row_dimensions[row].height = height


def _apply_table_header(ws, row, headers, start_col=1):
    """Aplica cabecera de tabla."""
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=start_col + i, value=h)
        cell.fill = SUBHEADER_FILL
        cell.font = SUBHEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 22


def _apply_table_row(ws, row, values, start_col=1, fill=None, bold=False, number_formats=None):
    """Llena una fila de datos con formato."""
    for i, val in enumerate(values):
        cell = ws.cell(row=row, column=start_col + i, value=val)
        cell.border = THIN_BORDER
        cell.font = Font(name='Calibri', size=10, bold=bold)
        if fill:
            cell.fill = fill
        if number_formats and i < len(number_formats) and number_formats[i]:
            cell.number_format = number_formats[i]
            cell.alignment = Alignment(horizontal='right', vertical='center')
        else:
            cell.alignment = Alignment(horizontal='left', vertical='center')


def _autosize_columns(ws, min_width=10, max_width=45):
    """Ajusta ancho de columnas segun contenido (skip merged cells)."""
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            # cell.column puede no estar para MergedCell
            col_letter = getattr(cell, 'column_letter', None)
            if not col_letter:
                continue
            try:
                length = len(str(cell.value)) if cell.value is not None else 0
            except Exception:
                length = 0
            if length > widths.get(col_letter, 0):
                widths[col_letter] = length
    for col_letter, length in widths.items():
        ws.column_dimensions[col_letter].width = min(max(length + 2, min_width), max_width)


def _periodo_label(anio, mes):
    return f'{MESES_ES[mes]} {anio}'


def _rango_mes(anio, mes):
    """Retorna (date_inicio, date_fin) inclusivo del mes."""
    ultimo = calendar.monthrange(anio, mes)[1]
    return date(anio, mes, 1), date(anio, mes, ultimo)


def _mes_anterior(anio, mes):
    if mes == 1:
        return anio - 1, 12
    return anio, mes - 1


def _fmt_num(v):
    """Devuelve string formateado o '—' si None/0 sin sentido."""
    if v is None:
        return '—'
    return v


# ══════════════════════════════════════════════════════════════════════
# Recoleccion de datos por seccion
# Cada funcion es defensiva: retorna dict con shape esperada o {}
# ══════════════════════════════════════════════════════════════════════

def _resumen_ejecutivo_data(anio, mes, empresa_id):
    """Calcula los 12 KPIs del resumen ejecutivo."""
    d_ini, d_fin = _rango_mes(anio, mes)
    data = {
        'headcount_total': 0,
        'activos': 0,
        'cesados_mes': 0,
        'altas_mes': 0,
        'vacantes_abiertas': 0,
        'postulaciones_mes': 0,
        'contratados_mes': 0,
        'total_nomina': Decimal('0'),
        'total_papeletas': 0,
        'certs_vigentes': 0,
        'certs_vencidas': 0,
        'capacitaciones_mes': 0,
    }

    def _q_personal():
        from personal.models import Personal
        qs = Personal.objects.all()
        if empresa_id:
            qs = qs.filter(empresa_id=empresa_id)
        data['headcount_total'] = qs.count()
        data['activos'] = qs.filter(estado='Activo').count()
        data['cesados_mes'] = qs.filter(
            fecha_cese__gte=d_ini, fecha_cese__lte=d_fin
        ).count()
        data['altas_mes'] = qs.filter(
            fecha_alta__gte=d_ini, fecha_alta__lte=d_fin
        ).count()
    _safe(_q_personal)

    def _q_reclu():
        from reclutamiento.models import Vacante, Postulacion
        data['vacantes_abiertas'] = Vacante.objects.filter(
            estado__in=['PUBLICADA', 'EN_PROCESO']
        ).count()
        data['postulaciones_mes'] = Postulacion.objects.filter(
            fecha_postulacion__date__gte=d_ini,
            fecha_postulacion__date__lte=d_fin,
        ).count()
        data['contratados_mes'] = Postulacion.objects.filter(
            estado='CONTRATADA',
            fecha_postulacion__date__gte=d_ini,
            fecha_postulacion__date__lte=d_fin,
        ).count()
    _safe(_q_reclu)

    def _q_nomina():
        from nominas.models import PeriodoNomina
        from django.db.models import Sum
        qs = PeriodoNomina.objects.filter(anio=anio, mes=mes, tipo='REGULAR')
        if empresa_id:
            qs = qs.filter(empresa_id=empresa_id)
        agg = qs.aggregate(total=Sum('total_bruto'))
        data['total_nomina'] = agg['total'] or Decimal('0')
    _safe(_q_nomina)

    def _q_papel():
        from asistencia.models import RegistroPapeleta
        qs = RegistroPapeleta.objects.filter(
            fecha_inicio__gte=d_ini, fecha_inicio__lte=d_fin
        )
        if empresa_id:
            qs = qs.filter(personal__empresa_id=empresa_id)
        data['total_papeletas'] = qs.count()
    _safe(_q_papel)

    def _q_certs():
        from capacitaciones.models import CertificacionTrabajador
        qs = CertificacionTrabajador.objects.all()
        if empresa_id:
            qs = qs.filter(personal__empresa_id=empresa_id)
        data['certs_vigentes'] = qs.filter(estado='VIGENTE').count()
        data['certs_vencidas'] = qs.filter(estado='VENCIDA').count()
    _safe(_q_certs)

    def _q_capac():
        from capacitaciones.models import Capacitacion
        data['capacitaciones_mes'] = Capacitacion.objects.filter(
            fecha_inicio__gte=d_ini, fecha_inicio__lte=d_fin,
            estado='COMPLETADA',
        ).count()
    _safe(_q_capac)

    return data


def _headcount_data(anio, mes, empresa_id):
    """Datos para tab Headcount."""
    d_ini, d_fin = _rango_mes(anio, mes)
    out = {
        'por_empresa': [],     # [(nombre, activos, cesados_mes, altas_mes, neto)]
        'por_area': [],        # [(area, count)]
        'por_contrato': [],    # [(tipo, count)]
        'piramide': [],        # [(rango, count)]
    }

    def _calc():
        from personal.models import Personal
        from django.db.models import Count, Q
        base = Personal.objects.all()
        if empresa_id:
            base = base.filter(empresa_id=empresa_id)

        # Por empresa
        try:
            from empresas.models import Empresa
            emp_qs = Empresa.objects.all()
            if empresa_id:
                emp_qs = emp_qs.filter(pk=empresa_id)
            for e in emp_qs:
                p_emp = base.filter(empresa=e)
                activos = p_emp.filter(estado='Activo').count()
                cesados = p_emp.filter(fecha_cese__gte=d_ini, fecha_cese__lte=d_fin).count()
                altas = p_emp.filter(fecha_alta__gte=d_ini, fecha_alta__lte=d_fin).count()
                if activos or cesados or altas:
                    out['por_empresa'].append((
                        e.razon_social or e.ruc, activos, cesados, altas, altas - cesados,
                    ))
        except Exception as exc:
            logger.warning('headcount por_empresa fallo: %s', exc)

        # Por area (top 10)
        try:
            por_area = (
                base.filter(estado='Activo')
                .values('subarea__area__nombre')
                .annotate(c=Count('id'))
                .order_by('-c')[:10]
            )
            out['por_area'] = [
                (r['subarea__area__nombre'] or 'Sin asignar', r['c'])
                for r in por_area
            ]
        except Exception as exc:
            logger.warning('headcount por_area fallo: %s', exc)

        # Por tipo de contrato
        try:
            por_contrato = (
                base.filter(estado='Activo')
                .exclude(tipo_contrato='')
                .values('tipo_contrato')
                .annotate(c=Count('id'))
                .order_by('-c')
            )
            out['por_contrato'] = [
                (r['tipo_contrato'] or 'Sin especificar', r['c'])
                for r in por_contrato
            ]
        except Exception as exc:
            logger.warning('headcount por_contrato fallo: %s', exc)

        # Piramide de edades
        try:
            hoy = date.today()
            rangos = [
                ('18-25', 18, 25),
                ('26-35', 26, 35),
                ('36-45', 36, 45),
                ('46-55', 46, 55),
                ('56+',   56, 120),
            ]
            for label, mn, mx in rangos:
                fecha_max = hoy.replace(year=hoy.year - mn)
                fecha_min = hoy.replace(year=hoy.year - mx - 1) + timedelta(days=1)
                c = base.filter(
                    estado='Activo',
                    fecha_nacimiento__gte=fecha_min,
                    fecha_nacimiento__lte=fecha_max,
                ).count()
                out['piramide'].append((label, c))
        except Exception as exc:
            logger.warning('headcount piramide fallo: %s', exc)

    _safe(_calc)
    return out


def _reclutamiento_data(anio, mes, empresa_id):
    """Datos para tab Reclutamiento."""
    d_ini, d_fin = _rango_mes(anio, mes)
    out = {
        'kpis': {'abiertas': 0, 'postulaciones': 0, 'contrataciones': 0, 'conversion': 0.0},
        'top_vacantes': [],   # [(titulo, total)]
        'funnel': [],         # [(etapa, count)]
        'fuentes': [],        # [(fuente, total, contratadas, %conv)]
    }

    def _calc():
        from reclutamiento.models import Vacante, Postulacion, EtapaPipeline
        from django.db.models import Count, Q

        out['kpis']['abiertas'] = Vacante.objects.filter(
            estado__in=['PUBLICADA', 'EN_PROCESO']
        ).count()
        postul_mes = Postulacion.objects.filter(
            fecha_postulacion__date__gte=d_ini,
            fecha_postulacion__date__lte=d_fin,
        )
        out['kpis']['postulaciones'] = postul_mes.count()
        contrats = postul_mes.filter(estado='CONTRATADA').count()
        out['kpis']['contrataciones'] = contrats
        if out['kpis']['postulaciones']:
            out['kpis']['conversion'] = round(
                100 * contrats / out['kpis']['postulaciones'], 1
            )

        # Top 10 vacantes con mas candidatos (lifetime)
        top = (
            Vacante.objects.annotate(total=Count('postulaciones'))
            .order_by('-total')[:10]
        )
        out['top_vacantes'] = [(v.titulo, v.total) for v in top]

        # Funnel por etapa (postulaciones ACTIVAS)
        funnel_q = (
            EtapaPipeline.objects.filter(activa=True)
            .annotate(total=Count('postulaciones',
                                  filter=Q(postulaciones__estado='ACTIVA')))
            .order_by('orden')
        )
        out['funnel'] = [(e.nombre, e.total) for e in funnel_q]

        # Fuentes con mejor conversion
        fuentes = {}
        for r in postul_mes.values('fuente').annotate(total=Count('id')):
            fuentes[r['fuente']] = {'total': r['total'], 'cont': 0}
        for r in (
            postul_mes.filter(estado='CONTRATADA')
            .values('fuente').annotate(c=Count('id'))
        ):
            if r['fuente'] in fuentes:
                fuentes[r['fuente']]['cont'] = r['c']
        for codigo, d in fuentes.items():
            conv = round(100 * d['cont'] / d['total'], 1) if d['total'] else 0
            out['fuentes'].append((codigo, d['total'], d['cont'], conv))
        out['fuentes'].sort(key=lambda x: x[3], reverse=True)

    _safe(_calc)
    return out


def _asistencia_data(anio, mes, empresa_id):
    """Datos para tab Asistencia."""
    d_ini, d_fin = _rango_mes(anio, mes)
    out = {
        'total_papeletas': 0,
        'por_tipo': [],          # [(tipo, count)]
        'top_personas': [],      # [(nombre, count)]
        'briefings_pub': 0,
        'briefings_borrador': 0,
    }

    def _calc():
        from asistencia.models import RegistroPapeleta
        from django.db.models import Count
        qs = RegistroPapeleta.objects.filter(
            fecha_inicio__gte=d_ini, fecha_inicio__lte=d_fin,
        )
        if empresa_id:
            qs = qs.filter(personal__empresa_id=empresa_id)
        out['total_papeletas'] = qs.count()

        # Por tipo
        por_tipo = qs.values('tipo_permiso').annotate(c=Count('id')).order_by('-c')
        # Mapeo legible
        choices = dict(RegistroPapeleta.TIPO_PERMISO_CHOICES)
        out['por_tipo'] = [
            (choices.get(r['tipo_permiso'], r['tipo_permiso']), r['c'])
            for r in por_tipo
        ]

        # Top 10 personas con mas papeletas
        top = (
            qs.exclude(personal__isnull=True)
            .values('personal__apellidos_nombres')
            .annotate(c=Count('id'))
            .order_by('-c')[:10]
        )
        out['top_personas'] = [(r['personal__apellidos_nombres'], r['c']) for r in top]
    _safe(_calc)

    def _bri():
        from asistencia.models import BriefingServicio
        qs = BriefingServicio.objects.filter(fecha__gte=d_ini, fecha__lte=d_fin)
        if empresa_id:
            qs = qs.filter(empresa_id=empresa_id)
        out['briefings_pub'] = qs.filter(estado__in=['PUBLICADO', 'CERRADO']).count()
        out['briefings_borrador'] = qs.filter(estado='BORRADOR').count()
    _safe(_bri)

    return out


def _nominas_data(anio, mes, empresa_id):
    """Datos para tab Nominas."""
    ant_a, ant_m = _mes_anterior(anio, mes)
    out = {
        'total_bruto': Decimal('0'),
        'total_descuentos': Decimal('0'),
        'total_neto': Decimal('0'),
        'total_costo_empresa': Decimal('0'),
        'total_trabajadores': 0,
        'bruto_anterior': Decimal('0'),
        'var_pct': None,
        'top_cargos': [],        # [(cargo, count, promedio_neto)]
        'descuentos': [],        # [(tipo, monto)]
    }

    def _calc():
        from nominas.models import PeriodoNomina, RegistroNomina, LineaNomina
        from django.db.models import Sum, Avg, Count

        qs = PeriodoNomina.objects.filter(anio=anio, mes=mes, tipo='REGULAR')
        if empresa_id:
            qs = qs.filter(empresa_id=empresa_id)
        agg = qs.aggregate(
            b=Sum('total_bruto'),
            d=Sum('total_descuentos'),
            n=Sum('total_neto'),
            c=Sum('total_costo_empresa'),
            t=Sum('total_trabajadores'),
        )
        out['total_bruto'] = agg['b'] or Decimal('0')
        out['total_descuentos'] = agg['d'] or Decimal('0')
        out['total_neto'] = agg['n'] or Decimal('0')
        out['total_costo_empresa'] = agg['c'] or Decimal('0')
        out['total_trabajadores'] = agg['t'] or 0

        # Mes anterior
        qs_ant = PeriodoNomina.objects.filter(anio=ant_a, mes=ant_m, tipo='REGULAR')
        if empresa_id:
            qs_ant = qs_ant.filter(empresa_id=empresa_id)
        agg_ant = qs_ant.aggregate(b=Sum('total_bruto'))
        out['bruto_anterior'] = agg_ant['b'] or Decimal('0')
        if out['bruto_anterior']:
            out['var_pct'] = float(
                (out['total_bruto'] - out['bruto_anterior']) / out['bruto_anterior'] * 100
            )

        # Top cargos por costo promedio
        regs = RegistroNomina.objects.filter(periodo__in=qs)
        top = (
            regs.values('personal__cargo')
            .annotate(c=Count('id'), promedio=Avg('neto_a_pagar'))
            .order_by('-promedio')[:10]
        )
        out['top_cargos'] = [
            (r['personal__cargo'] or 'Sin cargo', r['c'], r['promedio'] or Decimal('0'))
            for r in top
        ]

        # Descuentos por concepto (AFP, ONP, etc)
        try:
            from nominas.models import LineaNomina
            lineas = (
                LineaNomina.objects
                .filter(registro__periodo__in=qs, concepto__tipo='DESCUENTO')
                .values('concepto__nombre')
                .annotate(m=Sum('monto'))
                .order_by('-m')[:10]
            )
            out['descuentos'] = [(r['concepto__nombre'], r['m'] or Decimal('0')) for r in lineas]
        except Exception:
            pass

    _safe(_calc)
    return out


def _capacitaciones_data(anio, mes, empresa_id):
    """Datos para tab Capacitaciones."""
    d_ini, d_fin = _rango_mes(anio, mes)
    hoy = date.today()
    out = {
        'capacitaciones_mes': 0,
        'asistentes_unicos': 0,
        'asistentes_invitados': 0,
        'certs_vigentes': 0,
        'certs_por_vencer': 0,
        'certs_vencidas': 0,
        'lista_vencidas': [],   # [(personal, certificacion, dias_vencida)]
    }

    def _calc():
        from capacitaciones.models import (
            Capacitacion, AsistenciaCapacitacion, CertificacionTrabajador,
        )
        from django.db.models import Count

        capac_mes = Capacitacion.objects.filter(
            fecha_inicio__gte=d_ini, fecha_inicio__lte=d_fin,
        )
        out['capacitaciones_mes'] = capac_mes.filter(estado='COMPLETADA').count()

        asist = AsistenciaCapacitacion.objects.filter(capacitacion__in=capac_mes)
        if empresa_id:
            asist = asist.filter(personal__empresa_id=empresa_id)
        out['asistentes_invitados'] = asist.count()
        out['asistentes_unicos'] = asist.filter(estado='ASISTIO').values('personal').distinct().count()

        certs = CertificacionTrabajador.objects.all()
        if empresa_id:
            certs = certs.filter(personal__empresa_id=empresa_id)
        out['certs_vigentes'] = certs.filter(estado='VIGENTE').count()
        out['certs_por_vencer'] = certs.filter(estado='POR_VENCER').count()
        out['certs_vencidas'] = certs.filter(estado='VENCIDA').count()

        # Lista de vencidas (top 20 accion urgente)
        for c in (
            certs.filter(estado='VENCIDA')
            .select_related('personal', 'requerimiento')
            .order_by('fecha_vencimiento')[:20]
        ):
            dias = (hoy - c.fecha_vencimiento).days if c.fecha_vencimiento else 0
            out['lista_vencidas'].append((
                c.personal.apellidos_nombres if c.personal else '—',
                c.requerimiento.nombre if c.requerimiento else '—',
                dias,
            ))

    _safe(_calc)
    return out


def _alertas_data(anio, mes, empresa_id):
    """Recolecta listado de alertas / accion urgente."""
    hoy = date.today()
    out = {
        'certs_vencidas': 0,
        'papeletas_pendientes': 0,
        'vacantes_sin_actividad': 0,
        'briefings_pendientes': 0,
        'postulaciones_estancadas': 0,
        'contratos_por_vencer': 0,
    }

    def _certs():
        from capacitaciones.models import CertificacionTrabajador
        qs = CertificacionTrabajador.objects.filter(estado='VENCIDA')
        if empresa_id:
            qs = qs.filter(personal__empresa_id=empresa_id)
        out['certs_vencidas'] = qs.count()
    _safe(_certs)

    def _papeletas():
        from asistencia.models import RegistroPapeleta
        umbral = hoy - timedelta(days=7)
        qs = RegistroPapeleta.objects.filter(
            estado='PENDIENTE',
            creado_en__date__lte=umbral,
        )
        if empresa_id:
            qs = qs.filter(personal__empresa_id=empresa_id)
        out['papeletas_pendientes'] = qs.count()
    _safe(_papeletas)

    def _vacantes():
        from reclutamiento.models import Vacante
        umbral = hoy - timedelta(days=30)
        out['vacantes_sin_actividad'] = Vacante.objects.filter(
            estado__in=['PUBLICADA', 'EN_PROCESO'],
            actualizado_en__date__lte=umbral,
        ).count()
    _safe(_vacantes)

    def _briefings():
        from asistencia.models import BriefingServicio
        umbral = hoy - timedelta(days=7)
        qs = BriefingServicio.objects.filter(
            fecha__gte=umbral, fecha__lte=hoy,
            estado='BORRADOR',
        )
        if empresa_id:
            qs = qs.filter(empresa_id=empresa_id)
        out['briefings_pendientes'] = qs.count()
    _safe(_briefings)

    def _postul():
        from reclutamiento.models import Postulacion
        umbral = hoy - timedelta(days=14)
        out['postulaciones_estancadas'] = Postulacion.objects.filter(
            estado='ACTIVA',
            fecha_postulacion__date__lte=umbral,
        ).count()
    _safe(_postul)

    def _contratos():
        from personal.models import Personal
        umbral_max = hoy + timedelta(days=30)
        qs = Personal.objects.filter(
            estado='Activo',
            fecha_fin_contrato__gte=hoy,
            fecha_fin_contrato__lte=umbral_max,
        )
        if empresa_id:
            qs = qs.filter(empresa_id=empresa_id)
        out['contratos_por_vencer'] = qs.count()
    _safe(_contratos)

    return out


def _detalle_empresas_data(anio, mes, empresa_id):
    """Datos para tab detalle por empresa."""
    d_ini, d_fin = _rango_mes(anio, mes)
    rows = []
    try:
        from empresas.models import Empresa
        from personal.models import Personal
        from asistencia.models import RegistroPapeleta
        from capacitaciones.models import Capacitacion, CertificacionTrabajador

        qs = Empresa.objects.all().order_by('razon_social')
        if empresa_id:
            qs = qs.filter(pk=empresa_id)

        for e in qs:
            trab = Personal.objects.filter(empresa=e, estado='Activo').count()
            papel = _safe(
                lambda: RegistroPapeleta.objects.filter(
                    personal__empresa=e,
                    fecha_inicio__gte=d_ini, fecha_inicio__lte=d_fin,
                ).count(),
                0,
            )
            capac = _safe(
                lambda: Capacitacion.objects.filter(
                    fecha_inicio__gte=d_ini, fecha_inicio__lte=d_fin,
                    estado='COMPLETADA',
                ).count(),
                0,
            )
            cert_v = _safe(
                lambda: CertificacionTrabajador.objects.filter(
                    personal__empresa=e, estado='VENCIDA',
                ).count(),
                0,
            )
            rows.append((e.razon_social or e.ruc, trab, papel, capac, cert_v))
    except Exception as exc:
        logger.warning('detalle empresas fallo: %s', exc)
    return rows


# ══════════════════════════════════════════════════════════════════════
# Constructores de cada tab
# ══════════════════════════════════════════════════════════════════════

def _build_tab_resumen(wb, anio, mes, data):
    ws = wb.create_sheet('Resumen Ejecutivo')
    ws.sheet_properties.tabColor = COLOR_TEAL

    # Banner
    _apply_header(
        ws, 1, 6,
        f'Harmoni · Reporte BI Mensual — {_periodo_label(anio, mes)}',
        height=36,
    )
    ws.cell(row=2, column=1,
            value='Reporte ejecutivo de Recursos Humanos · uso interno Gerencia').font = Font(
        name='Calibri', size=10, italic=True, color=COLOR_GRIS,
    )
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)

    # KPI cards (2 columnas × 6 filas)
    kpis = [
        ('Headcount Total',         data['headcount_total'],       'fa-users',         COLOR_TEAL),
        ('Activos',                 data['activos'],                'fa-user-check',    COLOR_VERDE),
        ('Cesados (mes)',           data['cesados_mes'],            'fa-user-minus',    COLOR_ROJO),
        ('Altas (mes)',             data['altas_mes'],              'fa-user-plus',     COLOR_AZUL),
        ('Vacantes abiertas',       data['vacantes_abiertas'],      'fa-bullseye',      COLOR_VIOLETA),
        ('Postulaciones (mes)',     data['postulaciones_mes'],      'fa-file-signature', COLOR_TEAL_LT),
        ('Contratados (mes)',       data['contratados_mes'],        'fa-handshake',     COLOR_VERDE),
        ('Total Nómina S/',         float(data['total_nomina']),    'fa-money-bill',    COLOR_AMBAR),
        ('Papeletas (mes)',         data['total_papeletas'],        'fa-clipboard',     COLOR_AZUL),
        ('Certs. vigentes',         data['certs_vigentes'],         'fa-certificate',   COLOR_VERDE),
        ('Certs. vencidas',         data['certs_vencidas'],         'fa-exclamation',   COLOR_ROJO),
        ('Capacitaciones (mes)',    data['capacitaciones_mes'],     'fa-graduation-cap', COLOR_TEAL),
    ]

    start_row = 4
    icons_unicode = ['👥', '✅', '👤', '➕', '🎯', '📝', '🤝', '💰', '📋', '🏅', '⚠️', '🎓']
    for idx, ((label, value, _, color), icon) in enumerate(zip(kpis, icons_unicode)):
        col = 1 if idx % 2 == 0 else 4
        row = start_row + (idx // 2) * 3

        # Icon + label
        cell = ws.cell(row=row, column=col, value=f'{icon}  {label}')
        cell.font = KPI_LABEL_FONT
        cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 2)

        # Value
        val_cell = ws.cell(row=row + 1, column=col, value=value)
        val_cell.font = Font(name='Calibri', size=22, bold=True, color=color)
        val_cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        if isinstance(value, float):
            val_cell.number_format = '#,##0.00'
        else:
            val_cell.number_format = '#,##0'
        ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 2)

        # Decoracion: borde y fill
        for r in (row, row + 1):
            for c in range(col, col + 3):
                cc = ws.cell(row=r, column=c)
                cc.border = THIN_BORDER
                if not isinstance(cc, type(val_cell)) or cc.fill.fgColor.rgb in (None, '00000000'):
                    cc.fill = SOFT_FILL

        ws.row_dimensions[row].height = 18
        ws.row_dimensions[row + 1].height = 28

    # Anchos de columna
    for col_idx in range(1, 7):
        ws.column_dimensions[get_column_letter(col_idx)].width = 22

    ws.freeze_panes = 'A4'


def _build_tab_headcount(wb, anio, mes, data):
    ws = wb.create_sheet('Headcount')
    ws.sheet_properties.tabColor = COLOR_AZUL

    _apply_header(ws, 1, 5, f'👥 Headcount — {_periodo_label(anio, mes)}')

    row = 3
    _apply_table_header(
        ws, row,
        ['Empresa', 'Activos', 'Cesados (mes)', 'Altas (mes)', 'Neto'],
    )
    row += 1
    if data['por_empresa']:
        for nombre, act, ces, alt, neto in data['por_empresa']:
            _apply_table_row(
                ws, row,
                [nombre, act, ces, alt, neto],
                number_formats=[None, '#,##0', '#,##0', '#,##0', '+#,##0;-#,##0;0'],
            )
            row += 1
    else:
        _apply_table_row(ws, row, ['Sin datos', '—', '—', '—', '—'])
        row += 1

    # Distribucion por area
    row += 2
    _apply_header(ws, row, 5, 'Distribución por Área (Top 10)', fill=SUBHEADER_FILL,
                  font=SUBHEADER_FONT, height=22)
    row += 1
    _apply_table_header(ws, row, ['Área', 'Activos'])
    row += 1
    if data['por_area']:
        for area, c in data['por_area']:
            _apply_table_row(ws, row, [area, c], number_formats=[None, '#,##0'])
            row += 1
    else:
        _apply_table_row(ws, row, ['Sin datos', 0])
        row += 1

    # Por tipo de contrato
    row += 2
    _apply_header(ws, row, 5, 'Distribución por Tipo de Contrato', fill=SUBHEADER_FILL,
                  font=SUBHEADER_FONT, height=22)
    row += 1
    _apply_table_header(ws, row, ['Tipo de Contrato', 'Cantidad'])
    row += 1
    if data['por_contrato']:
        for tipo, c in data['por_contrato']:
            _apply_table_row(ws, row, [tipo, c], number_formats=[None, '#,##0'])
            row += 1
    else:
        _apply_table_row(ws, row, ['Sin datos', 0])
        row += 1

    # Piramide de edades
    row += 2
    _apply_header(ws, row, 5, 'Pirámide de Edades', fill=SUBHEADER_FILL,
                  font=SUBHEADER_FONT, height=22)
    row += 1
    _apply_table_header(ws, row, ['Rango', 'Trabajadores'])
    row += 1
    for rango, c in data['piramide']:
        _apply_table_row(ws, row, [rango, c], number_formats=[None, '#,##0'])
        row += 1

    _autosize_columns(ws)
    ws.freeze_panes = 'A4'


def _build_tab_reclutamiento(wb, anio, mes, data):
    ws = wb.create_sheet('Reclutamiento')
    ws.sheet_properties.tabColor = COLOR_VIOLETA

    _apply_header(ws, 1, 4, f'🎯 Reclutamiento — {_periodo_label(anio, mes)}')

    # KPIs en fila
    row = 3
    headers = ['Vacantes abiertas', 'Postulaciones (mes)', 'Contrataciones (mes)', 'Conversión %']
    _apply_table_header(ws, row, headers)
    row += 1
    k = data['kpis']
    _apply_table_row(
        ws, row,
        [k['abiertas'], k['postulaciones'], k['contrataciones'], k['conversion']],
        number_formats=['#,##0', '#,##0', '#,##0', '0.0"%"'],
        bold=True,
    )
    ws.row_dimensions[row].height = 28
    row += 2

    # Top 10 vacantes
    _apply_header(ws, row, 4, 'Top 10 Vacantes con más candidatos', fill=SUBHEADER_FILL,
                  font=SUBHEADER_FONT, height=22)
    row += 1
    _apply_table_header(ws, row, ['Vacante', 'Total candidatos'])
    row += 1
    if data['top_vacantes']:
        for titulo, total in data['top_vacantes']:
            _apply_table_row(ws, row, [titulo, total], number_formats=[None, '#,##0'])
            row += 1
    else:
        _apply_table_row(ws, row, ['Sin datos', 0])
        row += 1

    # Funnel
    row += 2
    _apply_header(ws, row, 4, 'Funnel por Etapa (postulaciones ACTIVAS)', fill=SUBHEADER_FILL,
                  font=SUBHEADER_FONT, height=22)
    row += 1
    _apply_table_header(ws, row, ['Etapa', 'Candidatos'])
    row += 1
    if data['funnel']:
        for etapa, c in data['funnel']:
            _apply_table_row(ws, row, [etapa, c], number_formats=[None, '#,##0'])
            row += 1
    else:
        _apply_table_row(ws, row, ['Sin datos', 0])
        row += 1

    # Fuentes
    row += 2
    _apply_header(ws, row, 4, 'Fuentes con mejor conversión', fill=SUBHEADER_FILL,
                  font=SUBHEADER_FONT, height=22)
    row += 1
    _apply_table_header(ws, row, ['Fuente', 'Postulaciones', 'Contratadas', 'Conv. %'])
    row += 1
    if data['fuentes']:
        for fuente, total, cont, conv in data['fuentes']:
            _apply_table_row(
                ws, row,
                [fuente, total, cont, conv],
                number_formats=[None, '#,##0', '#,##0', '0.0"%"'],
            )
            row += 1
    else:
        _apply_table_row(ws, row, ['Sin datos', 0, 0, 0])
        row += 1

    _autosize_columns(ws)
    ws.freeze_panes = 'A4'


def _build_tab_asistencia(wb, anio, mes, data):
    ws = wb.create_sheet('Asistencia')
    ws.sheet_properties.tabColor = COLOR_TEAL_LT

    _apply_header(ws, 1, 3, f'⏰ Asistencia — {_periodo_label(anio, mes)}')

    row = 3
    cell = ws.cell(row=row, column=1, value=f'Total papeletas del mes: {data["total_papeletas"]:,}')
    cell.font = Font(name='Calibri', size=14, bold=True, color=COLOR_TEAL)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 2

    # Por tipo
    _apply_header(ws, row, 3, 'Distribución por Tipo de Permiso', fill=SUBHEADER_FILL,
                  font=SUBHEADER_FONT, height=22)
    row += 1
    _apply_table_header(ws, row, ['Tipo de Permiso', 'Cantidad'])
    row += 1
    if data['por_tipo']:
        for tipo, c in data['por_tipo']:
            _apply_table_row(ws, row, [tipo, c], number_formats=[None, '#,##0'])
            row += 1
    else:
        _apply_table_row(ws, row, ['Sin datos', 0])
        row += 1

    # Top personas
    row += 2
    _apply_header(ws, row, 3, 'Top 10 Trabajadores con más papeletas (alerta)',
                  fill=SUBHEADER_FILL, font=SUBHEADER_FONT, height=22)
    row += 1
    _apply_table_header(ws, row, ['Trabajador', 'Papeletas'])
    row += 1
    if data['top_personas']:
        for nombre, c in data['top_personas']:
            fill = WARN_FILL if c >= 5 else None
            _apply_table_row(ws, row, [nombre, c], number_formats=[None, '#,##0'], fill=fill)
            row += 1
    else:
        _apply_table_row(ws, row, ['Sin datos', 0])
        row += 1

    # Briefings
    row += 2
    _apply_header(ws, row, 3, 'Briefings de Servicio (Gastronomía)',
                  fill=SUBHEADER_FILL, font=SUBHEADER_FONT, height=22)
    row += 1
    _apply_table_header(ws, row, ['Estado', 'Cantidad'])
    row += 1
    _apply_table_row(ws, row, ['Publicados / Cerrados', data['briefings_pub']],
                     number_formats=[None, '#,##0'], fill=OK_FILL)
    row += 1
    _apply_table_row(ws, row, ['Borrador (no publicado)', data['briefings_borrador']],
                     number_formats=[None, '#,##0'],
                     fill=WARN_FILL if data['briefings_borrador'] else None)
    row += 1

    _autosize_columns(ws)
    ws.freeze_panes = 'A4'


def _build_tab_nominas(wb, anio, mes, data):
    ws = wb.create_sheet('Nóminas')
    ws.sheet_properties.tabColor = COLOR_AMBAR

    _apply_header(ws, 1, 4, f'💰 Nóminas — {_periodo_label(anio, mes)}')

    # Totales
    row = 3
    _apply_table_header(ws, row, ['Concepto', 'Monto S/'])
    row += 1
    totales = [
        ('Total Bruto',            float(data['total_bruto'])),
        ('Total Descuentos',       float(data['total_descuentos'])),
        ('Total Neto pagado',      float(data['total_neto'])),
        ('Costo total empresa',    float(data['total_costo_empresa'])),
        ('Trabajadores en planilla', data['total_trabajadores']),
    ]
    for label, val in totales:
        _apply_table_row(
            ws, row, [label, val],
            number_formats=[None, '#,##0.00' if isinstance(val, float) else '#,##0'],
            bold=(label == 'Total Neto pagado'),
        )
        row += 1

    # Comparativo
    row += 1
    _apply_header(ws, row, 4, 'Comparativo vs Mes Anterior', fill=SUBHEADER_FILL,
                  font=SUBHEADER_FONT, height=22)
    row += 1
    _apply_table_header(ws, row, ['Concepto', 'Mes actual', 'Mes anterior', 'Var %'])
    row += 1
    var = data['var_pct']
    fill = None
    if var is not None:
        fill = OK_FILL if var <= 5 else WARN_FILL if var <= 15 else DANGER_FILL
    _apply_table_row(
        ws, row,
        ['Bruto', float(data['total_bruto']), float(data['bruto_anterior']),
         var if var is not None else '—'],
        number_formats=[None, '#,##0.00', '#,##0.00',
                        '+0.0"%";-0.0"%";0.0"%"' if var is not None else None],
        fill=fill,
    )
    row += 2

    # Top cargos
    _apply_header(ws, row, 4, 'Top 10 Cargos por Costo Promedio', fill=SUBHEADER_FILL,
                  font=SUBHEADER_FONT, height=22)
    row += 1
    _apply_table_header(ws, row, ['Cargo', 'Trabajadores', 'Neto promedio S/'])
    row += 1
    if data['top_cargos']:
        for cargo, c, prom in data['top_cargos']:
            _apply_table_row(
                ws, row, [cargo, c, float(prom)],
                number_formats=[None, '#,##0', '#,##0.00'],
            )
            row += 1
    else:
        _apply_table_row(ws, row, ['Sin datos', 0, 0])
        row += 1

    # Descuentos
    row += 2
    _apply_header(ws, row, 4, 'Top Descuentos por Concepto', fill=SUBHEADER_FILL,
                  font=SUBHEADER_FONT, height=22)
    row += 1
    _apply_table_header(ws, row, ['Concepto', 'Monto S/'])
    row += 1
    if data['descuentos']:
        for concepto, monto in data['descuentos']:
            _apply_table_row(ws, row, [concepto, float(monto)],
                             number_formats=[None, '#,##0.00'])
            row += 1
    else:
        _apply_table_row(ws, row, ['Sin datos', 0])
        row += 1

    _autosize_columns(ws)
    ws.freeze_panes = 'A4'


def _build_tab_capacitaciones(wb, anio, mes, data):
    ws = wb.create_sheet('Capacitaciones')
    ws.sheet_properties.tabColor = COLOR_VERDE

    _apply_header(ws, 1, 4, f'🎓 Capacitaciones — {_periodo_label(anio, mes)}')

    # Resumen
    row = 3
    _apply_table_header(ws, row, ['Indicador', 'Valor'])
    row += 1
    items = [
        ('Capacitaciones completadas (mes)', data['capacitaciones_mes']),
        ('Asistentes invitados',             data['asistentes_invitados']),
        ('Asistentes que asistieron (únicos)', data['asistentes_unicos']),
        ('Certificaciones vigentes',         data['certs_vigentes']),
        ('Certificaciones por vencer',       data['certs_por_vencer']),
        ('Certificaciones vencidas',         data['certs_vencidas']),
    ]
    for label, val in items:
        fill = None
        if 'vencer' in label.lower():
            fill = WARN_FILL
        elif 'vencidas' in label.lower():
            fill = DANGER_FILL if val else None
        elif 'vigentes' in label.lower():
            fill = OK_FILL if val else None
        _apply_table_row(ws, row, [label, val], number_formats=[None, '#,##0'], fill=fill)
        row += 1

    # Lista vencidas
    row += 2
    _apply_header(ws, row, 4,
                  'Certificaciones VENCIDAS — Acción Urgente (Top 20)',
                  fill=SUBHEADER_FILL, font=SUBHEADER_FONT, height=22)
    row += 1
    _apply_table_header(ws, row, ['Trabajador', 'Certificación', 'Días vencida'])
    row += 1
    if data['lista_vencidas']:
        for persona, cert, dias in data['lista_vencidas']:
            fill = DANGER_FILL if dias > 30 else WARN_FILL
            _apply_table_row(
                ws, row, [persona, cert, dias],
                number_formats=[None, None, '#,##0'],
                fill=fill,
            )
            row += 1
    else:
        _apply_table_row(ws, row, ['Sin certificaciones vencidas', '—', 0], fill=OK_FILL)
        row += 1

    _autosize_columns(ws)
    ws.freeze_panes = 'A4'


def _build_tab_alertas(wb, anio, mes, data):
    ws = wb.create_sheet('Alertas')
    ws.sheet_properties.tabColor = COLOR_ROJO

    _apply_header(ws, 1, 4, f'⚠️ Alertas y Acciones Urgentes — {_periodo_label(anio, mes)}')

    row = 3
    _apply_table_header(ws, row, ['Alerta', 'Cantidad', 'Severidad', 'Acción recomendada'])
    row += 1

    alertas = [
        ('Certificaciones vencidas',
         data['certs_vencidas'], 'Alta',
         'Renovar capacitaciones BPM/HACCP/SST inmediatamente.'),
        ('Papeletas pendientes > 7 días',
         data['papeletas_pendientes'], 'Media',
         'Revisar bandeja de aprobaciones de RRHH.'),
        ('Vacantes sin actividad > 30 días',
         data['vacantes_sin_actividad'], 'Media',
         'Reactivar fuentes o cerrar la vacante.'),
        ('Días sin briefing en últ. 7 días (gastronomía)',
         data['briefings_pendientes'], 'Media',
         'Jefes de local: publicar briefings diarios.'),
        ('Postulaciones estancadas > 14 días',
         data['postulaciones_estancadas'], 'Media',
         'Mover etapa o descartar candidato.'),
        ('Contratos por vencer próximos 30 días',
         data['contratos_por_vencer'], 'Alta',
         'Iniciar renovación o gestionar cese.'),
    ]

    for nombre, cant, sev, accion in alertas:
        fill = DANGER_FILL if (cant > 0 and sev == 'Alta') else (
            WARN_FILL if cant > 0 else OK_FILL
        )
        _apply_table_row(
            ws, row,
            [nombre, cant, sev, accion],
            number_formats=[None, '#,##0', None, None],
            fill=fill,
        )
        ws.row_dimensions[row].height = 30
        # Wrap accion
        ws.cell(row=row, column=4).alignment = Alignment(
            horizontal='left', vertical='center', wrap_text=True,
        )
        row += 1

    # Totales
    row += 1
    total = sum(v for _, v in [
        ('a', data['certs_vencidas']),
        ('b', data['papeletas_pendientes']),
        ('c', data['vacantes_sin_actividad']),
        ('d', data['briefings_pendientes']),
        ('e', data['postulaciones_estancadas']),
        ('f', data['contratos_por_vencer']),
    ])
    cell = ws.cell(row=row, column=1, value=f'TOTAL ALERTAS: {total}')
    cell.font = ALERT_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 55
    ws.freeze_panes = 'A4'


def _build_tab_detalle_empresa(wb, anio, mes, rows):
    ws = wb.create_sheet('Detalle por Empresa')
    ws.sheet_properties.tabColor = COLOR_GRIS

    _apply_header(ws, 1, 5, f'📋 Detalle por Empresa — {_periodo_label(anio, mes)}')

    row = 3
    _apply_table_header(
        ws, row,
        ['Empresa', 'Trabajadores activos', 'Papeletas (mes)',
         'Capacitaciones (mes)', 'Certs. vencidas'],
    )
    row += 1
    if rows:
        for nombre, trab, papel, capac, cert in rows:
            fill = DANGER_FILL if cert > 0 else None
            _apply_table_row(
                ws, row, [nombre, trab, papel, capac, cert],
                number_formats=[None, '#,##0', '#,##0', '#,##0', '#,##0'],
                fill=fill,
            )
            row += 1
        # Total general
        tot_trab = sum(r[1] for r in rows)
        tot_papel = sum(r[2] for r in rows)
        tot_capac = sum(r[3] for r in rows)
        tot_cert = sum(r[4] for r in rows)
        _apply_table_row(
            ws, row,
            ['TOTAL GENERAL', tot_trab, tot_papel, tot_capac, tot_cert],
            number_formats=[None, '#,##0', '#,##0', '#,##0', '#,##0'],
            bold=True,
            fill=SOFT_FILL,
        )
    else:
        _apply_table_row(ws, row, ['Sin empresas registradas', 0, 0, 0, 0])

    _autosize_columns(ws)
    ws.freeze_panes = 'A4'


# ══════════════════════════════════════════════════════════════════════
# Entry point
# ══════════════════════════════════════════════════════════════════════

def build_reporte_bi_mensual(anio: int, mes: int, empresa_id=None) -> bytes:
    """
    Genera Excel ejecutivo mensual con 8 tabs.

    Args:
        anio: año del periodo (ej. 2026)
        mes:  mes del periodo 1-12
        empresa_id: filtro opcional por empresa (None = todas)

    Returns:
        bytes del archivo .xlsx
    """
    wb = Workbook()
    # Remover hoja default
    default = wb.active
    wb.remove(default)

    # Datos
    data_resumen = _resumen_ejecutivo_data(anio, mes, empresa_id)
    data_head    = _headcount_data(anio, mes, empresa_id)
    data_reclu   = _reclutamiento_data(anio, mes, empresa_id)
    data_asis    = _asistencia_data(anio, mes, empresa_id)
    data_nom     = _nominas_data(anio, mes, empresa_id)
    data_capac   = _capacitaciones_data(anio, mes, empresa_id)
    data_alert   = _alertas_data(anio, mes, empresa_id)
    data_emp     = _detalle_empresas_data(anio, mes, empresa_id)

    # Tabs
    _build_tab_resumen(wb, anio, mes, data_resumen)
    _build_tab_headcount(wb, anio, mes, data_head)
    _build_tab_reclutamiento(wb, anio, mes, data_reclu)
    _build_tab_asistencia(wb, anio, mes, data_asis)
    _build_tab_nominas(wb, anio, mes, data_nom)
    _build_tab_capacitaciones(wb, anio, mes, data_capac)
    _build_tab_alertas(wb, anio, mes, data_alert)
    _build_tab_detalle_empresa(wb, anio, mes, data_emp)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
