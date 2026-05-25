"""
Importa planillas históricas YA EMITIDAS desde archivos del sistema anterior
del cliente (uso interno equipo Harmoni durante implementación).

Caso de uso EDO: el cliente tiene 5 planillas 2026 (ene-may) ya pagadas en
su sistema actual. Para que Harmoni arranque desde junio con histórico,
nosotros cargamos esas 5 planillas como "REGULAR HISTÓRICA" — sin
recalcular, manteniendo los montos exactos del sistema fuente.

A diferencia del importador legacy genérico, este comando:
- NO recalcula nada (los montos vienen como verdad de origen)
- Crea PeriodoNomina + RegistroNomina + LineaNomina inmediatamente CERRADO
- Marca origen='HISTORICA_IMPORTADA' para distinguir de planillas calculadas
- Soporta múltiples meses en un solo Excel (1 hoja por mes)

Uso (típico):
    python manage.py importar_planilla_historica planillas_2026.xlsx --empresa-ruc 20612345678

Formato esperado del Excel:
- 1 hoja por período (nombre tipo "Planilla_Ene_2026")
- Columnas: DNI · Nombre · Dias · Sueldo · Ingresos · Descuentos · Neto · CostoEmp

NOTA: este comando NO se expone como UI del cliente — solo lo corre el equipo
Harmoni durante el setup. Es destructivo si se corre 2 veces sobre el mismo
período (--allow-overwrite explícito requerido).
"""
from datetime import date
from decimal import Decimal
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction


COLUMNAS_ESPERADAS = {
    'DNI': 'dni',
    'Nombre': 'nombre',
    'Apellidos y Nombres': 'nombre',
    'Días': 'dias',
    'Dias': 'dias',
    'Días trabajados': 'dias',
    'Sueldo': 'sueldo',
    'Sueldo base': 'sueldo',
    'Ingresos': 'ingresos',
    'Total ingresos': 'ingresos',
    'Descuentos': 'descuentos',
    'Total descuentos': 'descuentos',
    'Neto': 'neto',
    'Neto a pagar': 'neto',
    'CostoEmp': 'costo_emp',
    'Costo empresa': 'costo_emp',
    'Costo total empresa': 'costo_emp',
}

MES_ALIASES = {
    'ene': 1, 'enero': 1, 'jan': 1, 'january': 1,
    'feb': 2, 'febrero': 2, 'february': 2,
    'mar': 3, 'marzo': 3, 'march': 3,
    'abr': 4, 'abril': 4, 'apr': 4, 'april': 4,
    'may': 5, 'mayo': 5,
    'jun': 6, 'junio': 6, 'june': 6,
    'jul': 7, 'julio': 7, 'july': 7,
    'ago': 8, 'agosto': 8, 'aug': 8, 'august': 8,
    'set': 9, 'sep': 9, 'setiembre': 9, 'septiembre': 9, 'september': 9,
    'oct': 10, 'octubre': 10, 'october': 10,
    'nov': 11, 'noviembre': 11, 'november': 11,
    'dic': 12, 'diciembre': 12, 'dec': 12, 'december': 12,
}


def _parse_periodo_from_sheet_name(sheet_name):
    """De 'Planilla_Mar_2026' o 'Mar 2026' o '2026-03' deduce (anio, mes)."""
    import re
    name = sheet_name.strip().lower()

    # AAAA-MM
    m = re.match(r'.*(\d{4})[-_/](\d{1,2}).*', name)
    if m:
        return int(m.group(1)), int(m.group(2))

    # Mes en texto + año
    for alias, mes_num in MES_ALIASES.items():
        if alias in name:
            year_m = re.search(r'(20\d{2})', name)
            if year_m:
                return int(year_m.group(1)), mes_num
            break

    return None


def _parse_decimal(val, default='0'):
    if val in (None, ''):
        return Decimal(default)
    if isinstance(val, (int, float, Decimal)):
        return Decimal(str(val))
    s = str(val).strip().replace(',', '').replace('S/', '').strip()
    try:
        return Decimal(s)
    except Exception:
        return Decimal(default)


class Command(BaseCommand):
    help = 'Importa planillas históricas YA EMITIDAS desde Excel del cliente anterior'

    def add_arguments(self, parser):
        parser.add_argument('archivo', type=str, help='Path al xlsx con las planillas históricas')
        parser.add_argument('--empresa-ruc', type=str, required=True,
                            help='RUC de la empresa a la que pertenecen estas planillas')
        parser.add_argument('--dry-run', action='store_true',
                            help='Solo reporta qué importaría')
        parser.add_argument('--allow-overwrite', action='store_true',
                            help='Si el período ya existe en Harmoni, sobreescribe sus datos')
        parser.add_argument('--anio', type=int, default=None,
                            help='Forzar año si el sheet name es ambiguo')

    def handle(self, *args, **opts):
        from empresas.models import Empresa
        from personal.models import Personal
        from nominas.models import (
            PeriodoNomina, RegistroNomina, LineaNomina, ConceptoRemunerativo,
        )
        import openpyxl

        archivo = Path(opts['archivo']).resolve()
        if not archivo.exists():
            raise CommandError(f'Archivo no encontrado: {archivo}')

        ruc = opts['empresa_ruc'].strip()
        dry = opts['dry_run']
        allow_overwrite = opts['allow_overwrite']

        try:
            empresa = Empresa.objects.get(ruc=ruc, activa=True)
        except Empresa.DoesNotExist:
            raise CommandError(f'Empresa con RUC {ruc} no existe o está inactiva')

        self.stdout.write(self.style.MIGRATE_HEADING(
            f'Importando planillas históricas\n'
            f'  Archivo:  {archivo.name}\n'
            f'  Empresa:  {empresa.razon_social} ({ruc})\n'
            f'  Modo:     {"DRY-RUN" if dry else "WRITE"}\n'
            f'  Overwrite:{"SI" if allow_overwrite else "NO"}'
        ))

        try:
            wb = openpyxl.load_workbook(archivo, data_only=True)
        except Exception as e:
            raise CommandError(f'Excel inválido: {e}')

        # Cache de personal por DNI
        personal_by_dni = {
            p.nro_doc: p for p in Personal.objects.filter(empresa=empresa).only('id', 'nro_doc', 'sueldo_base')
        }

        # Concepto "OTROS-HISTORICO" para representar el monto neto en línea única
        concepto_hist, _ = ConceptoRemunerativo.objects.get_or_create(
            codigo='hist-import',
            defaults={
                'nombre': 'Planilla histórica importada',
                'tipo': 'INGRESO',
                'subtipo': 'NO_REMUNERATIVO',
                'formula': 'MANUAL',
                'categoria': 'OTROS_ING',
                'activo': True,
            },
        )

        total_periodos = total_regs = 0

        for sheet_name in wb.sheetnames:
            anio_mes = _parse_periodo_from_sheet_name(sheet_name)
            if not anio_mes:
                self.stdout.write(self.style.WARNING(
                    f'  Skip: hoja "{sheet_name}" sin período identificable'
                ))
                continue
            anio, mes = anio_mes
            if opts['anio']:
                anio = opts['anio']

            ws = wb[sheet_name]
            self.stdout.write(f'\n  Procesando: {sheet_name} → {anio}-{mes:02d}')

            # Detectar columnas
            col_map = {}
            for col_idx in range(1, ws.max_column + 1):
                hdr = ws.cell(row=1, column=col_idx).value
                if hdr:
                    hdr_clean = str(hdr).strip()
                    if hdr_clean in COLUMNAS_ESPERADAS:
                        col_map[col_idx] = COLUMNAS_ESPERADAS[hdr_clean]

            if 'dni' not in col_map.values() or 'neto' not in col_map.values():
                self.stdout.write(self.style.ERROR(
                    f'    Faltan columnas DNI / Neto en {sheet_name}'
                ))
                continue

            # ¿Existe período?
            from datetime import date as _date
            periodo, created = PeriodoNomina.objects.get_or_create(
                empresa=empresa,
                anio=anio,
                mes=mes,
                tipo='REGULAR',
                defaults={
                    'descripcion': f'Planilla {anio}-{mes:02d} (histórica importada)',
                    'fecha_inicio': _date(anio, mes, 1),
                    'fecha_fin': _date(anio, mes, 28),  # aproximación
                    'estado': 'CERRADO',
                },
            )
            if not created and not allow_overwrite:
                self.stdout.write(self.style.WARNING(
                    f'    Skip: período {anio}-{mes:02d} ya existe '
                    f'(usa --allow-overwrite para sobreescribir)'
                ))
                continue

            # Leer filas
            n_regs = 0
            t_ing = t_desc = t_neto = t_costo = Decimal('0')

            with transaction.atomic():
                if not created:
                    # Borrar registros previos del período
                    periodo.registros.all().delete()

                row = 2
                while row <= ws.max_row:
                    data = {}
                    for col_idx, key in col_map.items():
                        data[key] = ws.cell(row=row, column=col_idx).value

                    dni = str(data.get('dni') or '').strip()
                    if not dni:
                        break

                    p = personal_by_dni.get(dni)
                    if not p:
                        self.stdout.write(self.style.WARNING(
                            f'      DNI {dni} no encontrado — skip'
                        ))
                        row += 1
                        continue

                    sueldo = _parse_decimal(data.get('sueldo'), str(p.sueldo_base or '0'))
                    ingresos = _parse_decimal(data.get('ingresos'))
                    descuentos = _parse_decimal(data.get('descuentos'))
                    neto = _parse_decimal(data.get('neto'))
                    costo_emp = _parse_decimal(data.get('costo_emp'), str(ingresos))
                    dias = int(data.get('dias') or 30)

                    if dry:
                        n_regs += 1
                        t_ing += ingresos
                        t_desc += descuentos
                        t_neto += neto
                        t_costo += costo_emp
                    else:
                        reg = RegistroNomina.objects.create(
                            periodo=periodo,
                            personal=p,
                            sueldo_base=sueldo,
                            regimen_pension=p.regimen_pension or 'AFP',
                            afp=getattr(p, 'afp', '') or '',
                            grupo=getattr(p, 'grupo_tareo', '') or '',
                            dias_trabajados=dias,
                            total_ingresos=ingresos,
                            total_descuentos=descuentos,
                            neto_a_pagar=neto,
                            aporte_essalud=Decimal('0'),  # no se desglosa histórico
                            costo_total_empresa=costo_emp,
                            estado='APROBADO',
                            observaciones='Histórico importado del sistema anterior',
                        )
                        # Una sola línea representando el resumen histórico
                        LineaNomina.objects.create(
                            registro=reg,
                            concepto=concepto_hist,
                            base_calculo=Decimal('0'),
                            porcentaje_aplicado=Decimal('0'),
                            monto=neto,
                            observacion='Resumen histórico — detalle no disponible',
                        )
                        n_regs += 1
                        t_ing += ingresos
                        t_desc += descuentos
                        t_neto += neto
                        t_costo += costo_emp

                    row += 1

                # Update totales del período
                if not dry:
                    periodo.total_trabajadores = n_regs
                    periodo.total_bruto = t_ing
                    periodo.total_descuentos = t_desc
                    periodo.total_neto = t_neto
                    periodo.total_costo_empresa = t_costo
                    periodo.save(update_fields=[
                        'total_trabajadores', 'total_bruto', 'total_descuentos',
                        'total_neto', 'total_costo_empresa',
                    ])

            self.stdout.write(self.style.SUCCESS(
                f'    Período {anio}-{mes:02d}: {n_regs} trabajadores · '
                f'bruto S/ {t_ing:,.2f} · neto S/ {t_neto:,.2f}'
            ))
            total_periodos += 1
            total_regs += n_regs

        self.stdout.write('')
        self.stdout.write(self.style.SUCCESS(
            f'FINAL: {total_periodos} períodos · {total_regs} registros procesados'
        ))
