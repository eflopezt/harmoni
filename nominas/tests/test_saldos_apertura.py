"""
Tests para Saldos de Apertura (modelo + wizard view).

Cubre:
- Modelo: OneToOne con Personal, defaults, str
- Wizard: render empty, render con saldos persistidos
- Plantilla Excel: descarga, headers, pre-llenado con trabajadores activos
"""
from datetime import date
from decimal import Decimal

import pytest
from django.contrib.auth.models import User
from django.urls import reverse

from personal.models import Area, Personal, SubArea
from nominas.models import SaldoAperturaTrabajador, ConfiguracionApertura


# ─── Fixtures ─────────────────────────────────────────────────────────────────

@pytest.fixture
def admin_user(db):
    return User.objects.create_superuser(
        username="admin_test", password="x", email="a@a.com",
    )


@pytest.fixture
def client_logged(client, admin_user):
    client.force_login(admin_user)
    return client


@pytest.fixture
def area(db):
    return Area.objects.create(nombre="Salón")


@pytest.fixture
def subarea(area):
    return SubArea.objects.create(nombre="Servicio", area=area)


@pytest.fixture
def worker(db, subarea):
    return Personal.objects.create(
        nro_doc="71234567",
        apellidos_nombres="QUISPE LOPEZ, MARIA",
        cargo="Mesera Senior",
        tipo_trab="Empleado",
        estado="Activo",
        subarea=subarea,
        fecha_alta=date(2024, 1, 15),
        sueldo_base=Decimal("1800.00"),
    )


# ─── Modelo ──────────────────────────────────────────────────────────────────

class TestModelo:
    def test_crear_saldo_apertura(self, worker):
        s = SaldoAperturaTrabajador.objects.create(
            personal=worker,
            fecha_corte=date(2026, 5, 31),
            prov_cts=Decimal("750.00"),
            prov_gratificacion=Decimal("1200.00"),
            prov_vacaciones=Decimal("600.00"),
            dias_vacaciones_pendientes=10,
            ir5_acumulado=Decimal("0.00"),
            prestamo_saldo=Decimal("0.00"),
        )
        assert s.personal == worker
        assert s.prov_cts == Decimal("750.00")
        assert s.dias_vacaciones_pendientes == 10
        assert "QUISPE LOPEZ" in str(s)

    def test_unique_personal_one_to_one(self, worker):
        SaldoAperturaTrabajador.objects.create(
            personal=worker, fecha_corte=date(2026, 5, 31),
        )
        with pytest.raises(Exception):
            SaldoAperturaTrabajador.objects.create(
                personal=worker, fecha_corte=date(2026, 6, 30),
            )

    def test_defaults_son_cero(self, worker):
        s = SaldoAperturaTrabajador.objects.create(
            personal=worker, fecha_corte=date(2026, 5, 31),
        )
        assert s.prov_cts == Decimal("0")
        assert s.prov_gratificacion == Decimal("0")
        assert s.prov_vacaciones == Decimal("0")
        assert s.dias_vacaciones_pendientes == 0
        assert s.ir5_acumulado == Decimal("0")
        assert s.prestamo_saldo == Decimal("0")


class TestConfiguracionApertura:
    def test_singleton_global_empresa_null(self, admin_user):
        cfg = ConfiguracionApertura.objects.create(
            empresa=None,
            fecha_corte=date(2026, 5, 31),
            completado=True,
            total_trabajadores=70,
            confirmado_por=admin_user,
        )
        assert cfg.empresa is None
        assert cfg.completado
        assert "Global" in str(cfg)


# ─── Wizard view ──────────────────────────────────────────────────────────────

class TestWizardView:
    def test_render_sin_apertura(self, client_logged):
        resp = client_logged.get(reverse("nominas_saldos_apertura"))
        assert resp.status_code == 200
        body = resp.content.decode()
        assert "Onboarding Express" in body
        # No debe mostrar "Onboarding Completado" si no está completada
        assert "Onboarding Completado" not in body

    def test_render_con_apertura_completada(self, client_logged, admin_user):
        ConfiguracionApertura.objects.create(
            empresa=None,
            fecha_corte=date(2026, 5, 31),
            completado=True,
            total_trabajadores=70,
            confirmado_por=admin_user,
        )
        resp = client_logged.get(reverse("nominas_saldos_apertura"))
        body = resp.content.decode()
        assert "Onboarding Completado" in body
        assert "2026-05-31" in body
        assert "70" in body

    def test_plantilla_excel_descarga(self, client_logged, worker):
        resp = client_logged.get(
            reverse("nominas_saldos_apertura_plantilla") + "?fecha_corte=2026-05-31"
        )
        assert resp.status_code == 200
        assert "spreadsheet" in resp["Content-Type"]
        assert "Saldos_Apertura" in resp["Content-Disposition"]
        # Contiene bytes Excel (zip magic)
        assert resp.content[:2] == b"PK"

    def test_plantilla_excel_incluye_workers(self, client_logged, worker):
        import openpyxl
        from io import BytesIO

        resp = client_logged.get(
            reverse("nominas_saldos_apertura_plantilla")
        )
        wb = openpyxl.load_workbook(BytesIO(resp.content))
        ws = wb.active
        # Header en fila 1
        headers = [c.value for c in ws[1]]
        assert "DNI" in headers
        assert any("CTS" in (h or "") for h in headers)
        # Worker en fila 2
        assert ws.cell(row=2, column=1).value == "71234567"
        assert "QUISPE LOPEZ" in ws.cell(row=2, column=2).value


# ─── Permisos ─────────────────────────────────────────────────────────────────

class TestPermisos:
    def test_anonimo_redirige(self, client):
        resp = client.get(reverse("nominas_saldos_apertura"))
        assert resp.status_code == 302
        assert "/login" in resp["Location"]

    def test_worker_no_pasa(self, db, client):
        u = User.objects.create_user(username="w", password="x", is_staff=False)
        client.force_login(u)
        resp = client.get(reverse("nominas_saldos_apertura"))
        # Decorador user_passes_test → redirige a login (302) o 403
        assert resp.status_code in (302, 403)
