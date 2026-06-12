"""
Tests del round-trip Excel de conceptos masivos + datos del registro.

Cubre:
- Export: plantilla incluye columnas `__campo` (días/HE/otros) con valores
  actuales del registro, además de los conceptos manuales.
- Import: columnas `__campo` actualizan RegistroNomina (HE, días, otros
  ingresos/descuentos) y los conceptos van a conceptos_manuales; la planilla
  se recalcula (HE valorizadas con sueldo/30/8).
- Celda vacía en columna `__campo` = no modificar; valores inválidos
  (negativos, días > 31) se ignoran.
- Período APROBADO/CERRADO rechaza el import.
"""
import io
from datetime import date
from decimal import Decimal

import openpyxl
import pytest
from django.contrib.auth.models import User
from django.urls import reverse

from nominas.models import ConceptoRemunerativo, PeriodoNomina, RegistroNomina
from personal.models import Area, Personal, SubArea


# ─── Fixtures ─────────────────────────────────────────────────────────────────

@pytest.fixture
def subarea(db):
    area = Area.objects.create(nombre="Operaciones")
    return SubArea.objects.create(nombre="Campo", area=area)


@pytest.fixture
def personal(subarea):
    return Personal.objects.create(
        nro_doc="71000010",
        apellidos_nombres="MASIVO PRUEBA, JUAN",
        cargo="Operario",
        tipo_trab="Empleado",
        estado="Activo",
        subarea=subarea,
        fecha_alta=date(2024, 1, 1),
        sueldo_base=Decimal("3000.00"),
    )


@pytest.fixture
def periodo(db):
    return PeriodoNomina.objects.create(
        tipo="REGULAR", anio=2026, mes=5,
        fecha_inicio=date(2026, 4, 22), fecha_fin=date(2026, 5, 21),
        estado="BORRADOR",
    )


@pytest.fixture
def registro(periodo, personal):
    return RegistroNomina.objects.create(
        periodo=periodo, personal=personal,
        sueldo_base=Decimal("3000.00"),
        regimen_pension="SIN_PENSION",
        dias_trabajados=30,
        horas_extra_25=Decimal("4.00"),
    )


@pytest.fixture
def concepto_bono(db):
    return ConceptoRemunerativo.objects.create(
        codigo="bono-prod-test",
        nombre="Bono Productividad Test",
        tipo="INGRESO",
        categoria="BONIFICACION",
        formula="MANUAL",
        activo=True,
    )


@pytest.fixture
def admin_client(db, client):
    User.objects.create_superuser("admin", "a@a.com", "x")
    client.login(username="admin", password="x")
    return client


def _build_xlsx(codigos, filas):
    """Arma un xlsx en memoria con la estructura de la plantilla.

    codigos: lista de códigos para la fila 2 (desde la columna 4).
    filas: lista de tuplas (dni, nombre, grupo, *valores).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="DNI")
    ws.cell(row=1, column=2, value="Apellidos y Nombres")
    ws.cell(row=1, column=3, value="Grupo")
    for idx, cod in enumerate(codigos):
        ws.cell(row=1, column=4 + idx, value=cod)
        ws.cell(row=2, column=4 + idx, value=cod)
    for row_idx, fila in enumerate(filas, 4):
        for col_idx, val in enumerate(fila, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    buf.name = "plantilla.xlsx"
    return buf


# ─── Export ───────────────────────────────────────────────────────────────────

class TestExportPlantilla:
    def test_incluye_columnas_registro_y_conceptos(
        self, admin_client, periodo, registro, concepto_bono
    ):
        url = reverse("nominas_periodo_conceptos_exportar", args=[periodo.pk])
        resp = admin_client.get(url)
        assert resp.status_code == 200

        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        codigos = []
        col = 4
        while ws.cell(row=2, column=col).value:
            codigos.append(ws.cell(row=2, column=col).value)
            col += 1

        assert "__he_25" in codigos
        assert "__dias_trabajados" in codigos
        assert "__otros_ingresos" in codigos
        assert "bono-prod-test" in codigos
        # Las columnas __ van antes que los conceptos
        assert codigos.index("__he_25") < codigos.index("bono-prod-test")

        # Pre-llenado con valores actuales del registro
        col_he25 = 4 + codigos.index("__he_25")
        assert ws.cell(row=4, column=col_he25).value == pytest.approx(4.0)
        col_dias = 4 + codigos.index("__dias_trabajados")
        assert ws.cell(row=4, column=col_dias).value == 30

    def test_export_sin_conceptos_manuales_funciona(self, admin_client, periodo, registro):
        # Sin conceptos manuales configurados, la plantilla igual sirve
        # para días/HE/otros.
        url = reverse("nominas_periodo_conceptos_exportar", args=[periodo.pk])
        resp = admin_client.get(url)
        assert resp.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        assert wb.active.cell(row=2, column=4).value == "__dias_trabajados"


# ─── Import ───────────────────────────────────────────────────────────────────

class TestImportRoundTrip:
    def test_actualiza_registro_y_conceptos_y_recalcula(
        self, admin_client, periodo, registro, concepto_bono
    ):
        codigos = ["__he_25", "__dias_falta", "__otros_ingresos", "bono-prod-test"]
        archivo = _build_xlsx(codigos, [
            ("71000010", "MASIVO PRUEBA, JUAN", "", 10, 2, 150, 200),
        ])
        url = reverse("nominas_periodo_conceptos_importar", args=[periodo.pk])
        resp = admin_client.post(url, {"archivo": archivo})
        assert resp.status_code == 302

        registro.refresh_from_db()
        assert registro.horas_extra_25 == Decimal("10.00")
        assert registro.dias_falta == 2
        assert registro.otros_ingresos == Decimal("150.00")
        assert registro.conceptos_manuales == {"bono-prod-test": "200"}

        # Recalculado: valor hora = 3000/30/8 = 12.50;
        # HE25 = 10 × 12.50 × 1.25 = 156.25; bruto = 3000 + 156.25 + 150 + 200
        assert registro.total_ingresos == Decimal("3506.25")
        assert registro.estado == "CALCULADO"

        periodo.refresh_from_db()
        assert periodo.total_bruto == Decimal("3506.25")

    def test_celda_vacia_no_modifica_campo(self, admin_client, periodo, registro):
        codigos = ["__he_25", "__dias_trabajados"]
        archivo = _build_xlsx(codigos, [
            ("71000010", "MASIVO PRUEBA, JUAN", "", None, 28),
        ])
        url = reverse("nominas_periodo_conceptos_importar", args=[periodo.pk])
        admin_client.post(url, {"archivo": archivo})

        registro.refresh_from_db()
        assert registro.horas_extra_25 == Decimal("4.00")  # sin cambio
        assert registro.dias_trabajados == 28

    def test_valores_invalidos_se_ignoran(self, admin_client, periodo, registro):
        codigos = ["__dias_trabajados", "__he_35", "__otros_descuentos"]
        archivo = _build_xlsx(codigos, [
            ("71000010", "MASIVO PRUEBA, JUAN", "", 45, -3, "abc"),
        ])
        url = reverse("nominas_periodo_conceptos_importar", args=[periodo.pk])
        admin_client.post(url, {"archivo": archivo})

        registro.refresh_from_db()
        assert registro.dias_trabajados == 30      # 45 > 31 → ignorado
        assert registro.horas_extra_35 == Decimal("0")  # negativo → ignorado
        assert registro.otros_descuentos == Decimal("0")  # no numérico → ignorado

    def test_periodo_cerrado_rechaza_import(self, admin_client, periodo, registro):
        periodo.estado = "CERRADO"
        periodo.save(update_fields=["estado"])
        archivo = _build_xlsx(["__he_25"], [("71000010", "X", "", 10)])
        url = reverse("nominas_periodo_conceptos_importar", args=[periodo.pk])
        admin_client.post(url, {"archivo": archivo})

        registro.refresh_from_db()
        assert registro.horas_extra_25 == Decimal("4.00")  # intacto
