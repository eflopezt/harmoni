"""
Revisión pre-aprobación del período (variance review).

Patrón estándar de payroll moderno (Deel/Rippling/Gusto): antes de aprobar,
el admin ve UNA pantalla con:
  - el monto que va a salir de la cuenta (cifra hero),
  - la variación por concepto vs el mes anterior (Δ y Δ%),
  - flags por trabajador (del detector de anomalías) que puede descartar
    explícitamente — advertencia ≠ bloqueo, pero queda registro de quién
    descartó qué (PeriodoNomina.flags_descartados),
  - el resumen en lenguaje humano (qué cobran, qué se retiene, qué paga
    la empresa) y los parámetros legales congelados del período.

URLs:
  /nominas/periodos/<pk>/revision/                         (GET)
  /nominas/periodos/<pk>/revision/inconsistencias.xlsx      (GET)
  /nominas/periodos/<pk>/revision/descartar/                (POST flag_key)
"""
from decimal import Decimal

from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Avg, Count, Max, Min, Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_POST

from .models import LineaNomina, PeriodoNomina
from .views_anomalias import detectar_anomalias_periodo

solo_admin = user_passes_test(lambda u: u.is_superuser)

# Umbral de resaltado en la tabla de variación por concepto
PCT_VARIACION_CONCEPTO = Decimal('15.00')

TIPO_ALERTA_LABELS = {
    'NUEVO': 'Trabajador nuevo en planilla',
    'FALTANTE': 'Trabajador faltante vs mes anterior',
    'NETO_CAMBIO': 'Cambio significativo de neto',
    'DESCUENTOS_ATIPICOS': 'Descuentos atípicos',
    'SUELDO_BASE_CAMBIO': 'Cambio de sueldo base',
}

ACCIONES_SUGERIDAS = {
    'NUEVO': (
        'Confirmar alta, contrato, tareo y conceptos manuales. '
        'Si corresponde al mes, descartar con motivo.'
    ),
    'FALTANTE': (
        'Validar si hubo cese/liquidación o si el trabajador quedó fuera del cálculo. '
        'Corregir personal/contrato y recalcular si aplica.'
    ),
    'NETO_CAMBIO': (
        'Abrir el registro, comparar conceptos, asistencia, días y descuentos contra '
        'el mes anterior. Si el cambio es correcto, descartar con sustento.'
    ),
    'DESCUENTOS_ATIPICOS': (
        'Revisar AFP/ONP, IR 5ta, préstamos, descuentos judiciales y otros descuentos '
        'manuales antes de aprobar.'
    ),
    'SUELDO_BASE_CAMBIO': (
        'Validar que exista contrato, adenda o historial salarial que sustente el cambio.'
    ),
}


def _totales_por_concepto(periodo):
    """Suma de LineaNomina por concepto del período → dict codigo → row."""
    qs = (
        LineaNomina.objects
        .filter(registro__periodo=periodo)
        .values('concepto__codigo', 'concepto__nombre', 'concepto__tipo')
        .annotate(total=Sum('monto'), n=Count('registro_id', distinct=True))
    )
    return {
        r['concepto__codigo']: {
            'codigo': r['concepto__codigo'],
            'nombre': r['concepto__nombre'],
            'tipo': r['concepto__tipo'],
            'total': r['total'] or Decimal('0'),
            'n': r['n'],
        }
        for r in qs
    }


def variance_conceptos(periodo, anterior):
    """
    Tabla de variación por concepto: actual vs anterior, Δ y Δ%.
    Incluye conceptos que aparecen o desaparecen entre meses.
    """
    act = _totales_por_concepto(periodo)
    ant = _totales_por_concepto(anterior) if anterior else {}

    rows = []
    for codigo in set(act) | set(ant):
        a = act.get(codigo)
        b = ant.get(codigo)
        total_act = a['total'] if a else Decimal('0')
        total_ant = b['total'] if b else Decimal('0')
        delta = total_act - total_ant
        if total_ant:
            pct = delta / total_ant * Decimal('100')
        else:
            pct = Decimal('100') if total_act else Decimal('0')
        base = a or b
        resaltar = bool(anterior) and (
            abs(pct) >= PCT_VARIACION_CONCEPTO or a is None or b is None
        )
        rows.append({
            'codigo': codigo,
            'nombre': base['nombre'],
            'tipo': base['tipo'],
            'total_act': total_act,
            'total_ant': total_ant,
            'n_act': a['n'] if a else 0,
            'n_ant': b['n'] if b else 0,
            'delta': delta,
            'pct': pct,
            'nuevo': b is None,
            'desaparecio': a is None,
            'resaltar': resaltar,
        })

    rows.sort(key=lambda r: (0 if r['resaltar'] else 1, -abs(r['delta'])))
    return rows


def _flag_key(anomalia):
    reg = anomalia.get('registro')
    pid = getattr(reg, 'personal_id', None) or 'NA'
    return f"{anomalia['tipo']}:{pid}"


def _registros_por_personal(periodo):
    if not periodo:
        return {}
    return {
        r.personal_id: r
        for r in periodo.registros.select_related(
            'personal',
            'personal__subarea',
            'personal__subarea__area',
        )
    }


def _periodo_label(periodo):
    if not periodo:
        return ''
    return f"{periodo.mes:02d}/{periodo.anio}"


def _decimal_to_float(value):
    if value in (None, ''):
        return None
    return float(value)


def _pct(antes, despues):
    antes = antes or Decimal('0')
    despues = despues or Decimal('0')
    if not antes:
        return Decimal('100') if despues else Decimal('0')
    return (despues - antes) / antes * Decimal('100')


def _format_descartado_en(value):
    if not value:
        return ''
    return str(value).replace('T', ' ')[:16]


def _flag_origen_path(flag):
    actual = flag.get('registro_actual')
    if actual:
        return reverse('nominas_registro_detalle', args=[actual.pk])

    registro = flag.get('registro')
    personal_id = getattr(registro, 'personal_id', None)
    if personal_id:
        return reverse('personal_detail', args=[personal_id])
    return ''


def _enriquecer_flag(anomalia, periodo, anterior, actuales, anteriores, request=None):
    registro = anomalia.get('registro')
    personal_id = getattr(registro, 'personal_id', None)
    actual = actuales.get(personal_id)
    previo = anteriores.get(personal_id)
    persona = getattr(actual or previo or registro, 'personal', None)

    flag = dict(anomalia)
    flag['tipo_label'] = TIPO_ALERTA_LABELS.get(flag.get('tipo'), flag.get('tipo', 'Alerta'))
    flag['accion_sugerida'] = ACCIONES_SUGERIDAS.get(
        flag.get('tipo'),
        'Revisar el origen de la alerta y decidir si se corrige o se descarta con sustento.',
    )
    flag['registro_actual'] = actual
    flag['registro_anterior'] = previo
    flag['personal_obj'] = persona
    flag['periodo_actual_label'] = _periodo_label(periodo)
    flag['periodo_anterior_label'] = _periodo_label(anterior)

    origen_path = _flag_origen_path(flag)
    flag['origen_url'] = origen_path
    flag['origen_excel_url'] = request.build_absolute_uri(origen_path) if request and origen_path else origen_path

    if persona and getattr(persona, 'subarea', None):
        flag['subarea_nombre'] = persona.subarea.nombre
        flag['area_nombre'] = persona.subarea.area.nombre if persona.subarea.area else ''
    else:
        flag['subarea_nombre'] = ''
        flag['area_nombre'] = ''

    neto_actual = getattr(actual, 'neto_a_pagar', None)
    neto_anterior = getattr(previo, 'neto_a_pagar', None)
    sueldo_actual = getattr(actual, 'sueldo_base', None)
    sueldo_anterior = getattr(previo, 'sueldo_base', None)

    flag['delta_neto'] = (
        neto_actual - neto_anterior
        if neto_actual is not None and neto_anterior is not None
        else None
    )
    flag['pct_neto'] = (
        _pct(neto_anterior, neto_actual)
        if neto_actual is not None and neto_anterior is not None
        else None
    )
    flag['delta_sueldo'] = (
        sueldo_actual - sueldo_anterior
        if sueldo_actual is not None and sueldo_anterior is not None
        else None
    )
    return flag


def _flags_revision(periodo, report, request=None):
    descartados = periodo.flags_descartados or {}
    actuales = _registros_por_personal(periodo)
    anteriores = _registros_por_personal(report['anterior'])

    flags = []
    pendientes = 0
    for anomalia in report['anomalias']:
        key = _flag_key(anomalia)
        desc = descartados.get(key)
        if not desc:
            pendientes += 1
        flag = _enriquecer_flag(anomalia, periodo, report['anterior'], actuales, anteriores, request)
        flag.update({'key': key, 'descartado': desc})
        flags.append(flag)
    return flags, pendientes


@login_required
@solo_admin
def periodo_revision(request, pk):
    """Pantalla de revisión previa a aprobar el período."""
    periodo = get_object_or_404(PeriodoNomina, pk=pk)
    report = detectar_anomalias_periodo(periodo)
    anterior = report['anterior']
    flags, pendientes = _flags_revision(periodo, report, request)

    stats = periodo.registros.aggregate(
        n=Count('id'),
        neto_prom=Avg('neto_a_pagar'),
        neto_min=Min('neto_a_pagar'),
        neto_max=Max('neto_a_pagar'),
        essalud=Sum('aporte_essalud'),
    )
    retenciones = (periodo.total_bruto or 0) - (periodo.total_neto or 0)

    return render(request, 'nominas/periodo_revision.html', {
        'periodo': periodo,
        'anterior': anterior,
        'variance': variance_conceptos(periodo, anterior),
        'flags': flags,
        'flags_pendientes': pendientes,
        'stats': stats,
        'retenciones': retenciones,
        'parametros': periodo.parametros_snapshot or {},
        'puede_aprobar': periodo.estado == 'CALCULADO',
        'delta_neto': (periodo.total_neto or 0) - (anterior.total_neto or 0) if anterior else None,
    })


@login_required
@solo_admin
def periodo_revision_inconsistencias_excel(request, pk):
    """Exporta todas las inconsistencias de revisión con detalle auditable."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    periodo = get_object_or_404(PeriodoNomina, pk=pk)
    report = detectar_anomalias_periodo(periodo)
    anterior = report['anterior']
    flags, pendientes = _flags_revision(periodo, report, request)
    variance = variance_conceptos(periodo, anterior)

    header_fill = PatternFill("solid", fgColor="0D2B27")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    title_fill = PatternFill("solid", fgColor="CCFBF1")
    pending_fill = PatternFill("solid", fgColor="FEF3C7")
    critical_fill = PatternFill("solid", fgColor="FEE2E2")
    discarded_fill = PatternFill("solid", fgColor="E2E8F0")
    thin = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1'),
    )

    def header_row(ws, row_num, headers):
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin
        ws.row_dimensions[row_num].height = 30

    def data_cell(ws, row_num, col_num, value, fill=None, number_format=None, align='left'):
        cell = ws.cell(row=row_num, column=col_num, value=value)
        if fill:
            cell.fill = fill
        cell.border = thin
        cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
        if number_format:
            cell.number_format = number_format
        return cell

    def auto_width(ws, max_width=58):
        for col in ws.columns:
            max_len = max((len(str(cell.value or '')) for cell in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, max_width)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Inconsistencias'

    ws['A1'] = f'Revisión de inconsistencias - {periodo}'
    ws['A1'].font = Font(size=13, bold=True, color='0D2B27')
    ws['A1'].fill = title_fill
    ws.merge_cells('A1:AA1')
    ws['A2'] = (
        f'Período actual: {_periodo_label(periodo)} | '
        f'Comparado con: {_periodo_label(anterior) or "sin período anterior"} | '
        f'Flags pendientes: {pendientes} | Generado: {timezone.localtime():%d/%m/%Y %H:%M}'
    )
    ws.merge_cells('A2:AA2')
    ws['A2'].font = Font(size=9, color='475569')

    headers = [
        'N°', 'Estado revisión', 'Severidad', 'Tipo alerta', 'Trabajador', 'DNI',
        'Área', 'SubÁrea', 'Grupo', 'Período actual', 'Período comparación',
        'Detalle encontrado', 'Neto actual', 'Neto anterior', 'Dif. neto S/',
        'Var. neto %', 'Sueldo actual', 'Sueldo anterior', 'Dif. sueldo S/',
        'Ingresos actual', 'Descuentos actual', 'Costo empresa actual',
        'Descartado por', 'Descartado en', 'Motivo descarte',
        'Acción sugerida', 'Origen',
    ]
    header_row(ws, 4, headers)
    ws.freeze_panes = 'A5'
    ws.auto_filter.ref = f'A4:AA{max(5, len(flags) + 4)}'

    if not flags:
        ws.cell(row=5, column=1, value='Sin inconsistencias detectadas en la revisión automática.')
        ws.merge_cells('A5:AA5')
        ws['A5'].font = Font(color='15803D', bold=True)
    else:
        for idx, flag in enumerate(flags, 1):
            row = idx + 4
            actual = flag.get('registro_actual')
            previo = flag.get('registro_anterior')
            persona = flag.get('personal_obj')
            desc = flag.get('descartado') or {}
            estado = 'Descartado' if desc else 'Pendiente'
            fill = discarded_fill if desc else (critical_fill if flag.get('severidad') == 'critico' else pending_fill)
            values = [
                idx,
                estado,
                flag.get('severidad', ''),
                flag.get('tipo_label', flag.get('tipo', '')),
                getattr(persona, 'apellidos_nombres', '') or '—',
                getattr(persona, 'nro_doc', '') or '—',
                flag.get('area_nombre', ''),
                flag.get('subarea_nombre', ''),
                getattr(actual or previo, 'grupo', '') or getattr(persona, 'grupo_tareo', '') or '',
                flag.get('periodo_actual_label', ''),
                flag.get('periodo_anterior_label', ''),
                flag.get('mensaje', ''),
                _decimal_to_float(getattr(actual, 'neto_a_pagar', None)),
                _decimal_to_float(getattr(previo, 'neto_a_pagar', None)),
                _decimal_to_float(flag.get('delta_neto')),
                _decimal_to_float(flag.get('pct_neto')),
                _decimal_to_float(getattr(actual, 'sueldo_base', None)),
                _decimal_to_float(getattr(previo, 'sueldo_base', None)),
                _decimal_to_float(flag.get('delta_sueldo')),
                _decimal_to_float(getattr(actual, 'total_ingresos', None)),
                _decimal_to_float(getattr(actual, 'total_descuentos', None)),
                _decimal_to_float(getattr(actual, 'costo_total_empresa', None)),
                desc.get('por', ''),
                _format_descartado_en(desc.get('en')),
                desc.get('mensaje', ''),
                flag.get('accion_sugerida', ''),
                flag.get('origen_excel_url', ''),
            ]
            for col_num, value in enumerate(values, 1):
                align = 'right' if col_num in (13, 14, 15, 16, 17, 18, 19, 20, 21, 22) else 'left'
                if col_num in (1, 2, 3, 9, 10, 11):
                    align = 'center'
                number_format = None
                if col_num in (13, 14, 15, 17, 18, 19, 20, 21, 22):
                    number_format = '"S/ "#,##0.00'
                elif col_num == 16:
                    number_format = '0.0%'
                    value = value / 100 if value is not None else None
                cell = data_cell(ws, row, col_num, value, fill=fill, number_format=number_format, align=align)
                if col_num == 6:
                    cell.number_format = '@'
                if col_num == 27 and value:
                    cell.hyperlink = value
                    cell.style = 'Hyperlink'

    auto_width(ws)

    ws2 = wb.create_sheet('Variación por concepto')
    ws2['A1'] = f'Variación por concepto - {periodo}'
    ws2['A1'].font = Font(size=13, bold=True, color='0D2B27')
    ws2['A1'].fill = title_fill
    ws2.merge_cells('A1:J1')
    headers2 = [
        'Concepto', 'Código', 'Tipo', 'Monto actual', 'Monto anterior', 'Dif. S/',
        'Var. %', 'Trab. actual', 'Trab. anterior', 'Estado',
    ]
    header_row(ws2, 3, headers2)
    ws2.freeze_panes = 'A4'
    ws2.auto_filter.ref = f'A3:J{max(4, len(variance) + 3)}'
    if not variance:
        ws2.cell(row=4, column=1, value='Sin conceptos calculados para comparar.')
        ws2.merge_cells('A4:J4')
    else:
        for idx, row_data in enumerate(variance, 4):
            estado = 'Nuevo' if row_data['nuevo'] else ('Ya no aparece' if row_data['desaparecio'] else ('Revisar' if row_data['resaltar'] else 'OK'))
            fill = pending_fill if row_data['resaltar'] else None
            values = [
                row_data['nombre'],
                row_data['codigo'],
                row_data['tipo'],
                _decimal_to_float(row_data['total_act']),
                _decimal_to_float(row_data['total_ant']),
                _decimal_to_float(row_data['delta']),
                _decimal_to_float(row_data['pct']),
                row_data['n_act'],
                row_data['n_ant'],
                estado,
            ]
            for col_num, value in enumerate(values, 1):
                align = 'right' if col_num in (4, 5, 6, 7, 8, 9) else 'left'
                number_format = None
                if col_num in (4, 5, 6):
                    number_format = '"S/ "#,##0.00'
                elif col_num == 7:
                    number_format = '0.0%'
                    value = value / 100 if value is not None else None
                data_cell(ws2, idx, col_num, value, fill=fill, number_format=number_format, align=align)
    auto_width(ws2)

    ws3 = wb.create_sheet('Guía de decisión')
    ws3['A1'] = 'Cómo usar este archivo'
    ws3['A1'].font = Font(size=13, bold=True, color='0D2B27')
    ws3['A1'].fill = title_fill
    ws3.merge_cells('A1:C1')
    header_row(ws3, 3, ['Paso', 'Qué revisar', 'Resultado esperado'])
    guia = [
        ('1', 'Filtrar Severidad = critico y Estado revisión = Pendiente.', 'Resolver primero lo que puede cambiar el pago o la declaración.'),
        ('2', 'Abrir el enlace de Origen de cada fila.', 'Corregir datos de trabajador, asistencia, conceptos o contrato desde la fuente.'),
        ('3', 'Revisar Variación por concepto.', 'Detectar si el cambio viene de sueldo, descuentos, horas extra u otros conceptos.'),
        ('4', 'Si la alerta es válida, descartarla en Harmoni con motivo.', 'Queda registrado quién decidió aprobarla y por qué.'),
        ('5', 'Recalcular y volver a descargar si hubo correcciones.', 'Aprobar solo con la versión final revisada.'),
    ]
    for row, values in enumerate(guia, 4):
        for col_num, value in enumerate(values, 1):
            data_cell(ws3, row, col_num, value, align='left')
    auto_width(ws3)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"revision_inconsistencias_{periodo.anio}_{periodo.mes:02d}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
@solo_admin
@require_POST
def periodo_flag_descartar(request, pk):
    """Descarta (o restaura) un flag de la revisión. Queda registro de quién."""
    periodo = get_object_or_404(PeriodoNomina, pk=pk)
    if periodo.estado in ('CERRADO', 'ANULADO'):
        messages.error(request, 'El período ya está cerrado.')
        return redirect('nominas_periodo_revision', pk=pk)

    key = (request.POST.get('flag_key') or '').strip()
    if not key:
        return redirect('nominas_periodo_revision', pk=pk)

    descartados = dict(periodo.flags_descartados or {})
    if key in descartados:
        # Restaurar: vuelve a quedar pendiente
        descartados.pop(key)
        messages.info(request, 'Flag restaurado — vuelve a quedar pendiente de revisión.')
    else:
        descartados[key] = {
            'por': request.user.get_username(),
            'en': timezone.now().isoformat(),
            'mensaje': (request.POST.get('mensaje') or '')[:300],
        }
        messages.success(request, 'Flag descartado. Queda registrado en la auditoría del período.')

    periodo.flags_descartados = descartados
    periodo.save(update_fields=['flags_descartados'])
    return redirect('nominas_periodo_revision', pk=pk)
