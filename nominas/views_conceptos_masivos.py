"""
Importar/exportar conceptos masivos por Excel para un período de nómina.

Permite al admin descargar un Excel con columnas por concepto manual
(propinas, bonificaciones, comisiones, etc.) por trabajador, llenarlo
y subirlo de vuelta para asignar montos en masa.

Round-trip extendido (2026-06): la plantilla incluye además columnas de
datos del registro (días, horas extra, otros ingresos/descuentos) con
códigos reservados `__campo` en la fila 2. Al reimportar, esas columnas
actualizan el RegistroNomina y la planilla se recalcula completa —
sirve para cargar HE, ajustar días o agregar montos (p.ej. utilidades
vía concepto manual) exportando, editando y volviendo a subir.
"""
from decimal import Decimal, InvalidOperation

from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db import models, transaction
from django.db.models import Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect

solo_admin = user_passes_test(lambda u: u.is_superuser)

from . import engine
from .models import (
    PeriodoNomina,
    ConceptoRemunerativo,
    LineaNomina,
)

# Columnas de datos del RegistroNomina incluidas en la plantilla (además de
# los conceptos manuales). Código reservado (fila 2) → (etiqueta, atributo,
# tipo de dato). Tipos: 'dias' = entero 0–31, 'horas' = decimal ≥ 0,
# 'monto' = decimal ≥ 0. Celda vacía al importar = no modificar el campo.
CAMPOS_REGISTRO = {
    '__dias_trabajados':  ('Días Trab.',           'dias_trabajados',  'dias'),
    '__dias_falta':       ('Días Falta',           'dias_falta',       'dias'),
    '__he_25':            ('HE 25% (hrs)',         'horas_extra_25',   'horas'),
    '__he_35':            ('HE 35% (hrs)',         'horas_extra_35',   'horas'),
    '__he_100':           ('HE 100% (hrs)',        'horas_extra_100',  'horas'),
    '__otros_ingresos':   ('Otros Ingresos (S/)',  'otros_ingresos',   'monto'),
    '__otros_descuentos': ('Otros Desctos. (S/)',  'otros_descuentos', 'monto'),
}


def _parse_valor_registro(val, kind):
    """Valida y convierte un valor de columna `__campo`. None = inválido/skip."""
    try:
        d = Decimal(str(val))
    except (ValueError, TypeError, InvalidOperation):
        return None
    if d < 0:
        return None
    if kind == 'dias':
        if d != d.to_integral_value() or d > 31:
            return None
        return int(d)
    return d.quantize(Decimal('0.01'))


@login_required
@solo_admin
def periodo_conceptos_exportar(request, pk):
    """Exporta plantilla Excel del período con columnas por cada concepto manual.

    Estructura:
      DNI | Apellidos y Nombres | Grupo | [Concepto1] | [Concepto2] | ... | [ConceptoN]

    Fila 2 contiene los códigos de los conceptos (referencia para reimportar).
    Pre-llena con valores actuales si los hay (de RegistroNomina.conceptos_manuales).
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    periodo = get_object_or_404(PeriodoNomina, pk=pk)
    registros = periodo.registros.select_related('personal').order_by(
        'personal__apellidos_nombres'
    )

    # Conceptos manuales = formula='MANUAL' o categorías editables
    conceptos = ConceptoRemunerativo.objects.filter(activo=True).filter(
        models.Q(formula='MANUAL')
        | models.Q(categoria__in=[
            'BONIFICACION', 'COMISION', 'PROPINAS',
            'MOVILIDAD', 'OTROS_ING', 'DESCUENTO',
        ])
    ).order_by('tipo', 'orden', 'nombre').distinct()

    conceptos = list(conceptos)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Conceptos {periodo.anio}-{periodo.mes:02d}'

    # Styles
    header_fill = PatternFill(start_color='0F766E', end_color='0F766E', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    sub_fill = PatternFill(start_color='99F6E4', end_color='99F6E4', fill_type='solid')
    sub_font = Font(italic=True, color='0D2B27', size=8)
    border_thin = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1'),
    )

    # Row 1: Títulos principales
    # Tras las 3 fijas van las columnas del registro (días/HE/otros) y
    # luego una por concepto manual. La fila 2 lleva el código de cada una.
    headers_fijos = ['DNI', 'Apellidos y Nombres', 'Grupo']
    cols_registro = list(CAMPOS_REGISTRO.items())  # (codigo, (label, attr, kind))
    headers = (
        headers_fijos
        + [label for _, (label, _, _) in cols_registro]
        + [c.nombre for c in conceptos]
    )
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border_thin

    # Row 2: códigos (REFERENCIA para parsing en import)
    codigos_fila2 = [''] * len(headers_fijos) \
        + [cod for cod, _ in cols_registro] \
        + [c.codigo for c in conceptos]
    for col_idx, cod in enumerate(codigos_fila2, 1):
        cell = ws.cell(row=2, column=col_idx, value=cod)
        cell.font = sub_font
        cell.fill = sub_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border_thin

    # Row 3: tipo (Asistencia / INGRESO / DESCUENTO) — referencia visual
    for col_idx in range(1, len(headers_fijos) + 1):
        ws.cell(row=3, column=col_idx, value='').border = border_thin
    col_idx = len(headers_fijos) + 1
    for _cod, (_label, _attr, kind) in cols_registro:
        tipo_label = 'Días/Horas' if kind in ('dias', 'horas') else 'Registro'
        cell = ws.cell(row=3, column=col_idx, value=tipo_label)
        cell.font = Font(italic=True, size=8, color='0369A1')
        cell.alignment = Alignment(horizontal='center')
        cell.border = border_thin
        col_idx += 1
    for c in conceptos:
        tipo_label = '+ Ingreso' if c.tipo == 'INGRESO' else 'Descuento'
        cell = ws.cell(row=3, column=col_idx, value=tipo_label)
        cell.font = Font(
            italic=True, size=8,
            color='059669' if c.tipo == 'INGRESO' else 'DC2626'
        )
        cell.alignment = Alignment(horizontal='center')
        cell.border = border_thin
        col_idx += 1

    # Datos: fila por trabajador (desde row=4)
    for row_idx, reg in enumerate(registros, 4):
        ws.cell(row=row_idx, column=1, value=reg.personal.nro_doc).border = border_thin
        ws.cell(row=row_idx, column=2, value=reg.personal.apellidos_nombres).border = border_thin
        ws.cell(row=row_idx, column=3, value=reg.grupo or '').border = border_thin

        # Columnas del registro (pre-llenas con el valor actual)
        col_idx = len(headers_fijos) + 1
        for _cod, (_label, attr, kind) in cols_registro:
            valor = getattr(reg, attr)
            if kind == 'dias':
                valor_num = int(valor)
            else:
                valor_num = float(valor) if valor else None
            cell = ws.cell(row=row_idx, column=col_idx, value=valor_num)
            cell.number_format = '0' if kind == 'dias' else '#,##0.00'
            cell.border = border_thin
            cell.alignment = Alignment(horizontal='right')
            col_idx += 1

        cm = reg.conceptos_manuales or {}
        for c in conceptos:
            valor = cm.get(c.codigo)
            try:
                valor_num = float(valor) if valor not in (None, '', 0) else None
            except (ValueError, TypeError):
                valor_num = None
            cell = ws.cell(row=row_idx, column=col_idx, value=valor_num)
            cell.number_format = '#,##0.00'
            cell.border = border_thin
            cell.alignment = Alignment(horizontal='right')
            col_idx += 1

    # Anchos
    ws.column_dimensions['A'].width = 13
    ws.column_dimensions['B'].width = 32
    ws.column_dimensions['C'].width = 10
    for col_idx in range(len(headers_fijos) + 1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    # Freeze panes
    ws.freeze_panes = 'D4'

    # Instrucción al final
    last_row = len(registros) + 6
    instr = ws.cell(
        row=last_row,
        column=1,
        value=(
            'Instrucciones: completá los montos en cada celda. '
            'Las columnas de días/HE/otros actualizan el registro del trabajador '
            '(celda vacía = no se modifica); las demás son conceptos manuales. '
            'Cargá este Excel desde "Importar conceptos masivos" en el período. '
            'La fila 2 contiene los CODIGOS (NO la modifiques).'
        )
    )
    instr.font = Font(italic=True, color='64748B', size=9)
    ws.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=len(headers))

    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    fname = f'conceptos_{periodo.anio}-{periodo.mes:02d}_{periodo.tipo or "REG"}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    wb.save(response)
    return response


@login_required
@solo_admin
def periodo_conceptos_importar(request, pk):
    """Importa Excel y aplica conceptos manuales a los registros del período."""
    import openpyxl

    periodo = get_object_or_404(PeriodoNomina, pk=pk)

    if periodo.estado in ('APROBADO', 'CERRADO', 'ANULADO'):
        messages.error(
            request,
            'El período está aprobado o cerrado. No se pueden importar conceptos.'
        )
        return redirect('nominas_periodo_detalle', pk=pk)

    if request.method != 'POST':
        return redirect('nominas_periodo_detalle', pk=pk)

    archivo = request.FILES.get('archivo')
    if not archivo:
        messages.error(request, 'Subí un archivo Excel.')
        return redirect('nominas_periodo_detalle', pk=pk)

    try:
        wb = openpyxl.load_workbook(archivo, data_only=True)
        ws = wb.active
    except Exception as e:
        messages.error(request, f'Excel inválido: {e}')
        return redirect('nominas_periodo_detalle', pk=pk)

    # Fila 2 tiene los códigos (columna >= 4)
    codigos = []
    col = 4
    while True:
        v = ws.cell(row=2, column=col).value
        if not v:
            break
        codigos.append(str(v).strip())
        col += 1

    if not codigos:
        messages.error(
            request,
            'No se detectaron códigos de conceptos en la fila 2. '
            'Descargá la plantilla del período primero.'
        )
        return redirect('nominas_periodo_detalle', pk=pk)

    # Separar códigos de registro (`__campo`) de códigos de concepto
    codigos_concepto = [c for c in codigos if not c.startswith('__')]
    conceptos_db = {
        c.codigo: c for c in ConceptoRemunerativo.objects.filter(
            codigo__in=codigos_concepto, activo=True
        )
    }
    invalidos = [
        c for c in codigos
        if not c.startswith('__') and c not in conceptos_db
        or c.startswith('__') and c not in CAMPOS_REGISTRO
    ]
    if invalidos:
        messages.warning(
            request,
            f'Códigos no encontrados (ignorados): {", ".join(invalidos)}'
        )

    conceptos_activos = ConceptoRemunerativo.objects.filter(activo=True).order_by('tipo', 'orden')

    # Procesar filas desde row=4
    registros_actualizados = 0
    row = 4
    while True:
        dni = ws.cell(row=row, column=1).value
        if not dni:
            break
        dni = str(dni).strip()

        reg = periodo.registros.filter(personal__nro_doc=dni).first()
        if not reg:
            row += 1
            continue

        nuevos = {}
        campos_reg_update = []
        for idx, cod in enumerate(codigos):
            val = ws.cell(row=row, column=4 + idx).value

            if cod.startswith('__'):
                # Columna de datos del registro: vacía = no modificar
                if cod not in CAMPOS_REGISTRO or val is None or val == '':
                    continue
                _label, attr, kind = CAMPOS_REGISTRO[cod]
                parsed = _parse_valor_registro(val, kind)
                if parsed is None:
                    continue
                if getattr(reg, attr) != parsed:
                    setattr(reg, attr, parsed)
                    campos_reg_update.append(attr)
                continue

            if cod not in conceptos_db:
                continue
            if val is None or val == '':
                continue
            try:
                monto = Decimal(str(val))
            except (ValueError, ArithmeticError):
                continue
            if monto != 0:
                nuevos[cod] = str(monto)

        reg.conceptos_manuales = nuevos

        # Recalcular registro
        with transaction.atomic():
            reg.save()
            resultado = engine.calcular_registro(reg, conceptos_activos)
            reg.lineas.all().delete()
            for l in resultado['lineas']:
                LineaNomina.objects.create(
                    registro=reg,
                    concepto=l['concepto'],
                    base_calculo=l['base_calculo'],
                    porcentaje_aplicado=l['porcentaje_aplicado'],
                    monto=l['monto'],
                    observacion=l['observacion'],
                )
            reg.total_ingresos = resultado['total_ingresos']
            reg.total_descuentos = resultado['total_descuentos']
            reg.neto_a_pagar = resultado['neto_a_pagar']
            reg.aporte_essalud = resultado['aporte_essalud']
            reg.costo_total_empresa = resultado['costo_total_empresa']
            reg.estado = 'CALCULADO'
            reg.save()

        registros_actualizados += 1
        row += 1

    # Resync totales del período
    agg = periodo.registros.aggregate(
        t_bruto=Sum('total_ingresos'),
        t_desc=Sum('total_descuentos'),
        t_neto=Sum('neto_a_pagar'),
        t_costo=Sum('costo_total_empresa'),
    )
    periodo.total_bruto = agg['t_bruto'] or Decimal('0')
    periodo.total_descuentos = agg['t_desc'] or Decimal('0')
    periodo.total_neto = agg['t_neto'] or Decimal('0')
    periodo.total_costo_empresa = agg['t_costo'] or Decimal('0')
    periodo.save(update_fields=[
        'total_bruto', 'total_descuentos', 'total_neto', 'total_costo_empresa'
    ])

    messages.success(
        request,
        f'Conceptos importados correctamente para {registros_actualizados} '
        f'trabajadores. Planilla recalculada.'
    )
    return redirect('nominas_periodo_detalle', pk=pk)
