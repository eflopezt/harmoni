"""
Tests del importador de personal en 2 pasos (preview → confirmar).

Cubre:
- Parseo: CREAR vs ACTUALIZAR, errores (DNI inválido, sueldo inválido,
  CCI mal formado, DNI duplicado en archivo), warnings (subárea inexistente).
- Columnas nuevas de planilla: SueldoBase, RegimenPension, AFP, Banco, CCI.
- Flujo web: subir → preview (nada en BD) → confirmar → solo filas sin
  error se aplican; token inválido expira con gracia.
- Fix del bug histórico: la SubArea ahora SÍ se asigna al importar.
"""
import io
from datetime import date
from decimal import Decimal

import openpyxl
import pytest
from django.contrib.auth.models import User
from django.urls import reverse

from personal.import_preview import aplicar_importacion, parsear_excel_personal
from personal.models import Area, Personal, SubArea


@pytest.fixture
def subarea(db):
    area = Area.objects.create(nombre="Operaciones")
    return SubArea.objects.create(nombre="Campo", area=area)


def _xlsx(filas, columnas):
    """Crea un xlsx en memoria con hoja 'Personal'."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Personal"
    for c, col in enumerate(columnas, 1):
        ws.cell(row=1, column=c, value=col)
    for r, fila in enumerate(filas, 2):
        for c, val in enumerate(fila, 1):
            ws.cell(row=r, column=c, value=val)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    buf.name = "personal.xlsx"
    return buf


COLS = ["NroDoc", "ApellidosNombres", "SubArea", "SueldoBase",
        "RegimenPension", "AFP", "Banco", "CuentaCCI", "FechaAlta"]


class TestParseo:
    def test_crear_actualizar_y_validaciones(self, subarea):
        Personal.objects.create(
            nro_doc="71000070", apellidos_nombres="EXISTE, YA",
            cargo="X", tipo_trab="Empleado", estado="Activo",
            subarea=subarea, fecha_alta=date(2024, 1, 1),
        )
        archivo = _xlsx([
            ("71000070", "EXISTE, YA",  "Campo",  "2500", "ONP", "", "BCP", "", "2024-01-01"),
            ("71000071", "NUEVA, ANA",  "Campo",  "3000", "AFP", "Prima", "BCP",
             "00219112345678901234", "2026-01-15"),
            ("123",      "DNI MALO, X", "",       "2000", "ONP", "", "", "", ""),
            ("71000072", "SUELDO MALO", "",       "abc",  "ONP", "", "", "", ""),
            ("71000073", "CCI MALO",    "",       "2000", "ONP", "", "BCP", "999", ""),
            ("71000071", "DUPLICADA",   "",       "1000", "ONP", "", "", "", ""),
            ("71000074", "SUB RARA",    "NoExiste", "2000", "ONP", "", "", "", ""),
        ], COLS)

        parse = parsear_excel_personal(archivo)
        assert parse["error_global"] is None
        por_dni = {}
        for f in parse["filas"]:
            por_dni.setdefault(f["nro_doc"], f)

        assert por_dni["71000070"]["accion"] == "ACTUALIZAR"
        assert por_dni["71000071"]["accion"] == "CREAR"
        assert not por_dni["71000071"]["errores"]
        assert por_dni["71000071"]["datos"]["sueldo_base"] == Decimal("3000")
        assert por_dni["71000071"]["datos"]["afp"] == "Prima"
        assert por_dni["71000071"]["datos"]["cuenta_cci"] == "00219112345678901234"
        assert por_dni["71000071"]["datos"]["subarea"] == subarea  # bug fix

        assert any("DNI inválido" in e for e in por_dni["123"]["errores"])
        assert any("SueldoBase inválido" in e for e in por_dni["71000072"]["errores"])
        assert any("CuentaCCI inválida" in e for e in por_dni["71000073"]["errores"])
        # fila 7 (segunda 71000071) duplicada
        dupes = [f for f in parse["filas"] if f["nro_doc"] == "71000071" and f["errores"]]
        assert dupes and any("repetido" in e for e in dupes[0]["errores"])
        assert any("no existe" in w for w in por_dni["71000074"]["warnings"])

        assert parse["con_error"] == 4
        assert parse["a_crear"] == 2   # 71000071 y 71000074
        assert parse["a_actualizar"] == 1

    def test_sin_hoja_personal(self, db):
        wb = openpyxl.Workbook()
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        parse = parsear_excel_personal(buf)
        assert parse["error_global"] and "Personal" in parse["error_global"]

    def test_aplicar_solo_filas_sin_error(self, subarea):
        archivo = _xlsx([
            ("71000080", "APLICA, UNO", "Campo", "2500", "ONP", "", "", "", "2026-01-01"),
            ("123",      "ERROR, DOS",  "",      "2000", "ONP", "", "", "", ""),
        ], COLS)
        parse = parsear_excel_personal(archivo)
        resultado = aplicar_importacion(parse)
        assert resultado == {"creados": 1, "actualizados": 0, "omitidos": 1}
        p = Personal.objects.get(nro_doc="71000080")
        assert p.subarea == subarea
        assert p.sueldo_base == Decimal("2500")
        assert not Personal.objects.filter(nro_doc="123").exists()


class TestFlujoWeb:
    @pytest.fixture
    def admin_client(self, db, client):
        User.objects.create_superuser("admin", "a@a.com", "x")
        client.login(username="admin", password="x")
        return client

    def test_preview_no_escribe_y_confirmar_si(self, admin_client, subarea):
        archivo = _xlsx([
            ("71000081", "WEB, FLUJO", "Campo", "2800", "AFP", "Habitat", "BCP",
             "00219112345678901234", "2026-02-01"),
        ], COLS)
        url = reverse("personal_import")

        resp = admin_client.post(url, {"archivo": archivo})
        assert resp.status_code == 200
        assert b"Revisa antes de importar" in resp.content
        assert not Personal.objects.filter(nro_doc="71000081").exists()  # nada en BD

        token = resp.context["token"]
        resp = admin_client.post(url, {"confirmar_token": token})
        assert resp.status_code == 302
        p = Personal.objects.get(nro_doc="71000081")
        assert p.afp == "Habitat"
        assert p.regimen_pension == "AFP"

    def test_token_invalido_expira_con_gracia(self, admin_client):
        resp = admin_client.post(reverse("personal_import"),
                                 {"confirmar_token": "deadbeef" * 4})
        assert resp.status_code == 302  # redirect con mensaje, sin crash
