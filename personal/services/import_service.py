"""
Servicio de importación masiva de personal desde Excel.

Soporta:
- Validación previa con preview antes de importar
- Detección de duplicados por nro_doc (DNI)
- Auto-creación de áreas y subáreas
- Múltiples formatos de fecha (DD/MM/YYYY, YYYY-MM-DD, D/M/YYYY)
- Codificación UTF-8 y latin-1
- Generación de plantilla con validaciones Excel
"""
import logging
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from django.db import transaction

from personal.models import Area, SubArea, Personal

logger = logging.getLogger('personal')


class PersonalImportService:
    """Import employees from Excel with validation and error reporting."""

    # Columnas requeridas (deben estar presentes y no vacías)
    REQUIRED_COLUMNS = ['nro_doc', 'apellido_paterno', 'apellido_materno', 'nombres']

    # Columnas opcionales soportadas
    OPTIONAL_COLUMNS = [
        'tipo_doc', 'fecha_nacimiento', 'sexo', 'email', 'correo_corporativo',
        'celular', 'direccion', 'cargo', 'tipo_trab', 'sueldo_base', 'bonos',
        'tipo_contrato', 'fecha_inicio_contrato', 'fecha_fin_contrato',
        'fecha_alta', 'area', 'subarea', 'regimen_pension', 'afp', 'cuspp',
        'condicion', 'grupo_tareo', 'asignacion_familiar', 'estado',
        'banco', 'cuenta_ahorros', 'cuenta_cci', 'cuenta_cts',
        'regimen_laboral', 'regimen_turno', 'jornada_horas',
        'codigo_fotocheck', 'codigo_sap', 'codigo_s10',
        'observaciones',
    ]

    # Mapeo de sinónimos de columnas (Excel header -> campo interno)
    COLUMN_ALIASES = {
        # nro_doc
        'nro_doc': 'nro_doc', 'nrodoc': 'nro_doc', 'dni': 'nro_doc',
        'documento': 'nro_doc', 'num_doc': 'nro_doc', 'numero_documento': 'nro_doc',
        'nro documento': 'nro_doc', 'n doc': 'nro_doc',
        # nombres
        'apellido_paterno': 'apellido_paterno', 'apellido paterno': 'apellido_paterno',
        'ap_paterno': 'apellido_paterno', 'paterno': 'apellido_paterno',
        'apellido_materno': 'apellido_materno', 'apellido materno': 'apellido_materno',
        'ap_materno': 'apellido_materno', 'materno': 'apellido_materno',
        'nombres': 'nombres', 'nombre': 'nombres',
        'apellidos_nombres': 'apellidos_nombres', 'apellidosnombres': 'apellidos_nombres',
        'apellidos y nombres': 'apellidos_nombres',
        # tipo_doc
        'tipo_doc': 'tipo_doc', 'tipodoc': 'tipo_doc', 'tipo documento': 'tipo_doc',
        # datos personales
        'fecha_nacimiento': 'fecha_nacimiento', 'fechanacimiento': 'fecha_nacimiento',
        'fec_nacimiento': 'fecha_nacimiento', 'fec nacimiento': 'fecha_nacimiento',
        'sexo': 'sexo', 'genero': 'sexo',
        'email': 'email', 'correo': 'email', 'correo_personal': 'email',
        'correopersonal': 'email', 'correo personal': 'email',
        'correo_corporativo': 'correo_corporativo', 'correocorporativo': 'correo_corporativo',
        'correo corporativo': 'correo_corporativo', 'email_corporativo': 'correo_corporativo',
        'celular': 'celular', 'telefono': 'celular', 'movil': 'celular',
        'direccion': 'direccion', 'domicilio': 'direccion',
        # datos laborales
        'cargo': 'cargo', 'puesto': 'cargo', 'posicion': 'cargo',
        'tipo_trab': 'tipo_trab', 'tipotrab': 'tipo_trab',
        'tipo trabajador': 'tipo_trab', 'tipo_trabajador': 'tipo_trab',
        'sueldo_base': 'sueldo_base', 'sueldobase': 'sueldo_base',
        'sueldo': 'sueldo_base', 'remuneracion': 'sueldo_base', 'basico': 'sueldo_base',
        'bonos': 'bonos', 'bono': 'bonos',
        # contrato
        'tipo_contrato': 'tipo_contrato', 'tipocontrato': 'tipo_contrato',
        'modalidad': 'tipo_contrato', 'tipo contrato': 'tipo_contrato',
        'fecha_inicio_contrato': 'fecha_inicio_contrato',
        'fechainiciocontrato': 'fecha_inicio_contrato',
        'inicio_contrato': 'fecha_inicio_contrato', 'inicio contrato': 'fecha_inicio_contrato',
        'fecha_fin_contrato': 'fecha_fin_contrato',
        'fechafincontrato': 'fecha_fin_contrato',
        'fin_contrato': 'fecha_fin_contrato', 'fin contrato': 'fecha_fin_contrato',
        'fecha_alta': 'fecha_alta', 'fechaalta': 'fecha_alta',
        'fecha alta': 'fecha_alta', 'fecha_ingreso': 'fecha_alta',
        'fecha ingreso': 'fecha_alta', 'ingreso': 'fecha_alta',
        # clasificación
        'area': 'area', 'área': 'area',
        'subarea': 'subarea', 'subárea': 'subarea', 'sub_area': 'subarea',
        'sub area': 'subarea', 'sub área': 'subarea',
        'regimen_pension': 'regimen_pension', 'regimenpension': 'regimen_pension',
        'regimen pension': 'regimen_pension', 'pension': 'regimen_pension',
        'regimen pensionario': 'regimen_pension',
        'afp': 'afp',
        'cuspp': 'cuspp',
        'condicion': 'condicion', 'condición': 'condicion',
        'grupo_tareo': 'grupo_tareo', 'grupotareo': 'grupo_tareo',
        'grupo tareo': 'grupo_tareo',
        'asignacion_familiar': 'asignacion_familiar',
        'asignacionfamiliar': 'asignacion_familiar',
        'asignacion familiar': 'asignacion_familiar', 'asig_familiar': 'asignacion_familiar',
        'estado': 'estado',
        # bancarios
        'banco': 'banco',
        'cuenta_ahorros': 'cuenta_ahorros', 'cta_ahorros': 'cuenta_ahorros',
        'cuenta ahorros': 'cuenta_ahorros',
        'cuenta_cci': 'cuenta_cci', 'cci': 'cuenta_cci',
        'cuenta_cts': 'cuenta_cts', 'cts': 'cuenta_cts', 'cuenta cts': 'cuenta_cts',
        # otros
        'regimen_laboral': 'regimen_laboral', 'regimen laboral': 'regimen_laboral',
        'regimen_turno': 'regimen_turno', 'regimen turno': 'regimen_turno',
        'turno': 'regimen_turno',
        'jornada_horas': 'jornada_horas', 'jornada': 'jornada_horas',
        'horas_jornada': 'jornada_horas',
        'codigo_fotocheck': 'codigo_fotocheck', 'fotocheck': 'codigo_fotocheck',
        'codigo_sap': 'codigo_sap', 'sap': 'codigo_sap',
        'codigo_s10': 'codigo_s10', 's10': 'codigo_s10',
        'observaciones': 'observaciones', 'notas': 'observaciones',
    }

    # Valores válidos para campos con choices
    VALID_CHOICES = {
        'tipo_doc': {v[0].lower(): v[0] for v in Personal.TIPO_DOC_CHOICES},
        'tipo_trab': {v[0].lower(): v[0] for v in Personal.TIPO_TRAB_CHOICES},
        'sexo': {'m': 'M', 'f': 'F', 'masculino': 'M', 'femenino': 'F'},
        'estado': {v[0].lower(): v[0] for v in Personal.ESTADO_CHOICES},
        'regimen_pension': {v[0].lower(): v[0] for v in Personal.REGIMEN_PENSION_CHOICES},
        'afp': {v[0].lower(): v[0] for v in Personal.AFP_CHOICES},
        'tipo_contrato': {v[0].lower(): v[0] for v in Personal.TIPO_CONTRATO_CHOICES},
        'condicion': {v[0].lower(): v[0] for v in Personal.CONDICION_CHOICES},
        'grupo_tareo': {v[0].lower(): v[0] for v in Personal.GRUPO_TAREO_CHOICES},
        'banco': {v[0].lower(): v[0] for v in Personal.BANCO_CHOICES},
    }

    # -----------------------------------------------------------------
    # Lectura del archivo Excel
    # -----------------------------------------------------------------

    def _read_excel(self, file) -> tuple:
        """
        Lee un archivo Excel y retorna (headers, rows, error).
        Intenta la hoja 'Personal' primero, luego la primera hoja.
        """
        try:
            wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        except Exception as e:
            return None, None, f'No se pudo abrir el archivo Excel: {e}'

        # Buscar hoja
        if 'Personal' in wb.sheetnames:
            ws = wb['Personal']
        else:
            ws = wb[wb.sheetnames[0]]

        rows_iter = ws.iter_rows(values_only=True)

        # Leer encabezados
        try:
            raw_headers = next(rows_iter)
        except StopIteration:
            wb.close()
            return None, None, 'El archivo Excel está vacío.'

        if not raw_headers or all(h is None for h in raw_headers):
            wb.close()
            return None, None, 'No se encontraron encabezados en la primera fila.'

        # Normalizar encabezados
        headers = []
        for h in raw_headers:
            if h is None:
                headers.append(None)
                continue
            normalized = str(h).strip().lower().replace(' ', '_')
            # Quitar tildes para matching
            normalized = (
                normalized
                .replace('á', 'a').replace('é', 'e').replace('í', 'i')
                .replace('ó', 'o').replace('ú', 'u').replace('ñ', 'n')
            )
            mapped = self.COLUMN_ALIASES.get(normalized)
            if not mapped:
                # Intentar sin guiones bajos
                no_underscore = normalized.replace('_', '')
                mapped = self.COLUMN_ALIASES.get(no_underscore, normalized)
            headers.append(mapped)

        # Leer filas de datos
        data_rows = []
        for row_values in rows_iter:
            # Saltar filas completamente vacías
            if all(v is None or str(v).strip() == '' for v in row_values):
                continue
            row_dict = {}
            for i, val in enumerate(row_values):
                if i < len(headers) and headers[i]:
                    row_dict[headers[i]] = val
            data_rows.append(row_dict)

        wb.close()
        return headers, data_rows, None

    # -----------------------------------------------------------------
    # Parseo de valores
    # -----------------------------------------------------------------

    def _parse_date(self, value) -> tuple:
        """Parsea un valor a date. Retorna (date|None, error|None)."""
        if value is None or str(value).strip() == '':
            return None, None

        # Si ya es date/datetime
        if isinstance(value, datetime):
            return value.date(), None
        if isinstance(value, date):
            return value, None

        text = str(value).strip()

        # Formatos comunes
        formats = [
            '%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%Y/%m/%d',
            '%d/%m/%y', '%d-%m-%y', '%m/%d/%Y', '%d.%m.%Y',
        ]
        for fmt in formats:
            try:
                return datetime.strptime(text, fmt).date(), None
            except ValueError:
                continue

        return None, f'Fecha no reconocida: "{text}"'

    def _parse_decimal(self, value, field_name='') -> tuple:
        """Parsea un valor a Decimal. Retorna (Decimal|None, error|None)."""
        if value is None or str(value).strip() == '':
            return None, None

        text = str(value).strip().replace(',', '')
        try:
            d = Decimal(text)
            if d < 0:
                return None, f'{field_name}: no puede ser negativo'
            return d, None
        except (InvalidOperation, ValueError):
            return None, f'{field_name}: valor numérico no válido: "{value}"'

    def _parse_bool(self, value) -> bool:
        """Parsea un valor a boolean."""
        if value is None:
            return False
        if isinstance(value, bool):
            return value
        text = str(value).strip().lower()
        return text in ('si', 'sí', 'yes', 'true', '1', 'x', 'v')

    def _parse_choice(self, value, field_name) -> tuple:
        """Valida un valor contra los choices del campo. Retorna (valor_correcto, error)."""
        if value is None or str(value).strip() == '':
            return '', None

        text = str(value).strip()
        choices = self.VALID_CHOICES.get(field_name, {})

        # Buscar coincidencia exacta (case-insensitive)
        lower = text.lower()
        if lower in choices:
            return choices[lower], None

        # Buscar coincidencia parcial
        for key, correct in choices.items():
            if key.startswith(lower) or lower.startswith(key):
                return correct, None

        valid_options = ', '.join(set(choices.values()))
        return text, f'{field_name}: "{text}" no es un valor válido. Opciones: {valid_options}'

    # -----------------------------------------------------------------
    # Validación de una fila
    # -----------------------------------------------------------------

    def _validate_row(self, row: dict, row_num: int, existing_docs: set) -> tuple:
        """
        Valida una fila de datos.
        Retorna (parsed_data: dict, errors: list[str], warnings: list[str]).
        """
        errors = []
        warnings = []
        data = {}

        # --- nro_doc ---
        nro_doc_raw = row.get('nro_doc')
        if nro_doc_raw is None or str(nro_doc_raw).strip() in ('', 'nan', 'None'):
            errors.append('nro_doc es obligatorio')
            return data, errors, warnings

        if isinstance(nro_doc_raw, (int, float)):
            nro_doc = str(int(nro_doc_raw))
        else:
            nro_doc = str(nro_doc_raw).strip()

        data['nro_doc'] = nro_doc
        data['is_update'] = nro_doc in existing_docs

        # --- apellidos_nombres ---
        # Soportar campo combinado o campos separados
        if 'apellidos_nombres' in row and row['apellidos_nombres']:
            data['apellidos_nombres'] = str(row['apellidos_nombres']).strip()
        else:
            paterno = str(row.get('apellido_paterno', '') or '').strip()
            materno = str(row.get('apellido_materno', '') or '').strip()
            nombres = str(row.get('nombres', '') or '').strip()

            if not paterno and not nombres:
                errors.append('Se requiere al menos apellido_paterno y nombres')
            else:
                parts = []
                if paterno:
                    parts.append(paterno.upper())
                if materno:
                    parts.append(materno.upper())
                if parts and nombres:
                    data['apellidos_nombres'] = f"{' '.join(parts)}, {nombres.upper()}"
                elif nombres:
                    data['apellidos_nombres'] = nombres.upper()
                else:
                    data['apellidos_nombres'] = ' '.join(parts)

        # --- tipo_doc ---
        val, err = self._parse_choice(row.get('tipo_doc', 'DNI'), 'tipo_doc')
        data['tipo_doc'] = val or 'DNI'
        if err:
            warnings.append(err)

        # Validar longitud de DNI
        if data['tipo_doc'] == 'DNI' and len(nro_doc) != 8:
            warnings.append(f'DNI debería tener 8 dígitos, tiene {len(nro_doc)}')

        # --- Fechas ---
        for field in ['fecha_nacimiento', 'fecha_alta', 'fecha_inicio_contrato', 'fecha_fin_contrato']:
            val, err = self._parse_date(row.get(field))
            if err:
                errors.append(f'{field}: {err}')
            elif val:
                data[field] = val

        # --- Decimales ---
        for field in ['sueldo_base', 'bonos', 'jornada_horas']:
            val, err = self._parse_decimal(row.get(field), field)
            if err:
                errors.append(err)
            elif val is not None:
                data[field] = val

        # --- Choices ---
        for field in ['tipo_trab', 'sexo', 'estado', 'regimen_pension', 'afp',
                       'tipo_contrato', 'condicion', 'grupo_tareo', 'banco']:
            raw = row.get(field)
            if raw is not None and str(raw).strip():
                val, err = self._parse_choice(raw, field)
                if err:
                    warnings.append(err)
                if val:
                    data[field] = val

        # --- Booleanos ---
        if 'asignacion_familiar' in row:
            data['asignacion_familiar'] = self._parse_bool(row['asignacion_familiar'])

        # --- Strings simples ---
        for field in ['celular', 'email', 'correo_corporativo', 'direccion', 'cargo',
                       'cuspp', 'regimen_laboral', 'regimen_turno',
                       'codigo_fotocheck', 'codigo_sap', 'codigo_s10',
                       'cuenta_ahorros', 'cuenta_cci', 'cuenta_cts',
                       'observaciones']:
            raw = row.get(field)
            if raw is not None and str(raw).strip() not in ('', 'nan', 'None'):
                data[field] = str(raw).strip()

        # --- Área y SubÁrea ---
        data['_area_name'] = str(row.get('area', '') or '').strip()
        data['_subarea_name'] = str(row.get('subarea', '') or '').strip()

        return data, errors, warnings

    # -----------------------------------------------------------------
    # Validación completa del Excel
    # -----------------------------------------------------------------

    def validate_excel(self, file) -> dict:
        """
        Validate Excel structure and data.
        Returns {valid: bool, errors: list, warnings: list, preview: list, stats: dict}
        """
        headers, rows, read_error = self._read_excel(file)

        if read_error:
            return {
                'valid': False,
                'errors': [read_error],
                'warnings': [],
                'preview': [],
                'stats': {},
                'headers': [],
            }

        if not rows:
            return {
                'valid': False,
                'errors': ['El archivo no contiene filas de datos.'],
                'warnings': [],
                'preview': [],
                'stats': {},
                'headers': [],
            }

        # Verificar columnas requeridas
        # Aceptar apellidos_nombres como alternativa a los 3 campos separados
        has_combined = 'apellidos_nombres' in headers
        has_separate = all(
            col in headers for col in ['apellido_paterno', 'apellido_materno', 'nombres']
        )
        has_nrodoc = 'nro_doc' in headers

        col_errors = []
        if not has_nrodoc:
            col_errors.append(
                'Falta columna obligatoria: nro_doc (o DNI, NroDoc, Documento)'
            )
        if not has_combined and not has_separate:
            col_errors.append(
                'Faltan columnas de nombre. Use "apellidos_nombres" o '
                '"apellido_paterno" + "apellido_materno" + "nombres"'
            )

        if col_errors:
            return {
                'valid': False,
                'errors': col_errors,
                'warnings': [],
                'preview': [],
                'stats': {},
                'headers': [h for h in headers if h],
            }

        # Obtener docs existentes para detección de duplicados
        existing_docs = set(
            Personal.objects.values_list('nro_doc', flat=True)
        )

        global_errors = []
        global_warnings = []
        preview = []
        seen_docs = set()
        creates = 0
        updates = 0
        error_count = 0

        for idx, row in enumerate(rows, start=2):  # Fila 2 en Excel (fila 1 = headers)
            parsed, row_errors, row_warnings = self._validate_row(row, idx, existing_docs)

            # Detectar duplicados dentro del mismo archivo
            nro_doc = parsed.get('nro_doc', '')
            if nro_doc and nro_doc in seen_docs:
                row_errors.append(f'DNI {nro_doc} duplicado dentro del archivo')
            if nro_doc:
                seen_docs.add(nro_doc)

            is_update = parsed.get('is_update', False)
            has_errors = len(row_errors) > 0

            if has_errors:
                error_count += 1
            elif is_update:
                updates += 1
            else:
                creates += 1

            # Prefixed errors/warnings with row number
            for e in row_errors:
                global_errors.append(f'Fila {idx}: {e}')
            for w in row_warnings:
                global_warnings.append(f'Fila {idx}: {w}')

            preview.append({
                'row_num': idx,
                'nro_doc': nro_doc,
                'apellidos_nombres': parsed.get('apellidos_nombres', ''),
                'cargo': parsed.get('cargo', ''),
                'area': parsed.get('_area_name', ''),
                'subarea': parsed.get('_subarea_name', ''),
                'sueldo_base': str(parsed.get('sueldo_base', '')),
                'estado': parsed.get('estado', 'Activo'),
                'is_update': is_update,
                'errors': row_errors,
                'warnings': row_warnings,
                'data': parsed,
            })

        return {
            'valid': error_count == 0,
            'errors': global_errors,
            'warnings': global_warnings,
            'preview': preview,
            'stats': {
                'total': len(rows),
                'creates': creates,
                'updates': updates,
                'errors': error_count,
            },
            'headers': [h for h in headers if h],
        }

    # -----------------------------------------------------------------
    # Importación
    # -----------------------------------------------------------------

    def import_excel(self, file, empresa, created_by) -> dict:
        """
        Import employees from Excel.
        Returns {created: int, updated: int, errors: list, details: list}
        """
        result = {
            'created': 0,
            'updated': 0,
            'errors': [],
            'details': [],
        }

        headers, rows, read_error = self._read_excel(file)
        if read_error:
            result['errors'].append(read_error)
            return result

        if not rows:
            result['errors'].append('El archivo no contiene filas de datos.')
            return result

        existing_docs = set(Personal.objects.values_list('nro_doc', flat=True))

        # Cache áreas y subáreas
        areas_cache = {a.nombre.lower(): a for a in Area.objects.all()}
        subareas_cache = {}
        for sa in SubArea.objects.select_related('area').all():
            key = f"{sa.area.nombre.lower()}|{sa.nombre.lower()}"
            subareas_cache[key] = sa
            # También indexar solo por nombre de subárea
            subareas_cache[sa.nombre.lower()] = sa

        for idx, row in enumerate(rows, start=2):
            parsed, row_errors, row_warnings = self._validate_row(row, idx, existing_docs)

            if row_errors:
                result['errors'].append(f'Fila {idx}: {"; ".join(row_errors)}')
                continue

            nro_doc = parsed['nro_doc']

            try:
                # Resolver subárea
                subarea = None
                area_name = parsed.pop('_area_name', '')
                subarea_name = parsed.pop('_subarea_name', '')

                if subarea_name:
                    # Buscar con área+subárea
                    if area_name:
                        key = f"{area_name.lower()}|{subarea_name.lower()}"
                        subarea = subareas_cache.get(key)

                    # Buscar solo por nombre de subárea
                    if not subarea:
                        subarea = subareas_cache.get(subarea_name.lower())

                    # Auto-crear si no existe
                    if not subarea:
                        area_obj = None
                        if area_name:
                            area_obj = areas_cache.get(area_name.lower())
                            if not area_obj:
                                area_obj = Area.objects.create(nombre=area_name)
                                areas_cache[area_name.lower()] = area_obj
                                result['details'].append(
                                    f'Fila {idx}: Área "{area_name}" creada automáticamente'
                                )
                        else:
                            # Usar o crear área "Sin Asignar"
                            area_obj = areas_cache.get('sin asignar')
                            if not area_obj:
                                area_obj = Area.objects.create(nombre='Sin Asignar')
                                areas_cache['sin asignar'] = area_obj

                        subarea = SubArea.objects.create(
                            nombre=subarea_name, area=area_obj
                        )
                        key = f"{area_obj.nombre.lower()}|{subarea_name.lower()}"
                        subareas_cache[key] = subarea
                        subareas_cache[subarea_name.lower()] = subarea
                        result['details'].append(
                            f'Fila {idx}: SubÁrea "{subarea_name}" creada automáticamente'
                        )

                elif area_name:
                    # Solo área, buscar primera subárea o crear default
                    area_obj = areas_cache.get(area_name.lower())
                    if not area_obj:
                        area_obj = Area.objects.create(nombre=area_name)
                        areas_cache[area_name.lower()] = area_obj
                        result['details'].append(
                            f'Fila {idx}: Área "{area_name}" creada automáticamente'
                        )
                    # Buscar subárea "General" dentro de esa área
                    general_key = f"{area_name.lower()}|general"
                    subarea = subareas_cache.get(general_key)
                    if not subarea:
                        first_sub = SubArea.objects.filter(area=area_obj, activa=True).first()
                        if first_sub:
                            subarea = first_sub
                        else:
                            subarea = SubArea.objects.create(
                                nombre='General', area=area_obj
                            )
                            subareas_cache[general_key] = subarea
                            result['details'].append(
                                f'Fila {idx}: SubÁrea "General" creada para área "{area_name}"'
                            )

                # Preparar datos para el modelo
                is_update = parsed.pop('is_update', False)
                model_data = {}

                # Mapear campos
                field_map = {
                    'apellidos_nombres': 'apellidos_nombres',
                    'tipo_doc': 'tipo_doc',
                    'fecha_nacimiento': 'fecha_nacimiento',
                    'sexo': 'sexo',
                    'email': 'correo_personal',
                    'correo_corporativo': 'correo_corporativo',
                    'celular': 'celular',
                    'direccion': 'direccion',
                    'cargo': 'cargo',
                    'tipo_trab': 'tipo_trab',
                    'sueldo_base': 'sueldo_base',
                    'bonos': 'bonos',
                    'tipo_contrato': 'tipo_contrato',
                    'fecha_inicio_contrato': 'fecha_inicio_contrato',
                    'fecha_fin_contrato': 'fecha_fin_contrato',
                    'fecha_alta': 'fecha_alta',
                    'regimen_pension': 'regimen_pension',
                    'afp': 'afp',
                    'cuspp': 'cuspp',
                    'condicion': 'condicion',
                    'grupo_tareo': 'grupo_tareo',
                    'asignacion_familiar': 'asignacion_familiar',
                    'estado': 'estado',
                    'banco': 'banco',
                    'cuenta_ahorros': 'cuenta_ahorros',
                    'cuenta_cci': 'cuenta_cci',
                    'cuenta_cts': 'cuenta_cts',
                    'regimen_laboral': 'regimen_laboral',
                    'regimen_turno': 'regimen_turno',
                    'jornada_horas': 'jornada_horas',
                    'codigo_fotocheck': 'codigo_fotocheck',
                    'codigo_sap': 'codigo_sap',
                    'codigo_s10': 'codigo_s10',
                    'observaciones': 'observaciones',
                }

                for src, dst in field_map.items():
                    if src in parsed:
                        model_data[dst] = parsed[src]

                if subarea:
                    model_data['subarea'] = subarea
                if empresa:
                    model_data['empresa'] = empresa

                # Defaults para campos requeridos en nuevos registros
                if not is_update:
                    model_data.setdefault('tipo_doc', 'DNI')
                    model_data.setdefault('tipo_trab', 'Empleado')
                    model_data.setdefault('cargo', 'Sin asignar')
                    model_data.setdefault('estado', 'Activo')

                personal_obj, created = Personal.objects.update_or_create(
                    nro_doc=nro_doc,
                    defaults=model_data
                )

                if created:
                    result['created'] += 1
                    result['details'].append(
                        f'Fila {idx}: {parsed.get("apellidos_nombres", nro_doc)} - CREADO'
                    )
                    existing_docs.add(nro_doc)
                else:
                    result['updated'] += 1
                    result['details'].append(
                        f'Fila {idx}: {parsed.get("apellidos_nombres", nro_doc)} - ACTUALIZADO'
                    )

            except Exception as e:
                logger.exception(f'Error importando fila {idx}')
                result['errors'].append(f'Fila {idx}: Error inesperado: {str(e)}')

        return result

    # -----------------------------------------------------------------
    # Generación de plantilla
    # -----------------------------------------------------------------

    def generate_template(self) -> bytes:
        """
        Generate blank Excel template with headers, example data,
        dropdown validations, and an instructions sheet.
        """
        wb = openpyxl.Workbook()

        # ── Hoja de Instrucciones ─────────────────────────────────────
        ws_inst = wb.active
        ws_inst.title = 'Instrucciones'
        ws_inst.sheet_properties.tabColor = '4472C4'

        instructions = [
            ['PLANTILLA DE IMPORTACION MASIVA DE PERSONAL'],
            [''],
            ['INSTRUCCIONES:'],
            ['1. Complete los datos del personal en la hoja "Personal".'],
            ['2. Los campos obligatorios estan marcados con * en el encabezado.'],
            ['3. Las columnas con listas desplegables tienen valores predefinidos.'],
            ['4. Puede usar apellidos_nombres (combinado) O los 3 campos separados.'],
            ['5. Si usa campos separados, el sistema armara: "PATERNO MATERNO, NOMBRES"'],
            ['6. Las fechas pueden estar en formato DD/MM/YYYY o YYYY-MM-DD.'],
            ['7. Si un DNI ya existe en el sistema, el registro se ACTUALIZARA.'],
            ['8. Si el Area/SubArea no existe, se creara automaticamente.'],
            [''],
            ['CAMPOS OBLIGATORIOS:'],
            ['  - nro_doc: Numero de documento (DNI, CE, Pasaporte)'],
            ['  - apellido_paterno + apellido_materno + nombres, o apellidos_nombres'],
            [''],
            ['CAMPOS OPCIONALES IMPORTANTES:'],
            ['  - cargo: Puesto del trabajador'],
            ['  - sueldo_base: Remuneracion basica mensual'],
            ['  - tipo_contrato: Modalidad contractual (ver hoja Catalogos)'],
            ['  - fecha_alta: Fecha de ingreso a la empresa'],
            ['  - area / subarea: Ubicacion organizacional'],
            ['  - regimen_pension: AFP u ONP'],
            ['  - asignacion_familiar: Si/No (10% RMV por hijo menor)'],
            [''],
            ['NOTAS:'],
            ['  - No modifique la hoja "Catalogos", es de referencia.'],
            ['  - Puede agregar hasta 5,000 filas en una sola importacion.'],
            ['  - El archivo debe guardarse como .xlsx'],
        ]

        title_font = Font(size=14, bold=True, color='1F4E79')
        section_font = Font(size=11, bold=True, color='2E75B6')
        normal_font = Font(size=10)

        for i, row in enumerate(instructions, 1):
            cell = ws_inst.cell(row=i, column=1, value=row[0] if row else '')
            if i == 1:
                cell.font = title_font
            elif row and row[0] and not row[0].startswith(' ') and row[0] != '':
                cell.font = section_font
            else:
                cell.font = normal_font

        ws_inst.column_dimensions['A'].width = 80

        # ── Hoja de Personal ─────────────────────────────────────────
        ws = wb.create_sheet('Personal')
        ws.sheet_properties.tabColor = '00B050'

        headers = [
            ('nro_doc *', 15),
            ('tipo_doc', 12),
            ('apellido_paterno *', 20),
            ('apellido_materno *', 20),
            ('nombres *', 20),
            ('cargo', 25),
            ('fecha_nacimiento', 16),
            ('sexo', 8),
            ('celular', 14),
            ('email', 28),
            ('correo_corporativo', 28),
            ('direccion', 35),
            ('sueldo_base', 14),
            ('bonos', 12),
            ('tipo_trab', 14),
            ('tipo_contrato', 22),
            ('fecha_alta', 14),
            ('fecha_inicio_contrato', 20),
            ('fecha_fin_contrato', 18),
            ('area', 20),
            ('subarea', 20),
            ('estado', 12),
            ('regimen_pension', 16),
            ('afp', 12),
            ('cuspp', 16),
            ('condicion', 12),
            ('grupo_tareo', 14),
            ('asignacion_familiar', 18),
            ('banco', 18),
            ('cuenta_ahorros', 20),
            ('cuenta_cci', 22),
            ('cuenta_cts', 20),
            ('regimen_turno', 14),
            ('jornada_horas', 14),
            ('codigo_fotocheck', 16),
            ('codigo_sap', 14),
            ('codigo_s10', 14),
            ('observaciones', 30),
        ]

        # Estilos
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        required_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=10)
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9'),
        )

        for col_idx, (header_name, width) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header_name)
            cell.font = header_font
            cell.fill = required_fill if '*' in header_name else header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Fila de ejemplo
        example_row = {
            'nro_doc *': '12345678',
            'tipo_doc': 'DNI',
            'apellido_paterno *': 'GARCIA',
            'apellido_materno *': 'LOPEZ',
            'nombres *': 'JUAN CARLOS',
            'cargo': 'Analista de Sistemas',
            'fecha_nacimiento': '15/03/1990',
            'sexo': 'M',
            'celular': '987654321',
            'email': 'juan.garcia@mail.com',
            'correo_corporativo': '',
            'direccion': 'Av. Arequipa 1234, Lima',
            'sueldo_base': '3500.00',
            'bonos': '500.00',
            'tipo_trab': 'Empleado',
            'tipo_contrato': 'PLAZO_FIJO',
            'fecha_alta': '01/02/2024',
            'fecha_inicio_contrato': '01/02/2024',
            'fecha_fin_contrato': '31/01/2025',
            'area': 'Tecnologia',
            'subarea': 'Desarrollo',
            'estado': 'Activo',
            'regimen_pension': 'AFP',
            'afp': 'Prima',
            'cuspp': '123456PRCBA00',
            'condicion': 'LIMA',
            'grupo_tareo': 'STAFF',
            'asignacion_familiar': 'Si',
            'banco': 'BCP',
            'cuenta_ahorros': '191-12345678-0-01',
            'cuenta_cci': '00219100123456780011',
            'cuenta_cts': '191-98765432-0-99',
            'regimen_turno': '',
            'jornada_horas': '8',
            'codigo_fotocheck': 'FC-001',
            'codigo_sap': '',
            'codigo_s10': '',
            'observaciones': '',
        }

        example_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        for col_idx, (header_name, _) in enumerate(headers, 1):
            val = example_row.get(header_name, '')
            cell = ws.cell(row=2, column=col_idx, value=val)
            cell.fill = example_fill
            cell.border = thin_border
            cell.font = Font(size=10, italic=True, color='808080')

        # Congelar encabezados
        ws.freeze_panes = 'A2'

        # ── Hoja de Catálogos ────────────────────────────────────────
        ws_cat = wb.create_sheet('Catalogos')
        ws_cat.sheet_properties.tabColor = 'FFC000'

        catalogs = {
            'tipo_doc': [c[0] for c in Personal.TIPO_DOC_CHOICES],
            'tipo_trab': [c[0] for c in Personal.TIPO_TRAB_CHOICES],
            'sexo': ['M', 'F'],
            'estado': [c[0] for c in Personal.ESTADO_CHOICES],
            'regimen_pension': [c[0] for c in Personal.REGIMEN_PENSION_CHOICES],
            'afp': [c[0] for c in Personal.AFP_CHOICES],
            'tipo_contrato': [c[0] for c in Personal.TIPO_CONTRATO_CHOICES],
            'condicion': [c[0] for c in Personal.CONDICION_CHOICES],
            'grupo_tareo': [c[0] for c in Personal.GRUPO_TAREO_CHOICES],
            'banco': [c[0] for c in Personal.BANCO_CHOICES],
        }

        cat_header_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        cat_col = 1
        for cat_name, values in catalogs.items():
            cell = ws_cat.cell(row=1, column=cat_col, value=cat_name)
            cell.font = Font(bold=True, size=10)
            cell.fill = cat_header_fill
            ws_cat.column_dimensions[get_column_letter(cat_col)].width = max(
                len(cat_name) + 4,
                max((len(str(v)) for v in values), default=8) + 4
            )
            for row_idx, val in enumerate(values, 2):
                ws_cat.cell(row=row_idx, column=cat_col, value=val)
            cat_col += 1

        # ── Validaciones dropdown en hoja Personal ────────────────────
        dropdown_fields = {
            'tipo_doc': 2,
            'tipo_trab': 15,
            'sexo': 8,
            'estado': 22,
            'regimen_pension': 23,
            'afp': 24,
            'tipo_contrato': 16,
            'condicion': 26,
            'grupo_tareo': 27,
            'banco': 29,
        }

        # Mapeo de catálogo a columna en Catalogos
        cat_col_map = {}
        cat_col_idx = 1
        for cat_name, values in catalogs.items():
            col_letter = get_column_letter(cat_col_idx)
            num_values = len(values)
            cat_col_map[cat_name] = f"Catalogos!${col_letter}$2:${col_letter}${num_values + 1}"
            cat_col_idx += 1

        for field_name, personal_col_idx in dropdown_fields.items():
            if field_name in cat_col_map:
                col_letter = get_column_letter(personal_col_idx)
                dv = DataValidation(
                    type='list',
                    formula1=cat_col_map[field_name],
                    allow_blank=True,
                    showErrorMessage=True,
                    errorTitle='Valor no valido',
                    error=f'Seleccione un valor de la lista para {field_name}'
                )
                dv.add(f'{col_letter}2:{col_letter}5001')
                ws.add_data_validation(dv)

        # Validación Si/No para asignacion_familiar (col 28)
        dv_bool = DataValidation(
            type='list',
            formula1='"Si,No"',
            allow_blank=True,
        )
        dv_bool.add('AB2:AB5001')
        ws.add_data_validation(dv_bool)

        # Guardar a bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()
