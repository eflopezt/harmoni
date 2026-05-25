"""
Genera plantilla Excel para que el cliente cargue el maestro de empleados.

Uso:
    python manage.py plantilla_maestro_empleados --output maestro_empleados_template.xlsx
    python manage.py plantilla_maestro_empleados --output ./out.xlsx --con-ejemplos

La plantilla tiene:
- Headers con colores corporativos Harmoni
- Validaciones de tipo (dropdowns) para campos enum (regimen pensión, tipo doc, etc.)
- Comentarios en celdas con instrucciones
- Hoja "Instrucciones" con guía completa
- Hoja "Catálogos" con valores válidos para que el cliente vea las opciones
- 2 filas de ejemplo (si --con-ejemplos)

Después de que el cliente la complete:
    python manage.py importar_maestro_empleados archivo.xlsx
"""
from pathlib import Path

from django.core.management.base import BaseCommand


# ── Colores marca Harmoni ──────────────────────────────────────────────
TEAL_DEEP = '042F2A'
TEAL_MAIN = '0F766E'
TEAL_LIGHT = '99F6E4'
TEAL_TINT = 'CCFBF1'
ACCENT_ORANGE = 'F59E0B'
INK = '0F172A'
GRAY = '94A3B8'
GRAY_SOFT = 'F1F5F9'
WHITE = 'FFFFFF'

# ── Columnas del maestro de empleados ──────────────────────────────────
# (header_visible, key_interna, tipo, requerido, ejemplo, ayuda)
COLUMNAS = [
    ('DNI / Carnet Ext.',         'nro_doc',           'text',  True,  '73000043', 'DNI peruano 8 dígitos o Carnet Extranjería'),
    ('Tipo Doc',                  'tipo_doc',          'enum',  True,  'DNI',      'DNI · CE · Pasaporte · RUC'),
    ('Apellidos y Nombres',       'apellidos_nombres', 'text',  True,  'PEREZ GARCIA, JUAN CARLOS', 'APELLIDOS PAT. MAT., NOMBRES (mayúsculas)'),
    ('Sexo',                      'sexo',              'enum',  True,  'M',        'M · F'),
    ('Fecha de nacimiento',       'fecha_nacimiento',  'date',  True,  '1990-05-15', 'Formato AAAA-MM-DD'),
    ('Celular',                   'celular',           'text',  False, '+51 987654321', 'Con código país +51'),
    ('Correo personal',           'correo_personal',   'email', False, 'juan@gmail.com', 'Para envío de boletas'),
    ('Dirección',                 'direccion',         'text',  False, 'Av. Javier Prado 1234, San Isidro', ''),

    # Empresa / cargo
    ('Empresa (RUC)',             'empresa_ruc',       'text',  True,  '20612345678', 'RUC al que pertenece (multi-empresa)'),
    ('Área',                      'area',              'text',  True,  'Cocina',   'Nombre del área. Crear si no existe.'),
    ('SubÁrea',                   'subarea',           'text',  False, 'Parrilla', 'Nombre subárea (opcional)'),
    ('Cargo',                     'cargo',             'text',  True,  'Cocinero', 'Cargo específico'),
    ('Grupo de Tareo',            'grupo_tareo',       'enum',  True,  'STAFF',    'STAFF (fijo) · RCO (rotativo construcción)'),
    ('Condición',                 'condicion',         'enum',  True,  'LOCAL',    'LOCAL · FORANEO · LIMA'),
    ('Categoría',                 'categoria',         'enum',  False, 'EMPLEADO', 'EMPLEADO · OBRERO · FUNCIONARIO · DIRECTIVO'),

    # Contrato
    ('Tipo contrato',             'tipo_contrato',     'enum',  True,  'INDEFINIDO', 'INDEFINIDO · PLAZO_FIJO · INTERMITENTE · EVENTUAL'),
    ('Fecha alta',                'fecha_alta',        'date',  True,  '2020-01-15', 'Fecha ingreso a la empresa'),
    ('Fecha fin contrato',        'fecha_fin_contrato','date',  False, '2026-12-31', 'Solo si es plazo fijo'),
    ('Régimen laboral',           'regimen_laboral',   'enum',  False, 'GENERAL',  'GENERAL · MICROEMPRESA · PEQUEÑA · CONSTRUCCION'),

    # Remuneración
    ('Sueldo base (S/)',          'sueldo_base',       'money', True,  '2500.00',  'Sueldo mensual base sin bonificaciones'),
    ('Asignación familiar',       'asignacion_familiar', 'bool', False, 'SI',      'SI si tiene hijos menores · NO'),
    ('Bonos fijos (S/)',          'bonos',             'money', False, '0.00',     'Bonos fijos mensuales (movilidad, alimentación)'),

    # Régimen pensión
    ('Régimen pensión',           'regimen_pension',   'enum',  True,  'AFP',      'AFP · ONP · SIN_PENSION'),
    ('AFP',                       'afp',               'enum',  False, 'INTEGRA',  'INTEGRA · PRIMA · PROFUTURO · HABITAT · null si ONP'),
    ('CUSPP',                     'cuspp',             'text',  False, '123456789012', 'Código único pensión AFP (12 dígitos)'),

    # Banco
    ('Banco depósito',            'banco',             'enum',  False, 'BCP',      'BCP · BBVA · SCOTIA · INTERBANK · BN · OTRO'),
    ('Cuenta ahorros',            'cuenta_ahorros',    'text',  False, '194-12345678-0-12', 'Para depósito de haberes'),
    ('CCI',                       'cuenta_cci',        'text',  False, '00219412345678012345', 'Cuenta interbancaria 20 dígitos'),
    ('Cuenta CTS',                'cuenta_cts',        'text',  False, '194-87654321-0-99', 'CTS separada (opcional)'),

    # Asistencia
    ('Jornada (horas/día)',       'jornada_horas',     'number', False, '8.00',    'Default 8h (Decreto Leg. 854)'),
    ('Régimen turno',             'regimen_turno',     'enum',  False, 'FIJO',     'FIJO · ROTATIVO · NOCHE · MADRUGADA · PARTIDO'),
    ('Día descanso semanal',      'dia_descanso_semanal', 'enum', False, 'DOM',    'LUN · MAR · MIE · JUE · VIE · SAB · DOM'),

    # EPS opcional
    ('Tiene EPS',                 'tiene_eps',         'bool',  False, 'NO',       'SI · NO'),
    ('Descuento EPS (S/)',        'eps_descuento_mensual', 'money', False, '0.00', 'Si aplica EPS particular'),

    # Códigos externos
    ('Código SAP',                'codigo_sap',        'text',  False, '',         'Si maneja codigo en sistema contable'),
    ('Código S10',                'codigo_s10',        'text',  False, '',         'Si proviene de S10 construcción civil'),
    ('Código fotocheck',          'codigo_fotocheck',  'text',  False, 'EDO-0042', 'Código interno fotocheck'),
]


# ── Valores válidos para dropdowns ─────────────────────────────────────
ENUMS = {
    'tipo_doc':         ['DNI', 'CE', 'PASAPORTE', 'RUC'],
    'sexo':             ['M', 'F'],
    'grupo_tareo':      ['STAFF', 'RCO'],
    'condicion':        ['LOCAL', 'FORANEO', 'LIMA'],
    'categoria':        ['EMPLEADO', 'OBRERO', 'FUNCIONARIO', 'DIRECTIVO'],
    'tipo_contrato':    ['INDEFINIDO', 'PLAZO_FIJO', 'INTERMITENTE', 'EVENTUAL', 'PRACTICAS'],
    'regimen_laboral':  ['GENERAL', 'MICROEMPRESA', 'PEQUEÑA', 'CONSTRUCCION', 'AGRARIO'],
    'regimen_pension':  ['AFP', 'ONP', 'SIN_PENSION'],
    'afp':              ['INTEGRA', 'PRIMA', 'PROFUTURO', 'HABITAT'],
    'banco':            ['BCP', 'BBVA', 'SCOTIA', 'INTERBANK', 'BN', 'PICHINCHA', 'CITIBANK', 'OTRO'],
    'regimen_turno':    ['FIJO', 'ROTATIVO', 'NOCHE', 'MADRUGADA', 'PARTIDO'],
    'dia_descanso_semanal': ['LUN', 'MAR', 'MIE', 'JUE', 'VIE', 'SAB', 'DOM'],
    'asignacion_familiar': ['SI', 'NO'],
    'tiene_eps':        ['SI', 'NO'],
}


# ── Filas de ejemplo (gastronomía coherente) ───────────────────────────
EJEMPLOS = [
    # Ejemplo 1: mozo STAFF con AFP
    {
        'nro_doc': '73000001', 'tipo_doc': 'DNI',
        'apellidos_nombres': 'GARCIA HUAMAN, MARIA ELENA',
        'sexo': 'F', 'fecha_nacimiento': '1995-03-12',
        'celular': '+51 987654321', 'correo_personal': 'maria.garcia@gmail.com',
        'direccion': 'Av. Brasil 1234, Pueblo Libre',
        'empresa_ruc': '20612345678',
        'area': 'Salón', 'subarea': 'Atención',
        'cargo': 'Mozo Senior', 'grupo_tareo': 'STAFF', 'condicion': 'LOCAL', 'categoria': 'EMPLEADO',
        'tipo_contrato': 'INDEFINIDO', 'fecha_alta': '2022-06-01',
        'regimen_laboral': 'GENERAL',
        'sueldo_base': '1800.00', 'asignacion_familiar': 'SI', 'bonos': '150.00',
        'regimen_pension': 'AFP', 'afp': 'INTEGRA', 'cuspp': '123456789012',
        'banco': 'BCP', 'cuenta_ahorros': '194-12345678-0-12',
        'cuenta_cci': '00219412345678012345', 'cuenta_cts': '',
        'jornada_horas': '8.00', 'regimen_turno': 'ROTATIVO', 'dia_descanso_semanal': 'LUN',
        'tiene_eps': 'NO', 'eps_descuento_mensual': '0.00',
        'codigo_sap': '', 'codigo_s10': '', 'codigo_fotocheck': 'EDO-0001',
    },
    # Ejemplo 2: cocinero ONP turno noche
    {
        'nro_doc': '73000002', 'tipo_doc': 'DNI',
        'apellidos_nombres': 'TORRES MAMANI, JORGE LUIS',
        'sexo': 'M', 'fecha_nacimiento': '1988-11-28',
        'celular': '+51 998877665', 'correo_personal': 'jorge.torres@hotmail.com',
        'direccion': 'Calle Los Pinos 567, Villa El Salvador',
        'empresa_ruc': '20612345678',
        'area': 'Cocina', 'subarea': 'Cocina caliente',
        'cargo': 'Cocinero', 'grupo_tareo': 'STAFF', 'condicion': 'LOCAL', 'categoria': 'EMPLEADO',
        'tipo_contrato': 'PLAZO_FIJO', 'fecha_alta': '2023-01-15', 'fecha_fin_contrato': '2026-12-31',
        'regimen_laboral': 'GENERAL',
        'sueldo_base': '2200.00', 'asignacion_familiar': 'SI', 'bonos': '200.00',
        'regimen_pension': 'ONP', 'afp': '', 'cuspp': '',
        'banco': 'BBVA', 'cuenta_ahorros': '0011-0123-0123456789',
        'cuenta_cci': '01101230123456789012', 'cuenta_cts': '',
        'jornada_horas': '8.00', 'regimen_turno': 'NOCHE', 'dia_descanso_semanal': 'MAR',
        'tiene_eps': 'NO', 'eps_descuento_mensual': '0.00',
        'codigo_sap': '', 'codigo_s10': '', 'codigo_fotocheck': 'EDO-0002',
    },
]


def _hex(color):
    """openpyxl quiere FF prefix para alpha"""
    return 'FF' + color


def build_plantilla(output_path: Path, con_ejemplos: bool = False):
    """Genera el xlsx en output_path."""
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, NamedStyle,
    )
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
    from openpyxl.comments import Comment

    wb = openpyxl.Workbook()
    # Borrar sheet default
    wb.remove(wb.active)

    # ── 1) HOJA INSTRUCCIONES ─────────────────────────────────────────
    ws_inst = wb.create_sheet('Instrucciones', 0)
    ws_inst.column_dimensions['A'].width = 110

    # Header con logo
    ws_inst['A1'] = 'PLANTILLA MAESTRO DE EMPLEADOS — HARMONI'
    ws_inst['A1'].font = Font(bold=True, size=18, color=_hex(WHITE))
    ws_inst['A1'].fill = PatternFill('solid', fgColor=_hex(TEAL_DEEP))
    ws_inst['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_inst.row_dimensions[1].height = 40

    instrucciones = [
        '',
        ('Esta plantilla sirve para cargar TODOS los empleados de su empresa a Harmoni en una sola operación.', 'normal'),
        '',
        ('CÓMO USARLA', 'h2'),
        ('1. Completar la hoja "EMPLEADOS" con un empleado por fila.', 'normal'),
        ('2. Las columnas marcadas con asterisco rojo (*) en el header son OBLIGATORIAS.', 'normal'),
        ('3. Para campos con valores predefinidos (régimen pensión, banco, etc.), ver la hoja "CATÁLOGOS" o usar los dropdowns dentro de las celdas.', 'normal'),
        ('4. NO cambie el orden de las columnas ni renombre los headers — el importador los reconoce por nombre.', 'normal'),
        ('5. Una vez completada, envíela a su contacto Harmoni o súbala desde Configuración → Importar Maestro.', 'normal'),
        '',
        ('REGLAS DE LLENADO', 'h2'),
        ('• Fechas en formato AAAA-MM-DD (ej. 2026-05-25), NO usar DD/MM/AAAA', 'normal'),
        ('• Montos sin símbolo S/ y con punto decimal (ej. 1500.00, no S/ 1,500.00)', 'normal'),
        ('• Apellidos y Nombres en MAYÚSCULAS, formato "PATERNO MATERNO, NOMBRES"', 'normal'),
        ('• DNI: 8 dígitos peruanos. Si es extranjero, use Carnet Extranjería + Tipo Doc = CE', 'normal'),
        ('• RUC empresa: 11 dígitos. Debe corresponder a una empresa ya creada en su Harmoni', 'normal'),
        ('• Asignación familiar: SI solo si tiene hijos menores de 18 (o hasta 24 si estudia)', 'normal'),
        ('• CUSPP: 12 caracteres alfanuméricos de AFP. Si está en ONP, dejar vacío', 'normal'),
        ('• CCI: 20 dígitos (Cuenta Interbancaria), para depósitos masivos', 'normal'),
        '',
        ('VALIDACIÓN AUTOMÁTICA', 'h2'),
        ('Al importar, Harmoni validará:', 'normal'),
        ('  – Que el DNI no esté duplicado', 'normal'),
        ('  – Que el RUC de la empresa exista', 'normal'),
        ('  – Que las áreas y subáreas existan (si no, las crea automáticamente)', 'normal'),
        ('  – Formato correcto de fechas, montos, AFP, etc.', 'normal'),
        ('  – CCI/CCI ahorros cumpla con check digit', 'normal'),
        ('Cualquier error se reporta en un Excel adjunto con el detalle por fila.', 'normal'),
        '',
        ('SOPORTE', 'h2'),
        ('Si tienes dudas durante el llenado:', 'normal'),
        ('  WhatsApp: +51 906 602 313 (Edwin López — Gerente de Operaciones)', 'normal'),
        ('  Email:    elopez@harmoni.pe', 'normal'),
        ('  Demo:     https://demo.harmoni.pe (demo / demo123)', 'normal'),
    ]

    row = 2
    for item in instrucciones:
        if isinstance(item, tuple):
            text, kind = item
        else:
            text, kind = item, 'normal'

        ws_inst.cell(row=row, column=1, value=text)
        cell = ws_inst.cell(row=row, column=1)
        if kind == 'h2':
            cell.font = Font(bold=True, size=12, color=_hex(TEAL_MAIN))
            ws_inst.row_dimensions[row].height = 22
        else:
            cell.font = Font(size=10, color=_hex(INK))
        row += 1

    # ── 2) HOJA EMPLEADOS ─────────────────────────────────────────────
    ws = wb.create_sheet('EMPLEADOS', 1)

    # Title row
    ws['A1'] = 'MAESTRO DE EMPLEADOS — Complete una fila por trabajador'
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNAS))
    ws['A1'].font = Font(bold=True, size=12, color=_hex(WHITE))
    ws['A1'].fill = PatternFill('solid', fgColor=_hex(TEAL_DEEP))
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 26

    # Headers fila 2
    header_fill = PatternFill('solid', fgColor=_hex(TEAL_MAIN))
    header_font = Font(bold=True, size=10, color=_hex(WHITE))
    header_req_font = Font(bold=True, size=10, color=_hex(WHITE))
    border_thin = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1'),
    )

    for col_idx, (header, key, ftype, required, ejemplo, ayuda) in enumerate(COLUMNAS, 1):
        cell = ws.cell(row=2, column=col_idx, value=header + (' *' if required else ''))
        cell.fill = header_fill
        cell.font = header_req_font if required else header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border_thin

        # Comment con ayuda
        comment_text = (ayuda or '') + (f'\n\nEjemplo: {ejemplo}' if ejemplo else '')
        if required:
            comment_text = '⚠ CAMPO OBLIGATORIO\n\n' + comment_text
        cell.comment = Comment(comment_text, 'Harmoni')

        # Ancho columna
        if ftype == 'text' and key == 'apellidos_nombres':
            ws.column_dimensions[get_column_letter(col_idx)].width = 32
        elif ftype == 'text' and key == 'direccion':
            ws.column_dimensions[get_column_letter(col_idx)].width = 28
        elif ftype == 'email':
            ws.column_dimensions[get_column_letter(col_idx)].width = 24
        elif ftype == 'date':
            ws.column_dimensions[get_column_letter(col_idx)].width = 13
        elif ftype == 'money' or ftype == 'number':
            ws.column_dimensions[get_column_letter(col_idx)].width = 12
        else:
            ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(header) + 2)

    ws.row_dimensions[2].height = 36

    # ── Validaciones (dropdowns) ───────────────────────────────────────
    for col_idx, (_, key, ftype, _req, _ex, _ay) in enumerate(COLUMNAS, 1):
        if ftype == 'enum' and key in ENUMS:
            valid_values = ENUMS[key]
            # openpyxl: formula1 con comillas para lista
            dv = DataValidation(
                type='list',
                formula1=f'"{",".join(valid_values)}"',
                allow_blank=True,
                error='Valor no válido',
                errorTitle='Error',
                prompt=f'Valores permitidos:\n{", ".join(valid_values)}',
                promptTitle=key,
            )
            dv.add(f'{get_column_letter(col_idx)}3:{get_column_letter(col_idx)}1000')
            ws.add_data_validation(dv)

    # ── Filas de ejemplo ───────────────────────────────────────────────
    row_idx = 3
    if con_ejemplos:
        for ejemplo in EJEMPLOS:
            for col_idx, (_, key, ftype, _req, _ex, _ay) in enumerate(COLUMNAS, 1):
                val = ejemplo.get(key, '')
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = border_thin
                if ftype == 'money' or ftype == 'number':
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                elif ftype == 'date':
                    cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill('solid', fgColor=_hex(TEAL_TINT))
                cell.font = Font(italic=True, size=9, color=_hex(GRAY))
            # Marcar con un comentario "EJEMPLO"
            ws.cell(row=row_idx, column=1).comment = Comment(
                f'EJEMPLO {row_idx-2} — Eliminar antes de importar', 'Harmoni'
            )
            row_idx += 1

    # Freeze panes en la primera fila editable
    ws.freeze_panes = f'A{row_idx if con_ejemplos else 3}'

    # ── 3) HOJA CATÁLOGOS ─────────────────────────────────────────────
    ws_cat = wb.create_sheet('CATÁLOGOS', 2)
    ws_cat['A1'] = 'VALORES VÁLIDOS PARA CADA COLUMNA CON DROPDOWN'
    ws_cat.merge_cells('A1:C1')
    ws_cat['A1'].font = Font(bold=True, size=12, color=_hex(WHITE))
    ws_cat['A1'].fill = PatternFill('solid', fgColor=_hex(TEAL_DEEP))
    ws_cat['A1'].alignment = Alignment(horizontal='center')
    ws_cat.row_dimensions[1].height = 26

    cat_headers = ['Campo', 'Valores válidos', 'Notas']
    for col_idx, h in enumerate(cat_headers, 1):
        cell = ws_cat.cell(row=2, column=col_idx, value=h)
        cell.font = Font(bold=True, color=_hex(WHITE))
        cell.fill = PatternFill('solid', fgColor=_hex(TEAL_MAIN))
        cell.alignment = Alignment(horizontal='center')

    notas_catalogo = {
        'tipo_doc':         'Si es Carnet Extranjería, use el código en columna DNI',
        'sexo':             'M = Masculino, F = Femenino',
        'grupo_tareo':      'STAFF para personal fijo (salón, cocina, admin) · RCO para personal por obra',
        'condicion':        'LOCAL trabaja en su sede · FORANEO viaja entre sedes · LIMA central',
        'tipo_contrato':    'EVENTUAL para personal de eventos · INTERMITENTE para temporadas',
        'regimen_laboral':  'Default GENERAL. MICROEMPRESA/PEQUEÑA si está acogido a régimen MYPE',
        'regimen_pension':  'AFP requiere CUSPP y nombre de AFP · ONP sin CUSPP',
        'afp':              'Solo las 4 AFPs vigentes en Perú (Mayo 2026)',
        'banco':            'Banco principal para depósito de haberes',
        'regimen_turno':    'NOCHE = trabajadores con jornada nocturna 22:00-06:00 (recargo 35% del valor hora)',
        'asignacion_familiar': '10% de RMV (S/ 102.50 a Mayo 2026) si tiene hijo(s) menor(es) de 18',
    }
    row = 3
    ws_cat.column_dimensions['A'].width = 22
    ws_cat.column_dimensions['B'].width = 65
    ws_cat.column_dimensions['C'].width = 60
    for key, valores in ENUMS.items():
        ws_cat.cell(row=row, column=1, value=key)
        ws_cat.cell(row=row, column=1).font = Font(bold=True, color=_hex(TEAL_MAIN))
        ws_cat.cell(row=row, column=2, value=' · '.join(valores))
        ws_cat.cell(row=row, column=3, value=notas_catalogo.get(key, ''))
        ws_cat.cell(row=row, column=3).alignment = Alignment(wrap_text=True)
        row += 1

    # Guardar
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


class Command(BaseCommand):
    help = 'Genera plantilla Excel para que el cliente cargue el maestro de empleados'

    def add_arguments(self, parser):
        parser.add_argument(
            '--output', type=str,
            default='plantilla_maestro_empleados_HARMONI.xlsx',
            help='Path de salida (default: ./plantilla_maestro_empleados_HARMONI.xlsx)',
        )
        parser.add_argument(
            '--con-ejemplos', action='store_true',
            help='Incluir 2 filas de ejemplo en la plantilla',
        )

    def handle(self, *args, **opts):
        output = Path(opts['output']).resolve()
        con_ejemplos = opts['con_ejemplos']

        build_plantilla(output, con_ejemplos=con_ejemplos)

        size = output.stat().st_size
        self.stdout.write(self.style.SUCCESS(
            f'\nPlantilla generada: {output}\n'
            f'   Tamaño: {size:,} bytes\n'
            f'   Hojas:  Instrucciones · EMPLEADOS · CATÁLOGOS\n'
            f'   Columnas: {len(COLUMNAS)} (con validaciones y comentarios)\n'
            f'   Ejemplos: {len(EJEMPLOS) if con_ejemplos else 0}\n'
        ))
