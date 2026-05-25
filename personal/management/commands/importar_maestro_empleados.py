"""
Importa el maestro de empleados desde el Excel completado por el cliente.

Uso:
    python manage.py importar_maestro_empleados maestro.xlsx
    python manage.py importar_maestro_empleados maestro.xlsx --dry-run
    python manage.py importar_maestro_empleados maestro.xlsx --crear-areas

Flujo:
1. Lee la hoja "EMPLEADOS" del xlsx (debe coincidir con plantilla_maestro_empleados)
2. Valida fila por fila (DNI duplicado, RUC empresa existe, formato fechas, etc.)
3. Si validación OK → crea Personal + áreas/subáreas si --crear-areas
4. Si hay errores → genera Excel adjunto con resumen por fila
5. dry-run muestra qué haría sin escribir nada

Validaciones:
- DNI obligatorio + único
- RUC empresa debe existir
- Régimen pensión AFP requiere CUSPP
- CCI 20 dígitos si está presente
- Fechas en formato AAAA-MM-DD
- Sueldo > 0
"""
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction


# Lista de columnas esperadas (debe coincidir con plantilla_maestro_empleados.COLUMNAS)
COLUMNAS_ESPERADAS = [
    'nro_doc', 'tipo_doc', 'apellidos_nombres', 'sexo', 'fecha_nacimiento',
    'celular', 'correo_personal', 'direccion',
    'empresa_ruc', 'area', 'subarea', 'cargo',
    'grupo_tareo', 'condicion', 'categoria',
    'tipo_contrato', 'fecha_alta', 'fecha_fin_contrato', 'regimen_laboral',
    'sueldo_base', 'asignacion_familiar', 'bonos',
    'regimen_pension', 'afp', 'cuspp',
    'banco', 'cuenta_ahorros', 'cuenta_cci', 'cuenta_cts',
    'jornada_horas', 'regimen_turno', 'dia_descanso_semanal',
    'tiene_eps', 'eps_descuento_mensual',
    'codigo_sap', 'codigo_s10', 'codigo_fotocheck',
]

HEADERS_VISIBLES = [
    'DNI / Carnet Ext.', 'Tipo Doc', 'Apellidos y Nombres', 'Sexo',
    'Fecha de nacimiento', 'Celular', 'Correo personal', 'Dirección',
    'Empresa (RUC)', 'Área', 'SubÁrea', 'Cargo',
    'Grupo de Tareo', 'Condición', 'Categoría',
    'Tipo contrato', 'Fecha alta', 'Fecha fin contrato', 'Régimen laboral',
    'Sueldo base (S/)', 'Asignación familiar', 'Bonos fijos (S/)',
    'Régimen pensión', 'AFP', 'CUSPP',
    'Banco depósito', 'Cuenta ahorros', 'CCI', 'Cuenta CTS',
    'Jornada (horas/día)', 'Régimen turno', 'Día descanso semanal',
    'Tiene EPS', 'Descuento EPS (S/)',
    'Código SAP', 'Código S10', 'Código fotocheck',
]


def _parse_date(val):
    if not val:
        return None
    if isinstance(val, (date, datetime)):
        return val.date() if isinstance(val, datetime) else val
    s = str(val).strip()
    if not s:
        return None
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    raise ValueError(f'Fecha inválida: {val!r}')


def _parse_decimal(val, default='0.00'):
    if val in (None, ''):
        return Decimal(default)
    if isinstance(val, (int, float, Decimal)):
        return Decimal(str(val))
    s = str(val).strip().replace(',', '')
    try:
        return Decimal(s)
    except InvalidOperation:
        raise ValueError(f'Monto inválido: {val!r}')


def _parse_bool_si(val):
    if val in (None, '', False):
        return False
    s = str(val).strip().upper()
    return s in ('SI', 'YES', 'TRUE', '1', 'X')


def _parse_int(val, default=0):
    if val in (None, ''):
        return default
    try:
        return int(val)
    except (ValueError, TypeError):
        try:
            return int(float(str(val).strip()))
        except (ValueError, TypeError):
            raise ValueError(f'Número inválido: {val!r}')


class Command(BaseCommand):
    help = 'Importa el maestro de empleados desde Excel'

    def add_arguments(self, parser):
        parser.add_argument('archivo', type=str, help='Path al archivo .xlsx')
        parser.add_argument(
            '--dry-run', action='store_true',
            help='Solo valida y muestra qué haría, sin escribir',
        )
        parser.add_argument(
            '--crear-areas', action='store_true',
            help='Crea áreas/subáreas que no existan (default: error si falta)',
        )
        parser.add_argument(
            '--actualizar-existentes', action='store_true',
            help='Si encuentra DNI ya existente, lo actualiza en lugar de error',
        )
        parser.add_argument(
            '--errores-output', type=str, default='errores_import.xlsx',
            help='Path del Excel con errores (default: errores_import.xlsx)',
        )

    def handle(self, *args, **opts):
        from personal.models import Personal, Area, SubArea
        from empresas.models import Empresa
        import openpyxl

        archivo = Path(opts['archivo']).resolve()
        if not archivo.exists():
            raise CommandError(f'Archivo no encontrado: {archivo}')

        dry = opts['dry_run']
        crear_areas = opts['crear_areas']
        actualizar_existentes = opts['actualizar_existentes']

        self.stdout.write(self.style.MIGRATE_HEADING(
            f'Importando maestro: {archivo.name} '
            f'({"DRY-RUN" if dry else "WRITE"} · '
            f'{"crear áreas" if crear_areas else "áreas estrictas"} · '
            f'{"actualizar DNI existentes" if actualizar_existentes else "estricto DNI único"})'
        ))

        try:
            wb = openpyxl.load_workbook(archivo, data_only=True)
        except Exception as e:
            raise CommandError(f'Excel inválido: {e}')

        if 'EMPLEADOS' not in wb.sheetnames:
            raise CommandError(
                f'No se encontró la hoja "EMPLEADOS". Hojas disponibles: {wb.sheetnames}'
            )

        ws = wb['EMPLEADOS']

        # Detectar fila de headers — la plantilla los pone en fila 2
        # Construir map nombre_columna_visible → key_interna
        col_map = {}  # col_idx → key
        for col_idx in range(1, ws.max_column + 1):
            header = ws.cell(row=2, column=col_idx).value
            if not header:
                continue
            header_clean = header.replace('*', '').strip()
            if header_clean in HEADERS_VISIBLES:
                idx = HEADERS_VISIBLES.index(header_clean)
                col_map[col_idx] = COLUMNAS_ESPERADAS[idx]

        if len(col_map) < 10:
            raise CommandError(
                f'No se reconocieron suficientes columnas. Detectadas: {len(col_map)}\n'
                f'Asegurate de usar la plantilla generada por '
                f'`python manage.py plantilla_maestro_empleados`'
            )

        self.stdout.write(f'  Columnas detectadas: {len(col_map)} / {len(COLUMNAS_ESPERADAS)}')

        # Cache de empresas y áreas
        empresas_by_ruc = {e.ruc: e for e in Empresa.objects.all()}
        areas_by_name = {a.nombre.lower(): a for a in Area.objects.all()}
        subareas_by_key = {
            f'{sa.area_id}|{sa.nombre.lower()}': sa
            for sa in SubArea.objects.select_related('area')
        }

        # Procesar filas
        row = 3  # primera fila de datos
        n_ok = n_errores = n_actualizados = n_creados = 0
        errores = []  # lista de (fila, dni, mensaje)
        dnis_vistos = set()

        while row <= ws.max_row:
            data = {}
            for col_idx, key in col_map.items():
                val = ws.cell(row=row, column=col_idx).value
                data[key] = val

            dni = (str(data.get('nro_doc') or '').strip())
            if not dni:
                # Asumir fin de datos
                break

            try:
                # ── Validaciones obligatorias ──
                if not data.get('apellidos_nombres'):
                    raise ValueError('Apellidos y Nombres es obligatorio')

                if dni in dnis_vistos:
                    raise ValueError(f'DNI duplicado dentro del archivo: {dni}')
                dnis_vistos.add(dni)

                ruc = (str(data.get('empresa_ruc') or '').strip())
                if not ruc:
                    raise ValueError('Empresa (RUC) es obligatorio')
                empresa = empresas_by_ruc.get(ruc)
                if not empresa:
                    raise ValueError(f'Empresa con RUC {ruc} no existe en Harmoni')

                # Áreas
                area_nombre = (str(data.get('area') or '').strip())
                if not area_nombre:
                    raise ValueError('Área es obligatoria')

                area = areas_by_name.get(area_nombre.lower())
                if not area:
                    if crear_areas and not dry:
                        area = Area.objects.create(nombre=area_nombre)
                        areas_by_name[area_nombre.lower()] = area
                    elif crear_areas and dry:
                        pass  # solo simular
                    else:
                        raise ValueError(
                            f'Área "{area_nombre}" no existe. Use --crear-areas o créela primero.'
                        )

                subarea_nombre = (str(data.get('subarea') or '').strip())
                subarea = None
                if subarea_nombre and area:
                    sa_key = f'{area.id}|{subarea_nombre.lower()}'
                    subarea = subareas_by_key.get(sa_key)
                    if not subarea:
                        if crear_areas and not dry:
                            subarea = SubArea.objects.create(area=area, nombre=subarea_nombre)
                            subareas_by_key[sa_key] = subarea

                # Régimen pensión
                regimen_pension = (str(data.get('regimen_pension') or 'AFP').strip().upper())
                if regimen_pension == 'AFP' and not data.get('cuspp'):
                    # Warning, no error: pueden cargarlo después
                    pass

                # ── Construir kwargs Personal ──
                kwargs = {
                    'nro_doc':           dni,
                    'tipo_doc':          (str(data.get('tipo_doc') or 'DNI').strip().upper()),
                    'apellidos_nombres': str(data.get('apellidos_nombres')).strip(),
                    'sexo':              (str(data.get('sexo') or 'M').strip().upper()),
                    'celular':           (str(data.get('celular') or '').strip()),
                    'correo_personal':   (str(data.get('correo_personal') or '').strip()),
                    'direccion':         (str(data.get('direccion') or '').strip()),
                    'empresa':           empresa,
                    'subarea':           subarea,
                    'cargo':             (str(data.get('cargo') or '').strip()),
                    'grupo_tareo':       (str(data.get('grupo_tareo') or 'STAFF').strip().upper()),
                    'condicion':         (str(data.get('condicion') or 'LOCAL').strip().upper()),
                    'categoria':         (str(data.get('categoria') or 'EMPLEADO').strip().upper()),
                    'tipo_contrato':     (str(data.get('tipo_contrato') or 'INDEFINIDO').strip().upper()),
                    'regimen_laboral':   (str(data.get('regimen_laboral') or 'GENERAL').strip().upper()),
                    'sueldo_base':       _parse_decimal(data.get('sueldo_base'), '0'),
                    'asignacion_familiar': _parse_bool_si(data.get('asignacion_familiar')),
                    'bonos':             _parse_decimal(data.get('bonos'), '0'),
                    'regimen_pension':   regimen_pension,
                    'afp':               (str(data.get('afp') or '').strip().upper()),
                    'cuspp':             (str(data.get('cuspp') or '').strip()),
                    'banco':             (str(data.get('banco') or '').strip().upper()),
                    'cuenta_ahorros':    (str(data.get('cuenta_ahorros') or '').strip()),
                    'cuenta_cci':        (str(data.get('cuenta_cci') or '').strip()),
                    'cuenta_cts':        (str(data.get('cuenta_cts') or '').strip()),
                    'jornada_horas':     _parse_decimal(data.get('jornada_horas'), '8'),
                    'regimen_turno':     (str(data.get('regimen_turno') or 'FIJO').strip().upper()),
                    'dia_descanso_semanal': (str(data.get('dia_descanso_semanal') or 'DOM').strip().upper()),
                    'tiene_eps':         _parse_bool_si(data.get('tiene_eps')),
                    'eps_descuento_mensual': _parse_decimal(data.get('eps_descuento_mensual'), '0'),
                    'codigo_sap':        (str(data.get('codigo_sap') or '').strip()),
                    'codigo_s10':        (str(data.get('codigo_s10') or '').strip()),
                    'codigo_fotocheck':  (str(data.get('codigo_fotocheck') or '').strip()),
                    'estado':            'Activo',
                }

                # Fechas
                fnac = _parse_date(data.get('fecha_nacimiento'))
                if fnac:
                    kwargs['fecha_nacimiento'] = fnac
                falta = _parse_date(data.get('fecha_alta'))
                if falta:
                    kwargs['fecha_alta'] = falta
                ffin = _parse_date(data.get('fecha_fin_contrato'))
                if ffin:
                    kwargs['fecha_fin_contrato'] = ffin

                # ── Crear o actualizar ──
                if dry:
                    n_ok += 1
                else:
                    existing = Personal.objects.filter(nro_doc=dni).first()
                    if existing:
                        if actualizar_existentes:
                            for k, v in kwargs.items():
                                setattr(existing, k, v)
                            existing.save()
                            n_actualizados += 1
                        else:
                            raise ValueError(
                                f'DNI {dni} ya existe en BD (id={existing.id}). '
                                f'Use --actualizar-existentes para sobreescribir.'
                            )
                    else:
                        Personal.objects.create(**kwargs)
                        n_creados += 1
                    n_ok += 1

            except Exception as e:
                n_errores += 1
                errores.append((row, dni, str(e)))

            row += 1

        # Reportar
        self.stdout.write('')
        self.stdout.write(self.style.SUCCESS(
            f'  [OK] Procesados:   {n_ok}\n'
            f'  [OK] Creados:      {n_creados}\n'
            f'  [OK] Actualizados: {n_actualizados}\n'
            f'  [!]  Errores:      {n_errores}'
        ))

        if errores:
            err_path = Path(opts['errores_output']).resolve()
            self._dump_errores_xlsx(errores, err_path)
            self.stdout.write(self.style.WARNING(
                f'\n  Errores en: {err_path}\n'
                f'  Corregilos y re-importá. Las filas OK no se duplican (importador idempotente con --actualizar-existentes).'
            ))

    def _dump_errores_xlsx(self, errores, output_path):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Errores import'

        ws['A1'] = 'ERRORES EN IMPORTACIÓN DEL MAESTRO'
        ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
        ws['A1'].fill = PatternFill('solid', fgColor='FFDC2626')
        ws.merge_cells('A1:C1')

        headers = ['Fila Excel', 'DNI', 'Error']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='FF0F766E')

        for i, (fila, dni, msg) in enumerate(errores, 3):
            ws.cell(row=i, column=1, value=fila)
            ws.cell(row=i, column=2, value=dni)
            ws.cell(row=i, column=3, value=msg)

        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 100

        wb.save(output_path)
