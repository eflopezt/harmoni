"""
Genera plantilla Excel para cargar saldos al cierre del año previo (vacaciones,
CTS, gratificación pendiente). Es el "estado inicial" del trabajador en Harmoni.

Uso:
    python manage.py plantilla_saldos_cierre --output saldos_2025.xlsx --anio 2025
    python manage.py plantilla_saldos_cierre --output saldos.xlsx --anio 2025 --con-ejemplos
"""
from pathlib import Path

from django.core.management.base import BaseCommand


# Colores marca
TEAL_DEEP = '042F2A'
TEAL_MAIN = '0F766E'
TEAL_TINT = 'CCFBF1'
ACCENT = 'F59E0B'
INK = '0F172A'
GRAY = '94A3B8'
WHITE = 'FFFFFF'


# (header, key, tipo, requerido, ejemplo, ayuda)
COLUMNAS = [
    ('DNI',                            'nro_doc', 'text', True, '73000043',
     'DNI del trabajador (debe existir en el maestro ya cargado)'),
    ('Apellidos y Nombres',            'apellidos_nombres', 'text', False, 'PEREZ GARCIA, JUAN CARLOS',
     'Solo referencia visual — el sistema cruza por DNI'),

    # ── Vacaciones ───────────────────────────────────────────────────
    ('Vacaciones — días ganados',      'vac_dias_ganados', 'number', True, '30',
     'Total días vacacionales ganados al cierre del año'),
    ('Vacaciones — días gozados',      'vac_dias_gozados', 'number', True, '15',
     'Días ya tomados/gozados al cierre del año'),
    ('Vacaciones — saldo (días)',      'vac_saldo_dias', 'number', True, '15',
     'Días pendientes de gozar (ganados - gozados)'),
    ('Vacaciones — record desde',      'vac_record_desde', 'date', False, '2024-01-15',
     'Fecha de inicio del récord vacacional vigente'),

    # ── CTS ──────────────────────────────────────────────────────────
    ('CTS — saldo acumulado (S/)',     'cts_saldo', 'money', True, '4500.00',
     'CTS acumulada al cierre, sin depositar todavía'),
    ('CTS — días computables',         'cts_dias_computables', 'number', False, '270',
     'Días computables hasta la fecha de cierre'),
    ('CTS — última base remunerativa', 'cts_base_remunerativa', 'money', False, '1800.00',
     'Sueldo + 1/6 gratif que se usó en último cálculo'),
    ('CTS — entidad depositaria',      'cts_entidad', 'enum', False, 'BCP',
     'Banco donde se deposita la CTS'),

    # ── Gratificaciones ───────────────────────────────────────────────
    ('Gratif Julio — saldo (S/)',      'grat_julio_saldo', 'money', False, '0.00',
     'Saldo pendiente de gratif. julio (si quedó algo)'),
    ('Gratif Diciembre — saldo (S/)',  'grat_dic_saldo', 'money', False, '0.00',
     'Saldo pendiente de gratif. diciembre (si quedó algo)'),
    ('Bonif. extraordinaria 9% saldo', 'bonif_extra_saldo', 'money', False, '0.00',
     'Saldo pendiente Ley 29351 (9% sobre gratif.)'),

    # ── Préstamos pendientes ─────────────────────────────────────────
    ('Préstamos — saldo pendiente',    'prestamos_saldo', 'money', False, '0.00',
     'Total préstamos que se deben descontar de planillas futuras'),
    ('Préstamos — cuota mensual',      'prestamos_cuota', 'money', False, '0.00',
     'Cuota mensual a descontar (debe respetar 30% RMV intangible)'),
    ('Préstamos — cuotas restantes',   'prestamos_cuotas', 'number', False, '0',
     'Número de cuotas mensuales restantes'),

    # ── Banco de Horas (solo STAFF) ──────────────────────────────────
    ('Banco horas — saldo (hrs)',      'banco_horas_saldo', 'number', False, '0.00',
     'Solo STAFF: saldo en banco de horas extras compensatorias'),

    # ── Otros saldos ──────────────────────────────────────────────────
    ('Adelantos no descontados (S/)',  'adelantos_saldo', 'money', False, '0.00',
     'Adelantos otorgados que aún no se descuentan'),
    ('Otros descuentos pendientes',    'otros_desc_pend', 'money', False, '0.00',
     'Embargos, retenciones judiciales, cuota sindical, etc.'),

    # Notas
    ('Observaciones',                  'observaciones', 'text', False, '',
     'Cualquier nota relevante para migración (ej. trabajador en licencia, etc.)'),
]

ENUMS = {
    'cts_entidad': ['BCP', 'BBVA', 'SCOTIA', 'INTERBANK', 'BN', 'MIBANCO', 'CITIBANK', 'OTRO'],
}

EJEMPLOS = [
    {
        'nro_doc': '73000001',
        'apellidos_nombres': 'GARCIA HUAMAN, MARIA ELENA',
        'vac_dias_ganados': 30, 'vac_dias_gozados': 15, 'vac_saldo_dias': 15,
        'vac_record_desde': '2024-06-01',
        'cts_saldo': '5400.00', 'cts_dias_computables': 365,
        'cts_base_remunerativa': '1950.00', 'cts_entidad': 'BCP',
        'grat_julio_saldo': '0.00', 'grat_dic_saldo': '0.00', 'bonif_extra_saldo': '0.00',
        'prestamos_saldo': '600.00', 'prestamos_cuota': '100.00', 'prestamos_cuotas': 6,
        'banco_horas_saldo': '12.50',
        'adelantos_saldo': '0.00', 'otros_desc_pend': '0.00',
        'observaciones': 'Vacaciones programadas para enero 2026',
    },
    {
        'nro_doc': '73000002',
        'apellidos_nombres': 'TORRES MAMANI, JORGE LUIS',
        'vac_dias_ganados': 30, 'vac_dias_gozados': 30, 'vac_saldo_dias': 0,
        'vac_record_desde': '2025-01-15',
        'cts_saldo': '3200.00', 'cts_dias_computables': 245,
        'cts_base_remunerativa': '2350.00', 'cts_entidad': 'BBVA',
        'grat_julio_saldo': '0.00', 'grat_dic_saldo': '150.00', 'bonif_extra_saldo': '13.50',
        'prestamos_saldo': '0.00', 'prestamos_cuota': '0.00', 'prestamos_cuotas': 0,
        'banco_horas_saldo': '0.00',
        'adelantos_saldo': '0.00', 'otros_desc_pend': '0.00',
        'observaciones': '',
    },
]


def _hex(c):
    return 'FF' + c


def build_plantilla(output_path: Path, anio: int, con_ejemplos: bool = False):
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
    from openpyxl.comments import Comment

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Hoja Instrucciones ──
    ws_inst = wb.create_sheet('Instrucciones', 0)
    ws_inst.column_dimensions['A'].width = 110

    ws_inst['A1'] = f'SALDOS AL CIERRE {anio} — HARMONI'
    ws_inst['A1'].font = Font(bold=True, size=18, color=_hex(WHITE))
    ws_inst['A1'].fill = PatternFill('solid', fgColor=_hex(TEAL_DEEP))
    ws_inst['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_inst.row_dimensions[1].height = 40

    blocks = [
        '',
        (f'Esta plantilla carga los saldos finales del año {anio} para cada trabajador. Es el "punto de partida" de su Harmoni — sin esto, vacaciones y CTS arrancarían en cero.', 'normal'),
        '',
        ('PREREQUISITO', 'h2'),
        ('El maestro de empleados YA debe estar cargado en Harmoni. El importador cruza por DNI — si un DNI no existe, esa fila se ignora con error legible.', 'normal'),
        '',
        ('CÓMO USARLA', 'h2'),
        ('1. Una fila por trabajador, identificado por DNI.', 'normal'),
        ('2. Complete solo los saldos que aplican; los demás pueden quedar en 0.', 'normal'),
        ('3. Las columnas con (*) son obligatorias.', 'normal'),
        ('4. Envíela a su contacto Harmoni o cárguela vía Configuración → Importar Saldos.', 'normal'),
        '',
        ('REGLAS DE LLENADO', 'h2'),
        ('• Vacaciones: complete días ganados, gozados y saldo. Validación: ganados − gozados = saldo.', 'normal'),
        ('• CTS: monto en S/ que tiene el trabajador acumulado pero AÚN no depositado al cierre.', 'normal'),
        ('• Préstamos: cuota mensual no puede exceder 30% del RMV intangible (Ley 9463).', 'normal'),
        ('• Banco de horas: solo para personal STAFF (rotativo no compensa, paga HE directo).', 'normal'),
        ('• Adelantos: si el trabajador recibió un adelanto a cuenta de planilla y aún no fue descontado.', 'normal'),
        '',
        ('VALIDACIÓN AUTOMÁTICA', 'h2'),
        ('Al importar, Harmoni:', 'normal'),
        ('  – Verifica que el DNI exista en el maestro', 'normal'),
        ('  – Suma "saldo_vac = ganados − gozados" y avisa si no coincide', 'normal'),
        ('  – Crea SaldoApertura por trabajador (visible en su legajo)', 'normal'),
        ('  – Carga préstamos como Prestamo activo con cuotas pendientes', 'normal'),
        ('  – Inicializa BancoHoras del mes de corte', 'normal'),
        '',
        ('SOPORTE', 'h2'),
        ('  WhatsApp:  +51 906 602 313 (Edwin López)', 'normal'),
        ('  Email:     elopez@harmoni.pe', 'normal'),
    ]

    row = 2
    for item in blocks:
        if isinstance(item, tuple):
            text, kind = item
        else:
            text, kind = item, 'normal'
        ws_inst.cell(row=row, column=1, value=text)
        cell = ws_inst.cell(row=row, column=1)
        if kind == 'h2':
            cell.font = Font(bold=True, size=12, color=_hex(TEAL_MAIN))
            ws_inst.row_dimensions[row].height = 22
        else:
            cell.font = Font(size=10, color=_hex(INK))
        row += 1

    # ── Hoja SALDOS ──
    ws = wb.create_sheet('SALDOS', 1)
    ws['A1'] = f'SALDOS AL CIERRE {anio} — Una fila por trabajador'
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNAS))
    ws['A1'].font = Font(bold=True, size=12, color=_hex(WHITE))
    ws['A1'].fill = PatternFill('solid', fgColor=_hex(TEAL_DEEP))
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 26

    border_thin = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1'),
    )
    header_fill = PatternFill('solid', fgColor=_hex(TEAL_MAIN))

    # Sub-categorías visuales (row 2)
    cat_row = 2
    categorias = []
    cat_pinta = {
        'nro_doc': 'IDENTIFICACIÓN', 'apellidos_nombres': 'IDENTIFICACIÓN',
        'vac_dias_ganados': 'VACACIONES', 'vac_dias_gozados': 'VACACIONES',
        'vac_saldo_dias': 'VACACIONES', 'vac_record_desde': 'VACACIONES',
        'cts_saldo': 'CTS', 'cts_dias_computables': 'CTS',
        'cts_base_remunerativa': 'CTS', 'cts_entidad': 'CTS',
        'grat_julio_saldo': 'GRATIF.', 'grat_dic_saldo': 'GRATIF.',
        'bonif_extra_saldo': 'GRATIF.',
        'prestamos_saldo': 'PRÉSTAMOS', 'prestamos_cuota': 'PRÉSTAMOS', 'prestamos_cuotas': 'PRÉSTAMOS',
        'banco_horas_saldo': 'BANCO HORAS',
        'adelantos_saldo': 'OTROS', 'otros_desc_pend': 'OTROS',
        'observaciones': 'NOTAS',
    }
    cat_colores = {
        'IDENTIFICACIÓN': '1E40AF',
        'VACACIONES':     '0F766E',
        'CTS':            '7C3AED',
        'GRATIF.':        'F59E0B',
        'PRÉSTAMOS':      'DC2626',
        'BANCO HORAS':    'EA580C',
        'OTROS':          '64748B',
        'NOTAS':          '94A3B8',
    }
    for col_idx, (_, key, *_rest) in enumerate(COLUMNAS, 1):
        cat = cat_pinta.get(key, 'OTROS')
        cell = ws.cell(row=cat_row, column=col_idx, value=cat)
        cell.font = Font(bold=True, size=8, color=_hex(WHITE))
        cell.fill = PatternFill('solid', fgColor=_hex(cat_colores.get(cat, GRAY)))
        cell.alignment = Alignment(horizontal='center')
        cell.border = border_thin
    ws.row_dimensions[cat_row].height = 16

    # Headers
    for col_idx, (header, key, ftype, required, ejemplo, ayuda) in enumerate(COLUMNAS, 1):
        cell = ws.cell(row=3, column=col_idx, value=header + (' *' if required else ''))
        cell.fill = header_fill
        cell.font = Font(bold=True, size=10, color=_hex(WHITE))
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border_thin
        comment_text = (ayuda or '') + (f'\n\nEjemplo: {ejemplo}' if ejemplo else '')
        if required:
            comment_text = '⚠ OBLIGATORIO\n\n' + comment_text
        cell.comment = Comment(comment_text, 'Harmoni')
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(header) + 2)
    ws.row_dimensions[3].height = 36

    # Validaciones (cts_entidad)
    for col_idx, (_, key, ftype, *_r) in enumerate(COLUMNAS, 1):
        if ftype == 'enum' and key in ENUMS:
            vals = ENUMS[key]
            dv = DataValidation(
                type='list', formula1=f'"{",".join(vals)}"', allow_blank=True,
                error='Valor no válido', errorTitle='Error',
                prompt=f'Valores: {", ".join(vals)}', promptTitle=key,
            )
            dv.add(f'{get_column_letter(col_idx)}4:{get_column_letter(col_idx)}1000')
            ws.add_data_validation(dv)

    # Ejemplos
    start_row = 4
    if con_ejemplos:
        for i, ejemplo in enumerate(EJEMPLOS):
            for col_idx, (_, key, ftype, *_r) in enumerate(COLUMNAS, 1):
                val = ejemplo.get(key, '')
                cell = ws.cell(row=start_row + i, column=col_idx, value=val)
                cell.border = border_thin
                if ftype == 'money':
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                elif ftype == 'number':
                    cell.alignment = Alignment(horizontal='right')
                elif ftype == 'date':
                    cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill('solid', fgColor=_hex(TEAL_TINT))
                cell.font = Font(italic=True, size=9, color=_hex(GRAY))
        ws.cell(row=start_row, column=1).comment = Comment('EJEMPLO — eliminar antes de importar', 'Harmoni')

    ws.freeze_panes = f'A{start_row + (len(EJEMPLOS) if con_ejemplos else 0)}'

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


class Command(BaseCommand):
    help = 'Genera plantilla Excel para cargar saldos al cierre del año previo'

    def add_arguments(self, parser):
        parser.add_argument('--output', type=str,
                            default='plantilla_saldos_cierre_HARMONI.xlsx',
                            help='Path de salida')
        parser.add_argument('--anio', type=int, default=2025, help='Año del cierre (default 2025)')
        parser.add_argument('--con-ejemplos', action='store_true')

    def handle(self, *args, **opts):
        output = Path(opts['output']).resolve()
        build_plantilla(output, opts['anio'], opts['con_ejemplos'])
        size = output.stat().st_size
        self.stdout.write(self.style.SUCCESS(
            f'\nPlantilla generada: {output}\n'
            f'   Tamaño: {size:,} bytes\n'
            f'   Año cierre: {opts["anio"]}\n'
            f'   Hojas: Instrucciones · SALDOS\n'
            f'   Columnas: {len(COLUMNAS)} agrupadas por categoría (Identif/Vac/CTS/Grat/Préstamos/Banco/Otros)\n'
        ))
