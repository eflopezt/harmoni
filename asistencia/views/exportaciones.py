"""
Vistas del módulo Tareo — Exportaciones.
"""
import calendar
from datetime import date
from decimal import Decimal
from io import BytesIO

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q, Sum
from django.http import HttpResponse
from django.shortcuts import redirect, render

from asistencia.views._common import solo_admin


# ---------------------------------------------------------------------------
# EXPORTAR CARGA S10
# ---------------------------------------------------------------------------

@login_required
@solo_admin
def exportar_carga_s10_view(request):
    """Genera y descarga el archivo CargaS10 para importar en el sistema S10."""
    from asistencia.services.exporters import CargaS10Exporter

    anio = int(request.GET.get('anio', date.today().year))
    mes = int(request.GET.get('mes', date.today().month))

    try:
        exporter = CargaS10Exporter(anio, mes)
        buffer = exporter.generar()

        MESES = ['enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio',
                 'julio', 'agosto', 'septiembre', 'octubre', 'noviembre', 'diciembre']
        filename = f'CargaS10_{MESES[mes-1]}_{anio}.xlsx'

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    except Exception as e:
        messages.error(request, f'Error generando CargaS10: {e}')
        return redirect('asistencia_dashboard')


# ---------------------------------------------------------------------------
# EXPORTAR REPORTE CIERRE
# ---------------------------------------------------------------------------

@login_required
@solo_admin
def exportar_cierre_view(request):
    """Genera el reporte de cierre de mes."""
    from asistencia.services.exporters import ReporteCierreExporter

    anio = int(request.GET.get('anio', date.today().year))
    mes = int(request.GET.get('mes', date.today().month))

    try:
        exporter = ReporteCierreExporter(anio, mes)
        buffer = exporter.generar()

        MESES = ['enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio',
                 'julio', 'agosto', 'septiembre', 'octubre', 'noviembre', 'diciembre']
        filename = f'Cierre_{MESES[mes-1]}_{anio}.xlsx'

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    except Exception as e:
        messages.error(request, f'Error generando reporte de cierre: {e}')
        return redirect('asistencia_dashboard')


@login_required
@solo_admin
def reportes_exportar_panel(request):
    """Panel central de reportes y exportaciones."""
    hoy = date.today()
    anio = int(request.GET.get('anio', hoy.year))
    mes = int(request.GET.get('mes', hoy.month))
    MESES = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
             'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    return render(request, 'asistencia/reportes_exportar.html', {
        'anio': anio, 'mes': mes,
        'anios': list(range(hoy.year - 2, hoy.year + 1)),
        'meses_list': [(i, MESES[i - 1]) for i in range(1, 13)],
    })


# ---------------------------------------------------------------------------
# EXPORTAR REPORTE DE HORAS RCO (Excel)
# ---------------------------------------------------------------------------

MESES_ES = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
            'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']


@login_required
@solo_admin
def exportar_horas_rco(request):
    """Excel con resumen de horas y HE por trabajador RCO del periodo."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from asistencia.models import RegistroTareo
    from personal.models import Personal

    anio = int(request.GET.get('anio', date.today().year))
    mes = int(request.GET.get('mes', date.today().month))

    mes_ini = date(anio, mes, 1)
    mes_fin = date(anio, mes, calendar.monthrange(anio, mes)[1])

    # Resumen por persona
    resumen = list(
        RegistroTareo.objects.filter(
            grupo='RCO', fecha__gte=mes_ini, fecha__lte=mes_fin
        ).values('dni', 'nombre_archivo', 'personal_id')
        .annotate(
            dias_trabajados=Count('id', filter=Q(
                codigo_dia__in=['T', 'NOR', 'TR', 'A', 'CDT', 'CPF', 'LCG', 'ATM', 'CHE', 'LIM', 'SS'])),
            dias_falta=Count('id', filter=Q(codigo_dia__in=['FA', 'F'])),
            dias_dl=Count('id', filter=Q(codigo_dia__in=['DL', 'DLA'])),
            dias_vac=Count('id', filter=Q(codigo_dia__in=['VAC', 'V'])),
            dias_dm=Count('id', filter=Q(codigo_dia='DM')),
            dias_sai=Count('id', filter=Q(codigo_dia='SAI')),
            total_horas=Sum('horas_marcadas'),
            total_hn=Sum('horas_normales'),
            total_he_25=Sum('he_25'),
            total_he_35=Sum('he_35'),
            total_he_100=Sum('he_100'),
        )
        .order_by('nombre_archivo')
    )

    # Enriquecer con datos de Personal
    pids = [r['personal_id'] for r in resumen if r['personal_id']]
    personal_map = {p.id: p for p in Personal.objects.filter(id__in=pids)}

    for r in resumen:
        r['total_he_25'] = r['total_he_25'] or Decimal('0')
        r['total_he_35'] = r['total_he_35'] or Decimal('0')
        r['total_he_100'] = r['total_he_100'] or Decimal('0')
        r['total_he'] = r['total_he_25'] + r['total_he_35'] + r['total_he_100']
        r['total_horas'] = r['total_horas'] or Decimal('0')
        r['total_hn'] = r['total_hn'] or Decimal('0')
        p = personal_map.get(r['personal_id'])
        r['cargo'] = p.cargo if p else ''
        r['area'] = p.subarea.area.nombre if p and p.subarea else ''

    # ── Excel ──
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Horas RCO {MESES_ES[mes - 1]} {anio}'

    # Estilos
    title_font = Font(bold=True, size=14, color='0f766e')
    sub_font = Font(size=9, color='64748b')
    header_font = Font(bold=True, size=9, color='FFFFFF')
    header_fill = PatternFill(start_color='0f766e', end_color='0f766e', fill_type='solid')
    data_font = Font(size=9)
    num_font = Font(size=9, bold=True)
    total_fill = PatternFill(start_color='134e4a', end_color='134e4a', fill_type='solid')
    total_font = Font(bold=True, size=9, color='FFFFFF')
    border = Border(bottom=Side(style='thin', color='e2e8f0'))
    center = Alignment(horizontal='center')

    # Titulo
    ws.cell(row=1, column=1, value=f'REPORTE DE HORAS — RCO — {MESES_ES[mes - 1].upper()} {anio}').font = title_font
    ws.cell(row=2, column=1, value=f'Periodo: {mes_ini.strftime("%d/%m/%Y")} al {mes_fin.strftime("%d/%m/%Y")}').font = sub_font

    # Headers
    headers = ['N°', 'DNI', 'Apellidos y Nombres', 'Cargo', 'Area',
               'Dias Trab.', 'Faltas', 'DL', 'VAC', 'DM', 'SAI',
               'Hrs Marcadas', 'Hrs Normales', 'HE 25%', 'HE 35%', 'HE 100%', 'Total HE']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # Data
    for i, r in enumerate(resumen, 1):
        row = i + 4
        ws.cell(row=row, column=1, value=i).font = data_font
        ws.cell(row=row, column=2, value=r['dni']).font = data_font
        ws.cell(row=row, column=3, value=r['nombre_archivo']).font = data_font
        ws.cell(row=row, column=4, value=r['cargo']).font = Font(size=8, color='64748b')
        ws.cell(row=row, column=5, value=r['area']).font = Font(size=8, color='64748b')
        ws.cell(row=row, column=6, value=r['dias_trabajados']).font = num_font
        ws.cell(row=row, column=7, value=r['dias_falta']).font = num_font
        ws.cell(row=row, column=8, value=r['dias_dl']).font = data_font
        ws.cell(row=row, column=9, value=r['dias_vac']).font = data_font
        ws.cell(row=row, column=10, value=r['dias_dm']).font = data_font
        ws.cell(row=row, column=11, value=r['dias_sai']).font = data_font
        ws.cell(row=row, column=12, value=float(r['total_horas'])).font = num_font
        ws.cell(row=row, column=13, value=float(r['total_hn'])).font = num_font
        ws.cell(row=row, column=14, value=float(r['total_he_25'])).font = num_font
        ws.cell(row=row, column=15, value=float(r['total_he_35'])).font = num_font
        ws.cell(row=row, column=16, value=float(r['total_he_100'])).font = num_font
        ws.cell(row=row, column=17, value=float(r['total_he'])).font = Font(bold=True, size=10, color='0f766e')
        for c in range(1, 18):
            ws.cell(row=row, column=c).border = border
            if c >= 6:
                ws.cell(row=row, column=c).alignment = center
                ws.cell(row=row, column=c).number_format = '#,##0.00' if c >= 12 else '0'

    # Totals row
    total_row = len(resumen) + 5
    for c in range(1, 18):
        cell = ws.cell(row=total_row, column=c)
        cell.fill = total_fill
        cell.font = total_font
        cell.alignment = center
    ws.cell(row=total_row, column=3, value='TOTALES')
    ws.cell(row=total_row, column=6, value=sum(r['dias_trabajados'] for r in resumen))
    ws.cell(row=total_row, column=7, value=sum(r['dias_falta'] for r in resumen))
    ws.cell(row=total_row, column=8, value=sum(r['dias_dl'] for r in resumen))
    ws.cell(row=total_row, column=9, value=sum(r['dias_vac'] for r in resumen))
    ws.cell(row=total_row, column=10, value=sum(r['dias_dm'] for r in resumen))
    ws.cell(row=total_row, column=11, value=sum(r['dias_sai'] for r in resumen))
    ws.cell(row=total_row, column=12, value=float(sum(r['total_horas'] for r in resumen)))
    ws.cell(row=total_row, column=13, value=float(sum(r['total_hn'] for r in resumen)))
    ws.cell(row=total_row, column=14, value=float(sum(r['total_he_25'] for r in resumen)))
    ws.cell(row=total_row, column=15, value=float(sum(r['total_he_35'] for r in resumen)))
    ws.cell(row=total_row, column=16, value=float(sum(r['total_he_100'] for r in resumen)))
    ws.cell(row=total_row, column=17, value=float(sum(r['total_he'] for r in resumen)))
    for c in [12, 13, 14, 15, 16, 17]:
        ws.cell(row=total_row, column=c).number_format = '#,##0.00'

    # Conteo
    ws.cell(row=total_row + 2, column=1,
            value=f'Total trabajadores: {len(resumen)}').font = Font(size=9, color='64748b')

    # Column widths
    widths = [5, 12, 38, 25, 20, 9, 8, 6, 6, 6, 6, 12, 12, 10, 10, 10, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Freeze panes
    ws.freeze_panes = 'A5'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="Horas_RCO_{MESES_ES[mes - 1]}_{anio}.xlsx"'
    return response
