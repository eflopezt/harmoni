"""
asistencia.views.reportes_avanzados
====================================

Reportes pivote cross-empresa + KPIs + exports Excel/PDF.

Endpoints:
    /asistencia/reportes/pivote/         → HTML tabla pivote
    /asistencia/reportes/pivote/excel/   → XLSX 3 hojas
    /asistencia/reportes/pivote/pdf/     → PDF A4 landscape
    /asistencia/reportes/kpis-cross/     → Panel KPIs cross-empresa
"""
from __future__ import annotations

import io
import json
from datetime import date, datetime, timedelta

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, HttpResponseBadRequest
from django.shortcuts import render

from asistencia.reportes_pivote import (
    COLOR_ESTADO,
    ESTADO_ASISTIO,
    ESTADO_DESCANSO,
    ESTADO_FALTA,
    ESTADO_FERIADO,
    ESTADO_PERMISO,
    ESTADO_TARDANZA,
    ESTADO_VACACIONES,
    ESTADO_VACIO,
    LABEL_ESTADO,
    construir_kpis,
    construir_pivote,
)
from asistencia.views._common import solo_admin


# ─────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────

def _parse_filtros(request):
    """Parsea querystring común a todas las vistas pivote."""
    hoy = date.today()
    desde_str = request.GET.get('desde', '')
    hasta_str = request.GET.get('hasta', '')

    try:
        desde = datetime.strptime(desde_str, '%Y-%m-%d').date() if desde_str else hoy.replace(day=1)
    except ValueError:
        desde = hoy.replace(day=1)
    try:
        hasta = datetime.strptime(hasta_str, '%Y-%m-%d').date() if hasta_str else hoy
    except ValueError:
        hasta = hoy

    if hasta < desde:
        hasta = desde

    empresa_ids_raw = request.GET.get('empresa_ids', '')
    empresa_ids = [int(x) for x in empresa_ids_raw.split(',') if x.strip().isdigit()]

    grupo_tareo = request.GET.get('grupo', '').strip().upper() or None
    if grupo_tareo not in ('STAFF', 'RCO', 'OTRO', None):
        grupo_tareo = None

    area_id = request.GET.get('area_id', '').strip()
    area_id = int(area_id) if area_id.isdigit() else None
    subarea_id = request.GET.get('subarea_id', '').strip()
    subarea_id = int(subarea_id) if subarea_id.isdigit() else None

    return {
        'desde': desde,
        'hasta': hasta,
        'empresa_ids': empresa_ids,
        'grupo_tareo': grupo_tareo,
        'area_id': area_id,
        'subarea_id': subarea_id,
    }


def _empresas_disponibles():
    """Lista para selector de empresas."""
    try:
        from empresas.models import Empresa
        return list(Empresa.objects.order_by('razon_social').values('id', 'razon_social'))
    except Exception:
        return []


# ─────────────────────────────────────────────────────────────────
# Vista HTML — Tabla Pivote
# ─────────────────────────────────────────────────────────────────

@login_required
@solo_admin
def reporte_pivote_view(request):
    """Tabla pivote: filas=trabajadores, columnas=días, celdas=estado."""
    f = _parse_filtros(request)
    pivote = construir_pivote(
        desde=f['desde'], hasta=f['hasta'],
        empresa_ids=f['empresa_ids'] or None,
        grupo_tareo=f['grupo_tareo'],
        area_id=f['area_id'],
        subarea_id=f['subarea_id'],
        limite_trabajadores=500,  # safety cap para la vista HTML
    )

    # Para template: convertir totales_por_dia (defaultdict) a list ordenada
    totales_dia_render = []
    for fecha in pivote['fechas']:
        d = pivote['totales_por_dia'].get(fecha, {})
        presentes = (d.get(ESTADO_ASISTIO, 0) + d.get(ESTADO_TARDANZA, 0))
        totales_dia_render.append({
            'fecha': fecha,
            'presentes': presentes,
            'faltas': d.get(ESTADO_FALTA, 0),
            'vacaciones': d.get(ESTADO_VACACIONES, 0),
            'permisos': d.get(ESTADO_PERMISO, 0),
            'tardanzas': d.get(ESTADO_TARDANZA, 0),
        })

    # Para celdas: enriquecer dict { date: {'codigo': X, 'color': X, 'label': X} }
    trabajadores_render = []
    for t in pivote['trabajadores']:
        dias_render = []
        for fecha in pivote['fechas']:
            codigo = t['dias'].get(fecha, ESTADO_VACIO)
            dias_render.append({
                'fecha': fecha,
                'codigo': codigo,
                'color': COLOR_ESTADO.get(codigo, '#fff'),
                'label': LABEL_ESTADO.get(codigo, ''),
                'es_fin_semana': fecha.weekday() >= 5,
            })
        trabajadores_render.append({
            **t,
            'dias_render': dias_render,
        })

    context = {
        'titulo': 'Reporte Pivote Cross-Empresa',
        'filtros': f,
        'empresas': _empresas_disponibles(),
        'fechas': pivote['fechas'],
        'trabajadores': trabajadores_render,
        'totales_dia': totales_dia_render,
        'totales_globales': dict(pivote['totales_globales']),
        'total_trabajadores': len(pivote['trabajadores']),
        'leyenda': [
            (ESTADO_ASISTIO, LABEL_ESTADO[ESTADO_ASISTIO], COLOR_ESTADO[ESTADO_ASISTIO]),
            (ESTADO_TARDANZA, LABEL_ESTADO[ESTADO_TARDANZA], COLOR_ESTADO[ESTADO_TARDANZA]),
            (ESTADO_FALTA, LABEL_ESTADO[ESTADO_FALTA], COLOR_ESTADO[ESTADO_FALTA]),
            (ESTADO_VACACIONES, LABEL_ESTADO[ESTADO_VACACIONES], COLOR_ESTADO[ESTADO_VACACIONES]),
            (ESTADO_PERMISO, LABEL_ESTADO[ESTADO_PERMISO], COLOR_ESTADO[ESTADO_PERMISO]),
            (ESTADO_DESCANSO, LABEL_ESTADO[ESTADO_DESCANSO], COLOR_ESTADO[ESTADO_DESCANSO]),
            (ESTADO_FERIADO, LABEL_ESTADO[ESTADO_FERIADO], COLOR_ESTADO[ESTADO_FERIADO]),
        ],
        'querystring': request.GET.urlencode(),
    }
    return render(request, 'asistencia/reporte_pivote.html', context)


# ─────────────────────────────────────────────────────────────────
# Vista HTML — KPIs Cross-Empresa
# ─────────────────────────────────────────────────────────────────

@login_required
@solo_admin
def reporte_kpis_cross_view(request):
    """Panel KPIs cross-empresa con stat cards + charts."""
    f = _parse_filtros(request)
    kpis = construir_kpis(
        desde=f['desde'], hasta=f['hasta'],
        empresa_ids=f['empresa_ids'] or None,
    )

    context = {
        'titulo': 'KPIs Cross-Empresa — Asistencia',
        'filtros': f,
        'empresas': _empresas_disponibles(),
        'kpis': kpis,
        'tendencia_json': json.dumps(kpis['tendencia_diaria']),
        'distribucion_json': json.dumps(kpis['distribucion_papeletas']),
    }
    return render(request, 'asistencia/reporte_kpis_cross.html', context)


# ─────────────────────────────────────────────────────────────────
# Export Excel
# ─────────────────────────────────────────────────────────────────

@login_required
@solo_admin
def reporte_pivote_excel(request):
    """Excel con 3 hojas: Pivote, Resumen, Detalle."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponseBadRequest('openpyxl no instalado')

    f = _parse_filtros(request)
    pivote = construir_pivote(
        desde=f['desde'], hasta=f['hasta'],
        empresa_ids=f['empresa_ids'] or None,
        grupo_tareo=f['grupo_tareo'],
        area_id=f['area_id'],
        subarea_id=f['subarea_id'],
    )

    wb = openpyxl.Workbook()

    # ── Estilos ──
    HEADER_FILL = PatternFill('solid', fgColor='0F766E')  # teal
    HEADER_FONT = Font(color='FFFFFF', bold=True, size=10)
    BORDER = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )
    CENTER = Alignment(horizontal='center', vertical='center')
    LEFT = Alignment(horizontal='left', vertical='center')

    color_fills = {
        ESTADO_ASISTIO: PatternFill('solid', fgColor='D4EDDA'),
        ESTADO_TARDANZA: PatternFill('solid', fgColor='FFF3CD'),
        ESTADO_FALTA: PatternFill('solid', fgColor='F8D7DA'),
        ESTADO_VACACIONES: PatternFill('solid', fgColor='CFE2FF'),
        ESTADO_PERMISO: PatternFill('solid', fgColor='E2D9F3'),
        ESTADO_DESCANSO: PatternFill('solid', fgColor='E9ECEF'),
        ESTADO_FERIADO: PatternFill('solid', fgColor='FFE5CC'),
    }

    # ── Hoja 1: Pivote ──
    ws1 = wb.active
    ws1.title = 'Pivote'

    # Header
    headers = ['#', 'Trabajador', 'DNI', 'Empresa', 'Grupo']
    for d in pivote['fechas']:
        headers.append(d.strftime('%d/%m'))
    headers += ['Trab', 'Faltas', 'Tard', 'Vac', 'Perm']
    for col, h in enumerate(headers, 1):
        c = ws1.cell(row=1, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = BORDER

    # Filas
    for idx, t in enumerate(pivote['trabajadores'], start=1):
        row = idx + 1
        p = t['personal']
        ws1.cell(row=row, column=1, value=idx).alignment = CENTER
        ws1.cell(row=row, column=2, value=p.apellidos_nombres).alignment = LEFT
        ws1.cell(row=row, column=3, value=p.nro_doc).alignment = CENTER
        ws1.cell(row=row, column=4, value=t['empresa_nombre']).alignment = LEFT
        ws1.cell(row=row, column=5, value=getattr(p, 'grupo_tareo', '') or '').alignment = CENTER

        col = 6
        for fecha in pivote['fechas']:
            codigo = t['dias'].get(fecha, ESTADO_VACIO)
            c = ws1.cell(row=row, column=col, value=codigo)
            c.alignment = CENTER
            c.font = Font(size=9, bold=True)
            if codigo in color_fills:
                c.fill = color_fills[codigo]
            col += 1

        ws1.cell(row=row, column=col, value=t['total_trabajados']).alignment = CENTER
        ws1.cell(row=row, column=col + 1, value=t['total_faltas']).alignment = CENTER
        ws1.cell(row=row, column=col + 2, value=t['total_tardanzas']).alignment = CENTER
        ws1.cell(row=row, column=col + 3, value=t['total_vacaciones']).alignment = CENTER
        ws1.cell(row=row, column=col + 4, value=t['total_permisos']).alignment = CENTER

    # Anchos
    ws1.column_dimensions['A'].width = 5
    ws1.column_dimensions['B'].width = 32
    ws1.column_dimensions['C'].width = 12
    ws1.column_dimensions['D'].width = 22
    ws1.column_dimensions['E'].width = 9
    for i in range(len(pivote['fechas'])):
        ws1.column_dimensions[get_column_letter(6 + i)].width = 5
    last = 6 + len(pivote['fechas'])
    for i in range(5):
        ws1.column_dimensions[get_column_letter(last + i)].width = 8

    # Freeze pane
    ws1.freeze_panes = 'F2'

    # ── Hoja 2: Resumen ──
    ws2 = wb.create_sheet('Resumen')
    resumen_headers = ['#', 'Trabajador', 'DNI', 'Empresa', 'Grupo',
                       'Días Trab', 'Faltas', 'Tardanzas', 'Vacaciones',
                       'Permisos', 'Descansos', 'Horas Totales', 'HE Totales']
    for col, h in enumerate(resumen_headers, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER

    for idx, t in enumerate(pivote['trabajadores'], start=1):
        p = t['personal']
        ws2.append([
            idx, p.apellidos_nombres, p.nro_doc, t['empresa_nombre'],
            getattr(p, 'grupo_tareo', '') or '',
            t['total_trabajados'], t['total_faltas'], t['total_tardanzas'],
            t['total_vacaciones'], t['total_permisos'], t['total_descansos'],
            round(t['total_horas'], 2), round(t['total_he'], 2),
        ])

    for col_letter, width in zip('ABCDEFGHIJKLM', [5, 32, 12, 22, 9, 10, 9, 11, 11, 10, 11, 13, 11]):
        ws2.column_dimensions[col_letter].width = width
    ws2.freeze_panes = 'A2'

    # ── Hoja 3: Detalle (papeletas + tardanzas) ──
    ws3 = wb.create_sheet('Detalle')
    detalle_headers = ['Trabajador', 'DNI', 'Empresa', 'Tipo', 'Detalle', 'Fecha Ini', 'Fecha Fin', 'Estado']
    for col, h in enumerate(detalle_headers, 1):
        c = ws3.cell(row=1, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER

    # Papeletas del período de los trabajadores del filtro
    try:
        from asistencia.models import RegistroPapeleta
        personal_ids = [t['personal'].id for t in pivote['trabajadores']]
        if personal_ids:
            pap_qs = RegistroPapeleta.objects.filter(
                personal_id__in=personal_ids,
                fecha_inicio__lte=f['hasta'],
                fecha_fin__gte=f['desde'],
            ).select_related('personal', 'personal__empresa').order_by('-fecha_inicio')
            for pap in pap_qs[:5000]:
                p = pap.personal
                emp = ''
                if p and p.empresa:
                    emp = getattr(p.empresa, 'nombre_comercial', '') or getattr(p.empresa, 'razon_social', '') or ''
                ws3.append([
                    p.apellidos_nombres if p else pap.dni,
                    p.nro_doc if p else pap.dni,
                    emp,
                    pap.get_tipo_permiso_display(),
                    (pap.detalle or '')[:200],
                    pap.fecha_inicio.strftime('%Y-%m-%d'),
                    pap.fecha_fin.strftime('%Y-%m-%d'),
                    pap.get_estado_display(),
                ])
    except Exception:
        pass

    for col_letter, width in zip('ABCDEFGH', [32, 12, 22, 28, 40, 12, 12, 12]):
        ws3.column_dimensions[col_letter].width = width
    ws3.freeze_panes = 'A2'

    # ── Response ──
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    response = HttpResponse(
        out.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    fname = f"pivote_asistencia_{f['desde']:%Y%m%d}_{f['hasta']:%Y%m%d}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    return response


# ─────────────────────────────────────────────────────────────────
# Export PDF
# ─────────────────────────────────────────────────────────────────

@login_required
@solo_admin
def reporte_pivote_pdf(request):
    """PDF A4 landscape, paginado a 50 trabajadores."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
        )
    except ImportError:
        return HttpResponseBadRequest('reportlab no instalado')

    f = _parse_filtros(request)
    pivote = construir_pivote(
        desde=f['desde'], hasta=f['hasta'],
        empresa_ids=f['empresa_ids'] or None,
        grupo_tareo=f['grupo_tareo'],
        area_id=f['area_id'],
        subarea_id=f['subarea_id'],
    )

    styles = getSampleStyleSheet()
    st_title = ParagraphStyle('t', parent=styles['Normal'],
                              fontSize=11, fontName='Helvetica-Bold',
                              textColor=colors.HexColor('#0F766E'))
    st_sub = ParagraphStyle('s', parent=styles['Normal'],
                            fontSize=8, textColor=colors.HexColor('#555555'))
    st_cell = ParagraphStyle('c', parent=styles['Normal'], fontSize=6.5)
    st_legend = ParagraphStyle('l', parent=styles['Normal'], fontSize=7)

    # Mapas de color rl
    rl_color = {
        ESTADO_ASISTIO: colors.HexColor('#D4EDDA'),
        ESTADO_TARDANZA: colors.HexColor('#FFF3CD'),
        ESTADO_FALTA: colors.HexColor('#F8D7DA'),
        ESTADO_VACACIONES: colors.HexColor('#CFE2FF'),
        ESTADO_PERMISO: colors.HexColor('#E2D9F3'),
        ESTADO_DESCANSO: colors.HexColor('#E9ECEF'),
        ESTADO_FERIADO: colors.HexColor('#FFE5CC'),
    }

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=10 * mm, rightMargin=10 * mm,
        topMargin=10 * mm, bottomMargin=12 * mm,
        title='Reporte Pivote Asistencia',
    )

    story = []
    PAGE_SIZE = 50
    fechas = pivote['fechas']
    trabajadores = pivote['trabajadores']

    if not trabajadores:
        story.append(Paragraph('Reporte Pivote Cross-Empresa', st_title))
        story.append(Spacer(1, 4))
        story.append(Paragraph(
            f"Período: {f['desde']:%d/%m/%Y} → {f['hasta']:%d/%m/%Y}", st_sub))
        story.append(Spacer(1, 10))
        story.append(Paragraph('Sin trabajadores que coincidan con los filtros.', st_cell))
    else:
        # Paginar
        for page_start in range(0, len(trabajadores), PAGE_SIZE):
            chunk = trabajadores[page_start:page_start + PAGE_SIZE]

            story.append(Paragraph('Reporte Pivote Asistencia Cross-Empresa', st_title))
            story.append(Paragraph(
                f"Período: {f['desde']:%d/%m/%Y} → {f['hasta']:%d/%m/%Y}  "
                f"| Trabajadores: {len(trabajadores)}  "
                f"| Página {page_start // PAGE_SIZE + 1} de {(len(trabajadores) - 1) // PAGE_SIZE + 1}",
                st_sub))
            story.append(Spacer(1, 4))

            # Construir tabla
            header = ['#', 'Trabajador', 'DNI']
            for d in fechas:
                header.append(d.strftime('%d'))
            header += ['Trab', 'F', 'T']

            data = [header]
            cell_bg = []  # list of (col, row, color)

            for i, t in enumerate(chunk, start=1):
                global_idx = page_start + i
                p = t['personal']
                nombre = (p.apellidos_nombres or '')[:28]
                fila = [str(global_idx), nombre, p.nro_doc or '']
                for col_idx, fecha in enumerate(fechas):
                    codigo = t['dias'].get(fecha, ESTADO_VACIO)
                    fila.append(codigo)
                    if codigo in rl_color:
                        cell_bg.append((3 + col_idx, len(data), rl_color[codigo]))
                fila.append(str(t['total_trabajados']))
                fila.append(str(t['total_faltas']))
                fila.append(str(t['total_tardanzas']))
                data.append(fila)

            # Column widths
            ndias = len(fechas)
            avail = landscape(A4)[0] - 20 * mm
            fixed = 9 + 65 + 28 + 18 + 14 + 14  # mm
            day_w_mm = max(4.5, (avail / mm - fixed) / max(ndias, 1))
            col_widths_mm = [9, 65, 28] + [day_w_mm] * ndias + [18, 14, 14]
            col_widths = [w * mm for w in col_widths_mm]

            tbl = Table(data, colWidths=col_widths, repeatRows=1)
            ts = [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0F766E')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 7),
                ('FONTSIZE', (0, 1), (-1, -1), 6.5),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (1, 1), (1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#CCCCCC')),
                ('FONTNAME', (3, 1), (3 + ndias - 1, -1), 'Helvetica-Bold'),
            ]
            for col, row, c in cell_bg:
                ts.append(('BACKGROUND', (col, row), (col, row), c))
            tbl.setStyle(TableStyle(ts))
            story.append(tbl)

            # Leyenda al final de cada página
            story.append(Spacer(1, 6))
            leg_data = [[
                Paragraph('<b>Leyenda:</b>', st_legend),
                Paragraph(f'<b>{ESTADO_ASISTIO}</b>=Asistió', st_legend),
                Paragraph(f'<b>{ESTADO_TARDANZA}</b>=Tarde', st_legend),
                Paragraph(f'<b>{ESTADO_FALTA}</b>=Falta', st_legend),
                Paragraph(f'<b>{ESTADO_VACACIONES}</b>=Vac', st_legend),
                Paragraph(f'<b>{ESTADO_PERMISO}</b>=Permiso', st_legend),
                Paragraph(f'<b>{ESTADO_DESCANSO}</b>=Descanso', st_legend),
                Paragraph(f'<b>{ESTADO_FERIADO}</b>=Feriado', st_legend),
            ]]
            leg_tbl = Table(leg_data, colWidths=[25 * mm] + [30 * mm] * 7)
            leg_tbl.setStyle(TableStyle([
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('BACKGROUND', (1, 0), (1, 0), rl_color[ESTADO_ASISTIO]),
                ('BACKGROUND', (2, 0), (2, 0), rl_color[ESTADO_TARDANZA]),
                ('BACKGROUND', (3, 0), (3, 0), rl_color[ESTADO_FALTA]),
                ('BACKGROUND', (4, 0), (4, 0), rl_color[ESTADO_VACACIONES]),
                ('BACKGROUND', (5, 0), (5, 0), rl_color[ESTADO_PERMISO]),
                ('BACKGROUND', (6, 0), (6, 0), rl_color[ESTADO_DESCANSO]),
                ('BACKGROUND', (7, 0), (7, 0), rl_color[ESTADO_FERIADO]),
            ]))
            story.append(leg_tbl)

            if page_start + PAGE_SIZE < len(trabajadores):
                story.append(PageBreak())

    doc.build(story)
    pdf_bytes = buf.getvalue()
    buf.close()

    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    fname = f"pivote_asistencia_{f['desde']:%Y%m%d}_{f['hasta']:%Y%m%d}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    return response
