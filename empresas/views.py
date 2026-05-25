"""Empresas — Vistas: CRUD de empresas y selección de empresa activa."""
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import get_object_or_404, redirect, render
from django.utils.http import url_has_allowed_host_and_scheme
from django.views.decorators.http import require_POST

from .models import Empresa

solo_admin = user_passes_test(lambda u: u.is_superuser)


@login_required
@solo_admin
def empresas_panel(request):
    """Lista de empresas."""
    empresas = Empresa.objects.all()
    return render(request, 'empresas/panel.html', {
        'titulo': 'Empresas',
        'empresas': empresas,
    })


@login_required
@solo_admin
def empresa_crear(request):
    """Crear nueva empresa."""
    if request.method == 'POST':
        ruc          = request.POST.get('ruc', '').strip()
        razon_social = request.POST.get('razon_social', '').strip()
        nombre_comercial = request.POST.get('nombre_comercial', '').strip()
        es_principal = request.POST.get('es_principal') == '1'

        if not ruc or not razon_social:
            messages.error(request, 'RUC y Razón Social son requeridos.')
            return redirect('empresa_crear')

        if Empresa.objects.filter(ruc=ruc).exists():
            messages.error(request, f'Ya existe una empresa con RUC {ruc}.')
            return redirect('empresa_crear')

        emp = Empresa.objects.create(
            ruc=ruc,
            razon_social=razon_social,
            nombre_comercial=nombre_comercial,
            es_principal=es_principal,
            creado_por=request.user,
        )
        messages.success(request, f'Empresa "{emp}" creada exitosamente.')
        return redirect('empresas_panel')

    return render(request, 'empresas/form.html', {
        'titulo': 'Nueva Empresa',
        'action': 'crear',
        'regimen_choices': Empresa.REGIMEN_CHOICES,
    })


@login_required
@solo_admin
def empresa_editar(request, pk):
    """Editar empresa existente."""
    empresa = get_object_or_404(Empresa, pk=pk)

    if request.method == 'POST':
        empresa.razon_social     = request.POST.get('razon_social', empresa.razon_social).strip()
        empresa.nombre_comercial = request.POST.get('nombre_comercial', '').strip()
        empresa.ruc              = request.POST.get('ruc', empresa.ruc).strip()
        empresa.direccion        = request.POST.get('direccion', '').strip()
        empresa.telefono         = request.POST.get('telefono', '').strip()
        empresa.email_rrhh       = request.POST.get('email_rrhh', '').strip()
        empresa.regimen_laboral  = request.POST.get('regimen_laboral', 'GENERAL')
        empresa.es_principal     = request.POST.get('es_principal') == '1'
        empresa.activa           = request.POST.get('activa') == '1'
        # Portal de beneficios externo (opcional)
        empresa.portal_beneficios_nombre   = request.POST.get('portal_beneficios_nombre', '').strip()
        empresa.portal_beneficios_url      = request.POST.get('portal_beneficios_url', '').strip()
        empresa.portal_beneficios_logo_url = request.POST.get('portal_beneficios_logo_url', '').strip()
        empresa.save()
        messages.success(request, f'Empresa "{empresa}" actualizada.')
        return redirect('empresas_panel')

    return render(request, 'empresas/form.html', {
        'titulo': f'Editar — {empresa.nombre_display}',
        'empresa': empresa,
        'action': 'editar',
        'regimen_choices': Empresa.REGIMEN_CHOICES,
    })


@login_required
@require_POST
def seleccionar_empresa(request):
    """
    Cambia la empresa activa en la sesión.
    Cualquier usuario autenticado puede cambiar su empresa activa.
    """
    empresa_id = request.POST.get('empresa_id')
    next_url   = request.POST.get('next', '/')
    # Prevent open redirect — only allow internal URLs
    if not url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}):
        next_url = '/'

    if empresa_id == 'all':
        # Vista consolidada — mostrar trabajadores/asistencia de TODAS las empresas.
        # Las planillas siguen siendo independientes por empresa.
        request.session['modo_consolidado'] = True
        request.session.pop('empresa_actual_id', None)
        request.session.pop('empresa_actual_nombre', None)
        messages.success(request, '🌐 Vista consolidada — Todas las empresas')
    elif empresa_id:
        try:
            emp = Empresa.objects.get(pk=empresa_id, activa=True)
            request.session['empresa_actual_id']     = emp.pk
            request.session['empresa_actual_nombre'] = emp.nombre_display
            request.session.pop('modo_consolidado', None)
            messages.success(request, f'Empresa activa: {emp.nombre_display}')
        except Empresa.DoesNotExist:
            messages.error(request, 'Empresa no encontrada.')
    else:
        # Limpiar selección (vuelve a la empresa principal)
        request.session.pop('empresa_actual_id', None)
        request.session.pop('empresa_actual_nombre', None)
        request.session.pop('modo_consolidado', None)
        messages.info(request, 'Empresa activa restablecida.')

    return redirect(next_url)


@login_required
@solo_admin
def configuracion_empresa(request, pk):
    """Configurar identidad visual de la empresa: logo, membrete, firma."""
    empresa = get_object_or_404(Empresa, pk=pk)

    if request.method == 'POST':
        # Campos de texto
        empresa.representante_legal = request.POST.get('representante_legal', '').strip()
        empresa.cargo_representante = request.POST.get('cargo_representante', '').strip()

        # Corte de planilla
        try:
            dia_ini = int(request.POST.get('dia_inicio_corte', 22))
            if 1 <= dia_ini <= 28:
                empresa.dia_inicio_corte = dia_ini
        except (ValueError, TypeError):
            pass
        try:
            dia_fin = int(request.POST.get('dia_fin_corte', 21))
            if 1 <= dia_fin <= 28:
                empresa.dia_fin_corte = dia_fin
        except (ValueError, TypeError):
            pass

        # Archivos de imagen
        if 'logo' in request.FILES:
            empresa.logo = request.FILES['logo']
        if 'membrete_header' in request.FILES:
            empresa.membrete_header = request.FILES['membrete_header']
        if 'firma_representante' in request.FILES:
            empresa.firma_representante = request.FILES['firma_representante']

        # Permitir borrar imágenes
        if request.POST.get('borrar_logo') == '1':
            empresa.logo = ''
        if request.POST.get('borrar_membrete') == '1':
            empresa.membrete_header = ''
        if request.POST.get('borrar_firma') == '1':
            empresa.firma_representante = ''

        empresa.save()
        messages.success(request, 'Configuración de empresa actualizada correctamente.')
        return redirect('configuracion_empresa', pk=empresa.pk)

    return render(request, 'empresas/configuracion.html', {
        'titulo': f'Configuración — {empresa.nombre_display}',
        'empresa': empresa,
    })


# ──────────────────────────────────────────────────────────────────
#  Pulse del Grupo — Dashboard multi-local para grupo corporativo
# ──────────────────────────────────────────────────────────────────

@login_required
@solo_admin
def pulse_del_grupo(request):
    """Dashboard ejecutivo: vista unificada de todas las empresas/locales.

    Para grupos corporativos con múltiples RUCs (gastronomía, retail, etc).
    Muestra cada empresa como tarjeta con KPIs operacionales + estado:
    - Headcount activo
    - Asistencia hoy (presentes / faltas / sin marcar)
    - Planilla del último período (neto + costo empresa)
    - Alertas (contratos por vencer, HE excedida, etc.)
    - Código de color: verde/amarillo/rojo según salud operativa

    El "cockpit de Isabel" — la pantalla que el dueño abre con su café.
    """
    from datetime import date, timedelta
    from django.db.models import Count, Q, Sum
    from personal.models import Personal
    from nominas.models import PeriodoNomina
    from empresas.models import Empresa

    hoy = date.today()
    en_30_dias = hoy + timedelta(days=30)

    empresas_qs = Empresa.objects.filter(activa=True).order_by('razon_social')

    locales = []
    for emp in empresas_qs:
        # Headcount activo
        personal_qs = Personal.objects.filter(empresa=emp, estado='Activo')
        headcount = personal_qs.count()

        # Asistencia hoy (solo si el modulo está activo)
        asistencia_hoy = None
        try:
            from asistencia.models import RegistroTareo
            tareo_qs = RegistroTareo.objects.filter(
                personal__empresa=emp, fecha=hoy,
            )
            agg = tareo_qs.aggregate(
                total=Count('id'),
                presentes=Count('id', filter=Q(codigo_dia__in=['T', 'NOR', 'TR', 'SS'])),
                faltas=Count('id', filter=Q(codigo_dia__in=['FA', 'LSG'])),
                permisos=Count('id', filter=Q(codigo_dia__in=[
                    'V', 'DL', 'DLA', 'DM', 'LCG', 'LF', 'LP', 'LM',
                ])),
            )
            sin_marcar = headcount - (agg['total'] or 0)
            asistencia_hoy = {**agg, 'sin_marcar': max(0, sin_marcar)}
        except Exception:
            pass

        # Último período de planilla (consolidado entre todas las empresas)
        # Usamos el último período REGULAR cerrado/aprobado, agregando por empresa
        ultimo_periodo = PeriodoNomina.objects.filter(
            tipo='REGULAR',
            estado__in=['CALCULADO', 'APROBADO', 'CERRADO'],
        ).order_by('-anio', '-mes').first()

        planilla_emp = None
        if ultimo_periodo:
            from nominas.models import RegistroNomina
            reg_qs = RegistroNomina.objects.filter(
                periodo=ultimo_periodo, personal__empresa=emp,
            )
            agg_pl = reg_qs.aggregate(
                trabajadores=Count('id'),
                neto=Sum('neto_a_pagar'),
                costo=Sum('costo_total_empresa'),
            )
            if agg_pl['trabajadores']:
                planilla_emp = {
                    'periodo': f'{ultimo_periodo.mes_nombre} {ultimo_periodo.anio}',
                    'trabajadores': agg_pl['trabajadores'],
                    'neto':         agg_pl['neto']  or 0,
                    'costo':        agg_pl['costo'] or 0,
                }

        # Alertas
        alertas = []
        # Contratos por vencer (30 días)
        contratos_vencen = personal_qs.filter(
            fecha_fin_contrato__isnull=False,
            fecha_fin_contrato__gte=hoy,
            fecha_fin_contrato__lte=en_30_dias,
        ).count()
        if contratos_vencen:
            alertas.append({
                'tipo':    'warning',
                'icono':   'fa-file-contract',
                'texto':   f'{contratos_vencen} contrato(s) por vencer en 30 días',
                'count':   contratos_vencen,
            })

        # Sin saldos de apertura cargados → bandera amarilla
        try:
            from nominas.models import SaldoAperturaTrabajador
            personal_con_saldo = SaldoAperturaTrabajador.objects.filter(
                personal__empresa=emp,
            ).count()
            sin_saldo = headcount - personal_con_saldo
            if headcount and sin_saldo > headcount * 0.5:
                alertas.append({
                    'tipo':  'info',
                    'icono': 'fa-rocket',
                    'texto': f'Saldos de apertura pendientes ({sin_saldo} trab.)',
                    'count': sin_saldo,
                })
        except Exception:
            pass

        # Determinar color de salud
        if asistencia_hoy and asistencia_hoy.get('faltas', 0) > headcount * 0.15:
            color = 'rojo'
        elif alertas or (asistencia_hoy and asistencia_hoy.get('sin_marcar', 0) > headcount * 0.3):
            color = 'amarillo'
        else:
            color = 'verde'

        locales.append({
            'empresa':        emp,
            'headcount':      headcount,
            'asistencia_hoy': asistencia_hoy,
            'planilla':       planilla_emp,
            'alertas':        alertas,
            'color':          color,
            'contratos_vencen': contratos_vencen,
        })

    # Stats agregadas globales
    total_locales = len(locales)
    total_headcount = sum(l['headcount'] for l in locales)
    total_planilla_neto = sum((l['planilla']['neto'] if l['planilla'] else 0) for l in locales)
    total_planilla_costo = sum((l['planilla']['costo'] if l['planilla'] else 0) for l in locales)
    total_alertas = sum(len(l['alertas']) for l in locales)
    locales_verde = sum(1 for l in locales if l['color'] == 'verde')
    locales_amarillo = sum(1 for l in locales if l['color'] == 'amarillo')
    locales_rojo = sum(1 for l in locales if l['color'] == 'rojo')

    # Briefings publicados HOY (globales)
    briefings_hoy_count = 0
    try:
        from asistencia.models import BriefingServicio
        briefings_hoy_count = BriefingServicio.objects.filter(
            fecha=hoy, estado='PUBLICADO',
        ).count()
    except Exception:
        pass

    # Export PDF (?format=pdf)
    if request.GET.get('format') == 'pdf':
        return _exportar_pulse_pdf(
            locales=locales, hoy=hoy,
            total_locales=total_locales, total_headcount=total_headcount,
            total_planilla_neto=total_planilla_neto,
            total_planilla_costo=total_planilla_costo,
            total_alertas=total_alertas, briefings_hoy_count=briefings_hoy_count,
            locales_verde=locales_verde, locales_amarillo=locales_amarillo,
            locales_rojo=locales_rojo,
            user=request.user,
        )

    return render(request, 'empresas/pulse_grupo.html', {
        'locales':              locales,
        'total_locales':        total_locales,
        'total_headcount':      total_headcount,
        'total_planilla_neto':  total_planilla_neto,
        'total_planilla_costo': total_planilla_costo,
        'total_alertas':        total_alertas,
        'briefings_hoy_count':  briefings_hoy_count,
        'locales_verde':        locales_verde,
        'locales_amarillo':     locales_amarillo,
        'locales_rojo':         locales_rojo,
        'hoy':                  hoy,
    })


def _exportar_pulse_pdf(*, locales, hoy, total_locales, total_headcount,
                        total_planilla_neto, total_planilla_costo,
                        total_alertas, briefings_hoy_count,
                        locales_verde, locales_amarillo, locales_rojo, user):
    """Genera PDF ejecutivo del Pulse del Grupo (1 página)."""
    from io import BytesIO
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable,
    )
    from django.http import HttpResponse
    from django.utils import timezone

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        topMargin=12*mm, bottomMargin=12*mm,
        leftMargin=15*mm, rightMargin=15*mm,
    )

    TEAL_DARK = colors.HexColor('#0d2b27')
    TEAL = colors.HexColor('#0f766e')
    GREEN = colors.HexColor('#10b981')
    AMBER = colors.HexColor('#f59e0b')
    RED = colors.HexColor('#ef4444')
    GRAY = colors.HexColor('#64748b')
    LIGHT = colors.HexColor('#f0fafa')

    styles = getSampleStyleSheet()
    st_title = ParagraphStyle('T', parent=styles['Heading1'],
        fontName='Helvetica-Bold', fontSize=18, textColor=TEAL_DARK, alignment=TA_LEFT,
        leading=22, spaceAfter=2)
    st_sub = ParagraphStyle('S', parent=styles['Normal'], fontName='Helvetica',
        fontSize=10, textColor=GRAY, alignment=TA_LEFT, spaceAfter=10)
    st_body = ParagraphStyle('B', parent=styles['Normal'], fontName='Helvetica',
        fontSize=9, textColor=colors.HexColor('#1f2937'), leading=12)

    story = []
    story.append(Paragraph('Pulse del Grupo — Snapshot Ejecutivo', st_title))
    story.append(Paragraph(
        f"Generado el {timezone.now().strftime('%d/%m/%Y %H:%M')} · "
        f"Usuario: {user.username}",
        st_sub,
    ))
    story.append(HRFlowable(width="100%", thickness=2, color=TEAL,
                            spaceBefore=0, spaceAfter=10))

    # KPIs globales
    kpis_data = [
        ['Locales activos',  str(total_locales),  'Verde',   f'{locales_verde}'],
        ['Headcount total',  f'{total_headcount:,}',  'Amarillo', f'{locales_amarillo}'],
        ['Alertas activas',  str(total_alertas), 'Rojo',    f'{locales_rojo}'],
        ['Briefings hoy',    str(briefings_hoy_count), 'Planilla', f'S/ {total_planilla_neto:,.0f}'],
    ]
    kpis = Table(kpis_data, colWidths=[35*mm, 25*mm, 35*mm, 35*mm])
    kpis.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('FONTSIZE', (1, 0), (1, -1), 11),
        ('FONTSIZE', (3, 0), (3, -1), 11),
        ('TEXTCOLOR', (0, 0), (0, -1), GRAY),
        ('TEXTCOLOR', (2, 0), (2, -1), GRAY),
        ('TEXTCOLOR', (1, 0), (1, 0), TEAL_DARK),
        ('TEXTCOLOR', (3, 0), (3, 0), GREEN),
        ('TEXTCOLOR', (3, 1), (3, 1), AMBER),
        ('TEXTCOLOR', (3, 2), (3, 2), RED),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('LINEBELOW', (0, 0), (-1, -2), 0.3, colors.HexColor('#cbd5e1')),
    ]))
    story.append(kpis)
    story.append(Spacer(1, 10))

    story.append(Paragraph('Detalle por Local', ParagraphStyle('h2', parent=st_title,
        fontSize=11, spaceAfter=4, textColor=TEAL)))

    # Tabla de locales
    locales_data = [
        ['Local', 'RUC', 'HC', 'Estado', 'Planilla (S/)', 'Alertas']
    ]
    for l in locales:
        emp = l['empresa']
        color_text = {'verde': 'OK', 'amarillo': 'ATENCIÓN', 'rojo': 'CRÍTICO'}.get(l['color'], '—')
        planilla = (
            f"{l['planilla']['neto']:,.0f}" if l.get('planilla') else '—'
        )
        alertas_str = (', '.join(a.get('texto', '')[:30] for a in l['alertas'][:2])
                       if l.get('alertas') else '—')
        locales_data.append([
            (emp.nombre_comercial or emp.razon_social)[:30],
            emp.ruc or '—',
            str(l['headcount']),
            color_text,
            planilla,
            alertas_str[:50],
        ])

    locales_t = Table(
        locales_data,
        colWidths=[50*mm, 28*mm, 12*mm, 22*mm, 25*mm, 47*mm],
        repeatRows=1,
    )
    style_cmds = [
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('BACKGROUND', (0, 0), (-1, 0), TEAL_DARK),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, LIGHT]),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#cbd5e1')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (2, 1), (2, -1), 'CENTER'),
        ('ALIGN', (4, 1), (4, -1), 'RIGHT'),
    ]
    # Colorear columna Estado según color del local
    for idx, l in enumerate(locales, start=1):
        color = {'verde': GREEN, 'amarillo': AMBER, 'rojo': RED}.get(l['color'])
        if color:
            style_cmds.append(('TEXTCOLOR', (3, idx), (3, idx), color))
            style_cmds.append(('FONTNAME', (3, idx), (3, idx), 'Helvetica-Bold'))
    locales_t.setStyle(TableStyle(style_cmds))
    story.append(locales_t)

    # Footer
    story.append(Spacer(1, 14))
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor('#cbd5e1'),
                            spaceBefore=0, spaceAfter=4))
    story.append(Paragraph(
        f'Harmoni ERP · Pulse del Grupo · Reporte ejecutivo generado automáticamente · '
        f'Costo empresa total: S/ {total_planilla_costo:,.0f}',
        ParagraphStyle('foot', parent=st_body, fontSize=7, textColor=GRAY, alignment=TA_CENTER),
    ))

    doc.build(story)
    pdf = buf.getvalue()
    buf.close()

    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = (
        f'inline; filename="pulse_grupo_{hoy.strftime("%Y%m%d")}.pdf"'
    )
    return response


@login_required
@solo_admin
def pulse_local_detalle(request, pk):
    """Drill-down de un local específico desde el Pulse del Grupo.

    Muestra detalle completo del local: trabajadores, briefings hoy,
    asistencia hoy, alertas, último período planilla.
    """
    from datetime import date, timedelta
    from django.db.models import Count, Q, Sum
    from personal.models import Personal
    from nominas.models import PeriodoNomina, RegistroNomina
    from empresas.models import Empresa

    emp = get_object_or_404(Empresa, pk=pk)
    hoy = date.today()
    en_30_dias = hoy + timedelta(days=30)

    # Trabajadores activos del local
    personal_qs = Personal.objects.filter(empresa=emp, estado='Activo').order_by('apellidos_nombres')
    headcount = personal_qs.count()

    # Distribución por grupo
    grupos = list(
        personal_qs.values('grupo_tareo')
        .annotate(n=Count('id'))
        .order_by('-n')
    )

    # Asistencia hoy
    asistencia_hoy = None
    try:
        from asistencia.models import RegistroTareo
        tareo_qs = RegistroTareo.objects.filter(personal__empresa=emp, fecha=hoy)
        agg = tareo_qs.aggregate(
            total=Count('id'),
            presentes=Count('id', filter=Q(codigo_dia__in=['T', 'NOR', 'TR', 'SS'])),
            faltas=Count('id', filter=Q(codigo_dia__in=['FA', 'LSG'])),
            permisos=Count('id', filter=Q(codigo_dia__in=['V', 'DL', 'DLA', 'DM', 'LCG', 'LF', 'LP', 'LM'])),
        )
        asistencia_hoy = {**agg, 'sin_marcar': max(0, headcount - (agg['total'] or 0))}
    except Exception:
        pass

    # Briefings de hoy/mañana
    briefings = []
    try:
        from asistencia.models import BriefingServicio
        briefings = list(
            BriefingServicio.objects
            .filter(empresa=emp, fecha__gte=hoy, fecha__lte=hoy + timedelta(days=2))
            .order_by('fecha', 'servicio')
        )
    except Exception:
        pass

    # Último período planilla
    ultimo_periodo = PeriodoNomina.objects.filter(
        tipo='REGULAR',
        estado__in=['CALCULADO', 'APROBADO', 'CERRADO'],
    ).order_by('-anio', '-mes').first()

    planilla_emp = None
    if ultimo_periodo:
        reg_qs = RegistroNomina.objects.filter(periodo=ultimo_periodo, personal__empresa=emp)
        agg_pl = reg_qs.aggregate(
            trabajadores=Count('id'),
            bruto=Sum('total_ingresos'),
            descuentos=Sum('total_descuentos'),
            neto=Sum('neto_a_pagar'),
            costo=Sum('costo_total_empresa'),
        )
        if agg_pl['trabajadores']:
            planilla_emp = {
                'periodo':      f'{ultimo_periodo.mes_nombre} {ultimo_periodo.anio}',
                'periodo_pk':   ultimo_periodo.pk,
                'trabajadores': agg_pl['trabajadores'],
                'bruto':        agg_pl['bruto']      or 0,
                'descuentos':   agg_pl['descuentos'] or 0,
                'neto':         agg_pl['neto']       or 0,
                'costo':        agg_pl['costo']      or 0,
            }

    # Contratos por vencer
    contratos_vencen = personal_qs.filter(
        fecha_fin_contrato__isnull=False,
        fecha_fin_contrato__gte=hoy,
        fecha_fin_contrato__lte=en_30_dias,
    ).order_by('fecha_fin_contrato')[:10]

    # Saldos apertura del local
    saldos_apertura_count = 0
    try:
        from nominas.models import SaldoAperturaTrabajador
        saldos_apertura_count = SaldoAperturaTrabajador.objects.filter(
            personal__empresa=emp,
        ).count()
    except Exception:
        pass

    return render(request, 'empresas/pulse_local_detalle.html', {
        'empresa':              emp,
        'headcount':            headcount,
        'grupos':               grupos,
        'asistencia_hoy':       asistencia_hoy,
        'briefings':            briefings,
        'planilla_emp':         planilla_emp,
        'ultimo_periodo':       ultimo_periodo,
        'contratos_vencen':     contratos_vencen,
        'saldos_apertura_count': saldos_apertura_count,
        'workers_sample':       personal_qs[:20],
        'hoy':                  hoy,
    })


@login_required
@solo_admin
def cuadricula_semanal_local(request, pk):
    """Cuadrícula Semanal del Local — vista turnos de la semana.

    Pantalla estilo brigade kitchen: eje X = 7 días (lun-dom), eje Y =
    trabajadores activos del local, celdas = código de turno asignado
    (M/T/N/Q/D/V/F) con color por tipo.

    Lee del modelo Roster (turnos planificados) y RegistroTareo (turnos
    reales si ya pasó el día). Permite seleccionar la semana a visualizar.
    """
    from datetime import date, timedelta
    from empresas.models import Empresa
    from personal.models import Personal, Roster

    emp = get_object_or_404(Empresa, pk=pk)
    hoy = date.today()

    # Parámetro ?fecha=YYYY-MM-DD para elegir cualquier semana
    fecha_str = request.GET.get('fecha', '')
    if fecha_str:
        try:
            anio, mes, dia = fecha_str.split('-')
            fecha_ref = date(int(anio), int(mes), int(dia))
        except (ValueError, AttributeError):
            fecha_ref = hoy
    else:
        fecha_ref = hoy

    # Calcular lunes de esa semana (weekday 0 = lunes)
    inicio_semana = fecha_ref - timedelta(days=fecha_ref.weekday())
    fin_semana = inicio_semana + timedelta(days=6)
    dias_semana = [inicio_semana + timedelta(days=i) for i in range(7)]
    DIA_NOMBRES = ['LUN', 'MAR', 'MIÉ', 'JUE', 'VIE', 'SÁB', 'DOM']

    # Trabajadores del local
    trabajadores = list(
        Personal.objects.filter(empresa=emp, estado='Activo')
        .order_by('apellidos_nombres')
    )

    # Roster de toda la semana para estos trabajadores
    roster_qs = Roster.objects.filter(
        personal__in=trabajadores,
        fecha__gte=inicio_semana,
        fecha__lte=fin_semana,
    )
    # Indexar por (personal_id, fecha) → código
    roster_map = {(r.personal_id, r.fecha): r.codigo for r in roster_qs}

    # Construir filas de la cuadrícula
    filas = []
    for t in trabajadores:
        celdas = []
        total_dias_trabajados = 0
        total_descansos = 0
        for d in dias_semana:
            codigo = roster_map.get((t.id, d), '')
            # Normalizar código a categoría
            cod_upper = (codigo or '').strip().upper()
            if cod_upper in ('M',):
                tipo = 'M'; color = 'manana'
            elif cod_upper in ('T', 'TR', 'NOR'):
                tipo = 'T'; color = 'tarde'
            elif cod_upper in ('N',):
                tipo = 'N'; color = 'noche'
            elif cod_upper in ('Q',):
                tipo = 'Q'; color = 'quebrado'
            elif cod_upper in ('D', 'DSO'):
                tipo = 'D'; color = 'descanso'; total_descansos += 1
            elif cod_upper in ('V',):
                tipo = 'V'; color = 'vacaciones'
            elif cod_upper in ('FA', 'F'):
                tipo = 'F'; color = 'falta'
            elif cod_upper in ('LSG',):
                tipo = 'LSG'; color = 'licencia'
            elif cod_upper:
                tipo = cod_upper[:3]; color = 'otro'
            else:
                tipo = ''; color = 'vacio'

            if color not in ('vacio', 'descanso', 'falta', 'licencia', 'vacaciones'):
                total_dias_trabajados += 1

            celdas.append({
                'fecha':  d,
                'codigo': codigo,
                'tipo':   tipo,
                'color':  color,
                'es_hoy': (d == hoy),
            })

        # Alertas por trabajador
        alerta_horas = total_dias_trabajados > 6  # más de 6 días seguidos
        alerta_sin_descanso = total_descansos == 0 and total_dias_trabajados > 0

        filas.append({
            'personal':          t,
            'celdas':            celdas,
            'dias_trabajados':   total_dias_trabajados,
            'dias_descanso':     total_descansos,
            'alerta_horas':      alerta_horas,
            'alerta_sin_descanso': alerta_sin_descanso,
        })

    # Stats globales
    total_asignados = sum(1 for f in filas for c in f['celdas'] if c['color'] not in ('vacio',))
    total_celdas = len(filas) * 7
    cobertura_pct = round(total_asignados / total_celdas * 100, 1) if total_celdas else 0
    alertas_count = sum(1 for f in filas if f['alerta_horas'] or f['alerta_sin_descanso'])

    # Navegación entre semanas
    semana_anterior = (inicio_semana - timedelta(days=7)).isoformat()
    semana_siguiente = (inicio_semana + timedelta(days=7)).isoformat()
    semana_actual = inicio_semana <= hoy <= fin_semana

    # ── Export Excel (?format=xlsx) ──────────────────────────────────────
    if request.GET.get('format') == 'xlsx':
        return _exportar_cuadricula_excel(emp, dias_semana, DIA_NOMBRES, filas,
                                          inicio_semana, fin_semana,
                                          total_trabajadores=len(trabajadores),
                                          total_asignados=total_asignados,
                                          cobertura_pct=cobertura_pct,
                                          alertas_count=alertas_count)

    return render(request, 'empresas/cuadricula_semanal.html', {
        'empresa':         emp,
        'inicio_semana':   inicio_semana,
        'fin_semana':      fin_semana,
        'dias_semana':     list(zip(DIA_NOMBRES, dias_semana)),
        'hoy':             hoy,
        'filas':           filas,
        'total_trabajadores': len(trabajadores),
        'total_asignados': total_asignados,
        'cobertura_pct':   cobertura_pct,
        'alertas_count':   alertas_count,
        'semana_anterior': semana_anterior,
        'semana_siguiente': semana_siguiente,
        'semana_actual':   semana_actual,
    })


def _exportar_cuadricula_excel(emp, dias_semana, DIA_NOMBRES, filas,
                                inicio_semana, fin_semana,
                                total_trabajadores, total_asignados,
                                cobertura_pct, alertas_count):
    """Genera xlsx de la cuadrícula semanal lista para imprimir."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    from django.http import HttpResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Cuadricula Semanal'

    TEAL = 'FF0D2B27'
    TEAL_LIGHT = 'FFCFFAFE'
    HDR_FONT = Font(color='FFFFFFFF', bold=True, size=10)
    HDR_FILL = PatternFill(fill_type='solid', fgColor=TEAL)
    DAY_FILL = PatternFill(fill_type='solid', fgColor=TEAL_LIGHT)
    THIN = Side(border_style='thin', color='FFE2E8F0')
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    # Colors por tipo de turno
    TURNO_COLORS = {
        'manana':    'FFFEF3C7',
        'tarde':     'FFDBEAFE',
        'noche':     'FF1E293B',
        'quebrado':  'FFFCE7F3',
        'descanso':  'FFE2E8F0',
        'vacaciones':'FFD1FAE5',
        'falta':     'FFFEE2E2',
        'licencia':  'FFFEF9C3',
        'otro':      'FFF3E8FF',
    }
    TURNO_FONT_LIGHT = Font(color='FFFFFFFF', bold=True, size=10)
    TURNO_FONT_DARK = Font(color='FF1F2937', bold=True, size=10)

    # ── Título ──────────────────────────────────────────────────
    ws['A1'] = f'Cuadricula Semanal — {emp.nombre_comercial or emp.razon_social}'
    ws['A1'].font = Font(bold=True, size=14, color=TEAL)
    ws.merge_cells('A1:J1')
    ws['A2'] = (
        f'Semana: {inicio_semana.strftime("%d/%m/%Y")} a {fin_semana.strftime("%d/%m/%Y")}  '
        f'| {total_trabajadores} trabajadores  | Cobertura: {cobertura_pct}%  '
        f'| Alertas: {alertas_count}'
    )
    ws['A2'].font = Font(size=10, italic=True, color='FF64748B')
    ws.merge_cells('A2:J2')

    # ── Encabezados ─────────────────────────────────────────────
    HEADER_ROW = 4
    ws.cell(row=HEADER_ROW, column=1, value='TRABAJADOR').font = HDR_FONT
    ws.cell(row=HEADER_ROW, column=1).fill = HDR_FILL
    ws.cell(row=HEADER_ROW, column=1).alignment = Alignment(horizontal='center', vertical='center')
    for i, (nombre, fecha) in enumerate(zip(DIA_NOMBRES, dias_semana), start=2):
        c = ws.cell(row=HEADER_ROW, column=i, value=f'{nombre}\n{fecha.strftime("%d/%m")}')
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.cell(row=HEADER_ROW, column=9, value='TRAB').font = HDR_FONT
    ws.cell(row=HEADER_ROW, column=9).fill = HDR_FILL
    ws.cell(row=HEADER_ROW, column=9).alignment = Alignment(horizontal='center')
    ws.cell(row=HEADER_ROW, column=10, value='DESC').font = HDR_FONT
    ws.cell(row=HEADER_ROW, column=10).fill = HDR_FILL
    ws.cell(row=HEADER_ROW, column=10).alignment = Alignment(horizontal='center')

    # ── Filas ───────────────────────────────────────────────────
    for r, fila in enumerate(filas, start=HEADER_ROW + 1):
        nombre = fila['personal'].apellidos_nombres
        alertas = []
        if fila['alerta_horas']: alertas.append('+6D')
        if fila['alerta_sin_descanso']: alertas.append('SD')
        if alertas:
            nombre += f"  [{','.join(alertas)}]"
        ws.cell(row=r, column=1, value=nombre).font = Font(bold=True, size=9)
        ws.cell(row=r, column=1).border = BORDER

        for c, celda in enumerate(fila['celdas'], start=2):
            cell = ws.cell(row=r, column=c, value=celda['tipo'] or '·')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = BORDER
            color = TURNO_COLORS.get(celda['color'])
            if color:
                cell.fill = PatternFill(fill_type='solid', fgColor=color)
                cell.font = TURNO_FONT_LIGHT if celda['color'] == 'noche' else TURNO_FONT_DARK

        ws.cell(row=r, column=9, value=fila['dias_trabajados']).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=9).font = Font(bold=True, size=9)
        ws.cell(row=r, column=10, value=fila['dias_descanso']).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=10).font = Font(bold=True, size=9)

    # ── Leyenda ─────────────────────────────────────────────────
    leyenda_row = HEADER_ROW + len(filas) + 3
    ws.cell(row=leyenda_row, column=1, value='LEYENDA:').font = Font(bold=True, size=9)
    leyenda_items = [
        ('M', 'manana', 'Mañana'),
        ('T', 'tarde', 'Tarde'),
        ('N', 'noche', 'Noche'),
        ('Q', 'quebrado', 'Quebrado'),
        ('D', 'descanso', 'Descanso'),
        ('V', 'vacaciones', 'Vacaciones'),
        ('F', 'falta', 'Falta'),
        ('LSG', 'licencia', 'Lic. s/goce'),
    ]
    for i, (cod, color, label) in enumerate(leyenda_items, start=2):
        cell = ws.cell(row=leyenda_row, column=i, value=f'{cod}')
        cell.fill = PatternFill(fill_type='solid', fgColor=TURNO_COLORS.get(color, 'FFFFFFFF'))
        cell.font = TURNO_FONT_LIGHT if color == 'noche' else TURNO_FONT_DARK
        cell.alignment = Alignment(horizontal='center')
        cell.border = BORDER
    ws.cell(row=leyenda_row + 1, column=1,
            value='+6D = más de 6 días seguidos (viola DS 003-97-TR Art. 1°)  ·  SD = sin descanso semanal').font = Font(size=8, italic=True, color='FF64748B')

    # Anchos de columna
    ws.column_dimensions['A'].width = 38
    for col in 'BCDEFGHI':
        ws.column_dimensions[col].width = 10
    ws.column_dimensions['J'].width = 10
    ws.row_dimensions[HEADER_ROW].height = 30

    buffer = BytesIO()
    wb.save(buffer)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    fname = f'Cuadricula_{emp.subdominio or emp.pk}_{inicio_semana.strftime("%Y%m%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    return response
